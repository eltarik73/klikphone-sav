#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KLIKPHONE SAV - Design inspiré du portail officiel
Couleurs orange, style moderne Tailwind
OPTIMISÉ: Logos inline, caching, fragments
"""

import streamlit as st
import sqlite3
import os
from datetime import datetime
import re
import functools
import unicodedata

# Option Postgres (Supabase)
try:
    import psycopg2
    import psycopg2.extras
except Exception:
    psycopg2 = None

import urllib.parse

# =============================================================================
# FONCTION DE NORMALISATION POUR RECHERCHE (sans accents, minuscules)
# =============================================================================
def normalize_search(text):
    """Normalise le texte pour recherche : supprime accents et met en minuscules"""
    if not text:
        return ""
    # Décompose les caractères accentués puis supprime les diacritiques
    text = unicodedata.normalize('NFD', str(text))
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    return text.lower().strip()

# =============================================================================
# CONFIG
# =============================================================================
st.set_page_config(
    page_title="Klikphone SAV",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="collapsed"
)

DB_PATH = "klikphone_sav.db"

# Logo Klikphone en base64
LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAGQAAABkCAMAAABHPGVmAAAAIGNIUk0AAHomAACAhAAA+gAAAIDoAAB1MAAA6mAAADqYAAAXcJy6UTwAAAIEUExURf9mAP5mAP9nAv1eAPzTuP////76+PzNrv5dAP9lAP5nAf9kAf1iBPzTt/79/P7+/v39/crKyv1oBvzp3f78+/zk1f1lAf1hA/zMq/7////9/MvLy5OTk/9hAPx5I/37+v7+/f349Px1HP9iAP5mAfvOrv/+/pKSkpqamv5nA/9dAPyRS//9/fyLQf9eAP5nAvvNrv78+pubm/5lAP9oBP5bAPusef/8+/umbv9oA5mZmf5cAPzIqPzCnf1kAPzi0fzcyP1hAPxyGP328f3y6v1uEf9jAPyHO//+/f3///yBMv9fAP5oA/9bAPulbf/8+vyfYvuNRP36+PyIPP9oAv5aAPuPSv3+/v38+/uJQf9pBP1ZAPvIqPvCnv5YAP9cAPueZP/6+PuYWfyORvyIO/yCMvyHOvyBMPzOrv1iA/9kAP1iAvvNrfvAmf7+//ubX/yCMfxtE/zt4vyBMf9gAPuqd/uWVv/7+fyPTP1lAvzj0v79/f1bAPvNsP77+PvHpv///v5ZAPuVVfuQTPuhZf5fAPuEOPzz7PuAMP9pBvy4i9XV1ZGRkf1xFPzv5vzr3/xtDvzDnqysrJaWlv14Iv77+f339Px0G/39/Px9KfyUT/uORfyUUP/9+/uORvzi0v5qBv5oBPxtEvubXvvAmP9lAf1hAvvMrPzNrfvLq/1hBPzTtv0V+7gAAAABYktHRAX4b+nHAAAACXBIWXMAAABIAAAASABGyWs+AAAAB3RJTUUH6gEcAQUILspFkwAAAHd0RVh0UmF3IHByb2ZpbGUgdHlwZSA4YmltAAo4YmltCiAgICAgIDQwCjM4NDI0OTRkMDQwNDAwMDAwMDAwMDAwMDM4NDI0OTRkMDQyNTAwMDAwMDAwMDAxMGQ0MWQ4Y2Q5OGYwMGIyMDRlOTgwMDk5OAplY2Y4NDI3ZQqmU8OOAAADkklEQVRo3rXa+VPTQBQH8Ka6QqoS1HosiIJgEdTGUivWeqMVbBAV8MCjKuKBqIh4IN73fd/3ff6TblNmbJ3J5r23If01mU+/+zabvMz6fM6H5h8xko1iBYV+zQc+9MDoMWNZkcHsw2BGMRsnO1/Tx09gQTZxkg5HhDF5Ci/JNUqnyi4IlE2bzspZxYzKKowxkwd5KCdH9SzZFTW1s8V/msPnhk1glIwxj0d4KJRj1MmQ6PzYAl5v1POFsTgMcTAWyUoSTyxmwVCQLQEiToYsieZbuowVid/yFT4I4mjIkKrKlQ2i7uVs1eokoPLOhgwxw2t4ozg51LQ2ZaoYEkSzmtfxFnF6C1/fbKkYMsRMbdgochisotU1idSQTuG29k2bRUm2bI21Rd2NAkdDejNGO7ZtF2vXjp0dAMM5hxzRrfQuvpvvSVtqhhQR60onr+etYVPNkCNmeC8v4V3ysrsbbsg+3ti0X4pkjANyQxmBGKqIMA66GooIKIciYhvM1VBCss9ad0MFARsKCNygIwiDjGAMKgKcu0oIKgcRydznCIOEYA0KgjYICLIeJMTOAVivVBD8WOERwlihEZqBQyj1wCKkeiAR4lihED3QTcuBQHRNP9RDM+CI5i88zCNHKAYYicYTR3mLQTIgSK9AtKrksT7Rq5AMFySVQY73myJI7IR4wacZbv3JSX6Kn661RNUHzrBBoiHvtOJnz7Hz7EJ7R9RMdTWFjIs2gjakPWP80mXRY4earrRbVvNVXp/99GMYWEOCRP3XrjeI7rec3bjZfyt2m0Wyo1XM7tzFGRKkKnnvPhsU86mIPXjoSzxiwaHRYo+f1KEMCRIoe/pMBGGikX/+onuYEDPVy0OGPZ1evkomhme4rPTr7Gc+o4S/6S/MLXypZ4W30m/5uyHkfaosbwojFRjSGfYPfGBF/25GnAJEaj/+t6ygFCjyKfm5z57QFAWK1MQTX/KXeoQCRsRD6yuP0BQwomvfvv/If/yCFTBiv6z8pClwREFBIHQFg2QVhldQCDULDiEqSISmYBGSgkYoCh4hKAQEr1AQtEJCsAoNQSpEBKdQEZRCRjAKHUEoCghcUUHAihJiKz3uihoCVBQRmKKKgBRlBKKoIwDFA8Rd8QJxVTxB3BRvEBfFI0SueIVIFc8QmeIdIlE8RJwVLxFHxVPESfEWGdpmkLf1qrTaa8RWfsE3kdGQjPL7T/52uNK/Zd9DQV9nvogAAABEZVhJZk1NACoAAAAIAAGHaQAEAAAAAQAAABoAAAAAAAOgAQADAAAAAQABAACgAgAEAAAAAQAAB9CgAwAEAAAAAQAAB9AAAAAAxqEN6QAAABF0RVh0ZXhpZjpDb2xvclNwYWNlADEPmwJJAAAAEnRFWHRleGlmOkV4aWZPZmZzZXQAMjZTG6JlAAAAGXRFWHRleGlmOlBpeGVsWERpbWVuc2lvbgAyMDAw1StfagAAABl0RVh0ZXhpZjpQaXhlbFlEaW1lbnNpb24AMjAwMGzQhIIAAAAASUVORK5CYII="

CATEGORIES = ["Smartphone", "Tablette", "PC Portable", "Console", "Commande", "Autre"]

PANNES = ["Écran casse", "Batterie", "Connecteur de charge", 
          "Camera avant", "Camera arriere", 
          "Bouton volume", "Bouton power", 
          "Haut-parleur (je n'entends pas les gens ou la musique)", 
          "Microphone (les gens ne m'entendent pas)", 
          "Vitre arriere", "Désoxydation", "Problème logiciel", "Diagnostic", "Autre"]

STATUTS = ["En attente de diagnostic", "En attente de pièce", "En attente d'accord client",
           "En cours de réparation", "Réparation terminée", "Rendu au client", "Clôturé"]

# Membres équipe par défaut
MEMBRES_EQUIPE_DEFAUT = [
    {"nom": "Marina", "role": "Technicien Apple", "couleur": "#EC4899"},  # Rose
    {"nom": "Jonathan", "role": "Technicien Multimarque", "couleur": "#22C55E"},  # Vert
    {"nom": "Tarik", "role": "Manager", "couleur": "#8B5CF6"},  # Violet
    {"nom": "Oualid", "role": "Accueil", "couleur": "#3B82F6"},  # Bleu
]

MARQUES = {
    "Smartphone": ["Apple", "Samsung", "Xiaomi", "Huawei", "OnePlus", "Google", "Oppo", "Autre"],
    "Tablette": ["Apple", "Samsung", "Huawei", "Lenovo", "Microsoft", "Autre"],
    "PC Portable": ["Apple", "HP", "Dell", "Lenovo", "Asus", "Acer", "MSI", "Autre"],
    "Console": ["Sony", "Microsoft", "Nintendo", "Autre"]
}

MODELES = {
    ("Smartphone", "Apple"): [
        "iPhone 16 Pro Max", "iPhone 16 Pro", "iPhone 16 Plus", "iPhone 16",
        "iPhone 15 Pro Max", "iPhone 15 Pro", "iPhone 15 Plus", "iPhone 15",
        "iPhone 14 Pro Max", "iPhone 14 Pro", "iPhone 14 Plus", "iPhone 14",
        "iPhone 13 Pro Max", "iPhone 13 Pro", "iPhone 13", "iPhone 13 Mini",
        "iPhone 12 Pro Max", "iPhone 12 Pro", "iPhone 12", "iPhone 12 Mini",
        "iPhone 11 Pro Max", "iPhone 11 Pro", "iPhone 11",
        "iPhone XS Max", "iPhone XS", "iPhone XR",
        "iPhone X", "iPhone 8 Plus", "iPhone 8", "iPhone 7 Plus", "iPhone 7",
        "iPhone 6s Plus", "iPhone 6s", "iPhone 6 Plus", "iPhone 6",
        "iPhone SE 2022", "iPhone SE 2020", "iPhone SE",
        "Autre"
    ],
    ("Smartphone", "Samsung"): [
        "Galaxy S24 Ultra", "Galaxy S24+", "Galaxy S24",
        "Galaxy S23 Ultra", "Galaxy S23+", "Galaxy S23",
        "Galaxy S22 Ultra", "Galaxy S22+", "Galaxy S22",
        "Galaxy S21 Ultra", "Galaxy S21+", "Galaxy S21",
        "Galaxy Z Fold 5", "Galaxy Z Fold 4", "Galaxy Z Fold 3",
        "Galaxy Z Flip 5", "Galaxy Z Flip 4", "Galaxy Z Flip 3",
        "Galaxy A54", "Galaxy A53", "Galaxy A34", "Galaxy A14",
        "Autre"
    ],
    ("Smartphone", "Xiaomi"): [
        "Xiaomi 14 Ultra", "Xiaomi 14 Pro", "Xiaomi 14",
        "Xiaomi 13 Ultra", "Xiaomi 13 Pro", "Xiaomi 13",
        "Redmi Note 13 Pro", "Redmi Note 13", "Redmi Note 12",
        "POCO F5", "POCO X5 Pro",
        "Autre"
    ],
    ("Smartphone", "Huawei"): [
        "P60 Pro", "P50 Pro", "P40 Pro", "P30 Pro", "P30",
        "Mate 60 Pro", "Mate 50 Pro", "Mate 40 Pro",
        "Nova 11", "Nova 10",
        "Autre"
    ],
    ("Tablette", "Apple"): [
        "iPad Pro 12.9 M2", "iPad Pro 11 M2",
        "iPad Pro 12.9 M1", "iPad Pro 11 M1",
        "iPad Air M2", "iPad Air M1",
        "iPad 10", "iPad 9", "iPad 8",
        "iPad Mini 6", "iPad Mini 5",
        "Autre"
    ],
    ("Tablette", "Samsung"): [
        "Galaxy Tab S9 Ultra", "Galaxy Tab S9+", "Galaxy Tab S9",
        "Galaxy Tab S8 Ultra", "Galaxy Tab S8+", "Galaxy Tab S8",
        "Galaxy Tab A9", "Galaxy Tab A8",
        "Autre"
    ],
    ("PC Portable", "Apple"): [
        "MacBook Pro 16 M3", "MacBook Pro 14 M3",
        "MacBook Pro 16 M2", "MacBook Pro 14 M2",
        "MacBook Pro 16 M1", "MacBook Pro 14 M1",
        "MacBook Air 15 M3", "MacBook Air 13 M3",
        "MacBook Air 15 M2", "MacBook Air 13 M2",
        "MacBook Air M1",
        "Autre"
    ],
    ("Console", "Sony"): ["PlayStation 5", "PlayStation 5 Digital", "PlayStation 4 Pro", "PlayStation 4", "PS Vita", "Autre"],
    ("Console", "Nintendo"): ["Switch OLED", "Switch V2", "Switch V1", "Switch Lite", "3DS XL", "3DS", "2DS", "Autre"],
    ("Console", "Microsoft"): ["Xbox Series X", "Xbox Series S", "Xbox One X", "Xbox One S", "Xbox One", "Autre"],
}

# =============================================================================
# CSS DESIGN SYSTEM - SAAS PREMIUM (Notion/Stripe/Linear inspired)
# =============================================================================

# Logos des marques en SVG (pour un rendu pro)
# Logos des marques en SVG inline (évite les requêtes réseau)
# Format: data:image/svg+xml pour performance maximale
BRAND_LOGOS = {
    "Apple": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%23000'%3E%3Cpath d='M18.71 19.5c-.83 1.24-1.71 2.45-3.05 2.47-1.34.03-1.77-.79-3.29-.79-1.53 0-2 .77-3.27.82-1.31.05-2.3-1.32-3.14-2.53C4.25 17 2.94 12.45 4.7 9.39c.87-1.52 2.43-2.48 4.12-2.51 1.28-.02 2.5.87 3.29.87.78 0 2.26-1.07 3.81-.91.65.03 2.47.26 3.64 1.98-.09.06-2.17 1.28-2.15 3.81.03 3.02 2.65 4.03 2.68 4.04-.03.07-.42 1.44-1.38 2.83M13 3.5c.73-.83 1.94-1.46 2.94-1.5.13 1.17-.34 2.35-1.04 3.19-.69.85-1.83 1.51-2.95 1.42-.15-1.15.41-2.35 1.05-3.11z'/%3E%3C/svg%3E",
    "Samsung": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%231428a0'%3E%3Cpath d='M2.61 11.4c0 .94.47 1.52 1.63 1.85l1.22.34c.59.17.81.39.81.72 0 .41-.36.66-.96.66-.69 0-1.01-.31-1.04-.78H2.61c.03.98.74 1.61 1.96 1.61 1.14 0 1.96-.63 1.96-1.61 0-.89-.48-1.41-1.56-1.72l-1.17-.33c-.64-.18-.88-.42-.88-.77 0-.39.33-.65.88-.65.58 0 .91.29.95.71h1.56c-.04-.94-.79-1.56-1.91-1.56-1.19 0-1.79.67-1.79 1.53zm5.56 3.25h1.58v-4.98H8.17v4.98zm3.29-4.98l1.68 4.98h1.72l1.68-4.98h-1.62l-.92 3.13-.92-3.13h-1.62zm6.3 0v4.98h1.58v-4.98h-1.58zm4.24 0h-1.58v4.98h1.58v-1.77l.44-.47 1.25 2.24h1.76l-1.93-3.22 1.82-1.76h-1.83l-1.51 1.66v-1.66z'/%3E%3C/svg%3E",
    "Xiaomi": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%23ff6900'%3E%3Cpath d='M19.913 13.317V7.332c0-.636-.509-1.145-1.145-1.145h-6.462c-.636 0-1.145.509-1.145 1.145v11.523c0 .636.509 1.145 1.145 1.145h1.717c.636 0 1.145-.509 1.145-1.145v-5.538h3.6c.636 0 1.145-.509 1.145-1.145v-1.717c0-.636-.509-1.145-1.145-1.145h-3.6v-.573h4.745c.636 0 1.145-.509 1.145-1.145v-.13zm-9.324 5.538c0 .636-.509 1.145-1.145 1.145H7.727c-.636 0-1.145-.509-1.145-1.145V7.332c0-.636.509-1.145 1.145-1.145h1.717c.636 0 1.145.509 1.145 1.145v11.523z'/%3E%3C/svg%3E",
    "Huawei": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%23ff0000'%3E%3Cpath d='M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2zm0 18c-4.41 0-8-3.59-8-8s3.59-8 8-8 8 3.59 8 8-3.59 8-8 8zm-1-13h2v6h-2zm0 8h2v2h-2z'/%3E%3C/svg%3E",
    "OnePlus": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%23f5010c'%3E%3Cpath d='M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2zm5 11h-4v4h-2v-4H7v-2h4V7h2v4h4v2z'/%3E%3C/svg%3E",
    "Google": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%234285f4'%3E%3Cpath d='M12.24 10.285V14.4h6.806c-.275 1.765-2.056 5.174-6.806 5.174-4.095 0-7.439-3.389-7.439-7.574s3.345-7.574 7.439-7.574c2.33 0 3.891.989 4.785 1.849l3.254-3.138C18.189 1.186 15.479 0 12.24 0c-6.635 0-12 5.365-12 12s5.365 12 12 12c6.926 0 11.52-4.869 11.52-11.726 0-.788-.085-1.39-.189-1.989H12.24z'/%3E%3C/svg%3E",
    "Oppo": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%231a8f3e'%3E%3Cpath d='M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2zm0 18c-4.41 0-8-3.59-8-8s3.59-8 8-8 8 3.59 8 8-3.59 8-8 8z'/%3E%3C/svg%3E",
    "Sony": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%23000'%3E%3Cpath d='M2 8.5v7h2.5v-2.5h2v2.5H9v-7H6.5v2.5h-2V8.5H2zm8 0v7h5v-2H12.5v-1h2.5v-2h-2.5v-1H15v-1h-5zm6 0v7h2.5v-5h2v5H23v-7h-7z'/%3E%3C/svg%3E",
    "Microsoft": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%2300a4ef'%3E%3Cpath d='M11.4 11.4H2V2h9.4v9.4zm10.6 0H12.6V2H22v9.4zM11.4 22H2v-9.4h9.4V22zm10.6 0H12.6v-9.4H22V22z'/%3E%3C/svg%3E",
    "Nintendo": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%23e60012'%3E%3Cpath d='M5 3a2 2 0 0 0-2 2v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2V5a2 2 0 0 0-2-2H5zm2 4a3 3 0 1 1 0 6 3 3 0 0 1 0-6zm10 0a3 3 0 1 1 0 6 3 3 0 0 1 0-6z'/%3E%3C/svg%3E",
    "HP": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%230096d6'%3E%3Cpath d='M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2zm-2 15H8v-4H6v4H4V7h2v4h2V7h2v10zm6-4h-2v4h-2V7h4c1.1 0 2 .9 2 2v2c0 1.1-.9 2-2 2zm0-4h-2v2h2V9z'/%3E%3C/svg%3E",
    "Dell": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%23007db8'%3E%3Cpath d='M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2zm-2 15H6V7h4c2.21 0 4 1.79 4 4v2c0 2.21-1.79 4-4 4zm0-2c1.1 0 2-.9 2-2v-2c0-1.1-.9-2-2-2H8v6h2z'/%3E%3C/svg%3E",
    "Lenovo": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%23e2231a'%3E%3Cpath d='M2 8v8h3v-6h2v6h3V8H2zm10 0v8h8v-3h-5v-1h5v-4h-8zm3 2h2v1h-2v-1z'/%3E%3C/svg%3E",
    "Asus": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%23000'%3E%3Cpath d='M23 12l-11 7-11-7 2.5-1.5L12 16l8.5-5.5L23 12zm-11-9l11 7-2.5 1.5L12 6 3.5 11.5 1 10l11-7z'/%3E%3C/svg%3E",
    "Acer": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%2383b81a'%3E%3Cpath d='M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2zm0 18c-4.41 0-8-3.59-8-8s3.59-8 8-8 8 3.59 8 8-3.59 8-8 8zm-4-8l4-4 4 4-4 4-4-4z'/%3E%3C/svg%3E",
    "MSI": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%23ff0000'%3E%3Cpath d='M2 7v10h3V9h2l2 8h2l2-8h2v8h3V7H2z'/%3E%3C/svg%3E",
}

def load_css():
    st.markdown("""
<style>
/* ============================================
   KLIKPHONE SAV - DESIGN SYSTEM v3.0
   Premium • Modern • Glass Morphism
   ============================================ */

/* === FORCE LIGHT MODE === */
html, body, [data-testid="stAppViewContainer"], [data-testid="stApp"] {
    background: #f7f8fb !important;
    color: #1e293b !important;
    min-height: 100vh;
}

/* Force inputs en mode clair */
input, textarea, select, [data-baseweb="input"], [data-baseweb="textarea"], [data-baseweb="select"] {
    background-color: #ffffff !important;
    color: #1e293b !important;
    -webkit-text-fill-color: #1e293b !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 12px !important;
    transition: all 0.2s ease !important;
}

input:focus, textarea:focus, select:focus {
    border-color: #f97316 !important;
    box-shadow: 0 0 0 3px rgba(249, 115, 22, 0.1) !important;
}

input::placeholder, textarea::placeholder {
    color: #94a3b8 !important;
    -webkit-text-fill-color: #94a3b8 !important;
}

/* === TYPOGRAPHY === */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* === DESIGN TOKENS === */
:root {
    --brand-50: #fff7ed;
    --brand-100: #ffedd5;
    --brand-200: #fed7aa;
    --brand-300: #fdba74;
    --brand-400: #fb923c;
    --brand-500: #f97316;
    --brand-600: #ea580c;
    --brand-700: #c2410c;
    
    --neutral-0: #ffffff;
    --neutral-50: #f8fafc;
    --neutral-100: #f1f5f9;
    --neutral-200: #e2e8f0;
    --neutral-300: #cbd5e1;
    --neutral-400: #94a3b8;
    --neutral-500: #64748b;
    --neutral-600: #475569;
    --neutral-700: #334155;
    --neutral-800: #1e293b;
    --neutral-900: #0f172a;
    
    --success: #10b981;
    --warning: #f59e0b;
    --error: #ef4444;
    --info: #3b82f6;
    
    --font: 'Inter', -apple-system, BlinkMacSystemFont, system-ui, sans-serif;
    
    --shadow-sm: 0 1px 2px 0 rgb(0 0 0 / 0.05);
    --shadow-md: 0 4px 6px -1px rgb(0 0 0 / 0.1), 0 2px 4px -2px rgb(0 0 0 / 0.1);
    --shadow-lg: 0 10px 15px -3px rgb(0 0 0 / 0.1), 0 4px 6px -4px rgb(0 0 0 / 0.1);
    --shadow-xl: 0 20px 25px -5px rgb(0 0 0 / 0.1), 0 8px 10px -6px rgb(0 0 0 / 0.1);
    
    --glass-bg: rgba(255, 255, 255, 0.7);
    --glass-border: rgba(255, 255, 255, 0.3);
}

*, *::before, *::after {
    font-family: var(--font) !important;
    box-sizing: border-box;
}

/* === HIDE STREAMLIT ELEMENTS === */
#MainMenu, footer, header, [data-testid="stToolbar"] {visibility: hidden !important;}
.stDeployButton {display: none !important;}

/* === GLASS CARD === */
.glass-card {
    background: var(--glass-bg);
    backdrop-filter: blur(20px);
    -webkit-backdrop-filter: blur(20px);
    border: 1px solid var(--glass-border);
    border-radius: 20px;
    padding: 24px;
    box-shadow: var(--shadow-lg);
}

/* === PREMIUM BUTTONS === */
.stButton > button {
    background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%) !important;
    color: #1e293b !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 14px !important;
    padding: 14px 24px !important;
    font-weight: 600 !important;
    font-size: 15px !important;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    box-shadow: 0 2px 4px rgba(0,0,0,0.05), 0 1px 2px rgba(0,0,0,0.1) !important;
    position: relative;
    overflow: hidden;
}

.stButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 25px rgba(0,0,0,0.1), 0 4px 10px rgba(0,0,0,0.08) !important;
    border-color: #cbd5e1 !important;
}

.stButton > button:active {
    transform: translateY(0) !important;
}

/* Primary Button */
.stButton > button[kind="primary"],
.stButton > button[data-testid="baseButton-primary"] {
    background: linear-gradient(135deg, #f97316 0%, #ea580c 50%, #c2410c 100%) !important;
    color: white !important;
    border: none !important;
    box-shadow: 0 4px 15px rgba(249, 115, 22, 0.4), 0 2px 6px rgba(249, 115, 22, 0.3) !important;
}

.stButton > button[kind="primary"]:hover,
.stButton > button[data-testid="baseButton-primary"]:hover {
    background: linear-gradient(135deg, #ea580c 0%, #c2410c 50%, #9a3412 100%) !important;
    box-shadow: 0 8px 30px rgba(249, 115, 22, 0.5), 0 4px 12px rgba(249, 115, 22, 0.4) !important;
    transform: translateY(-3px) !important;
}

/* Secondary Button */
.stButton > button[kind="secondary"],
.stButton > button[data-testid="baseButton-secondary"] {
    background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%) !important;
    color: #64748b !important;
    border: 1px solid #e2e8f0 !important;
}

/* === BRAND BUTTON (pour les marques) === */
.brand-btn {
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 12px;
    background: white;
    border: 2px solid #e2e8f0;
    border-radius: 16px;
    padding: 20px;
    cursor: pointer;
    transition: all 0.3s ease;
    min-height: 80px;
}

.brand-btn:hover {
    border-color: #f97316;
    background: linear-gradient(135deg, #fff7ed 0%, #ffffff 100%);
    transform: translateY(-4px);
    box-shadow: 0 12px 30px rgba(249, 115, 22, 0.15);
}

.brand-btn img {
    width: 32px;
    height: 32px;
    object-fit: contain;
}

.brand-btn span {
    font-weight: 600;
    font-size: 16px;
    color: #1e293b;
}

/* === CATEGORY CARDS (Smartphone, Tablette, etc) === */
.category-card {
    background: white;
    border: 2px solid #e2e8f0;
    border-radius: 20px;
    padding: 32px 24px;
    text-align: center;
    cursor: pointer;
    transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
    position: relative;
    overflow: hidden;
}

.category-card::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 4px;
    background: linear-gradient(90deg, #f97316, #ea580c);
    transform: scaleX(0);
    transition: transform 0.3s ease;
}

.category-card:hover {
    border-color: #f97316;
    transform: translateY(-6px);
    box-shadow: 0 20px 40px rgba(249, 115, 22, 0.15);
}

.category-card:hover::before {
    transform: scaleX(1);
}

.category-card .icon {
    font-size: 48px;
    margin-bottom: 12px;
}

.category-card .title {
    font-size: 18px;
    font-weight: 700;
    color: #1e293b;
}

/* === HEADER PREMIUM === */
.header-premium {
    text-align: center;
    padding: 40px 20px 30px;
    background: linear-gradient(180deg, rgba(255,255,255,0.9) 0%, rgba(248,250,252,0.8) 100%);
    border-bottom: 1px solid rgba(226, 232, 240, 0.5);
    margin-bottom: 30px;
}

.header-premium .logo-container {
    width: 90px;
    height: 90px;
    margin: 0 auto 16px;
    background: linear-gradient(135deg, #fff7ed 0%, #ffedd5 100%);
    border-radius: 24px;
    display: flex;
    align-items: center;
    justify-content: center;
    box-shadow: 0 10px 30px rgba(249, 115, 22, 0.2);
}

.header-premium h1 {
    font-size: 2.5rem;
    font-weight: 800;
    background: linear-gradient(135deg, #f97316 0%, #c2410c 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    margin: 0 0 8px 0;
    letter-spacing: -1px;
}

.header-premium .subtitle {
    color: #64748b;
    font-size: 1.1rem;
    font-weight: 500;
    margin-bottom: 8px;
}

.header-premium .info {
    color: #94a3b8;
    font-size: 0.9rem;
}

/* === STEP INDICATOR === */
.step-indicator {
    display: flex;
    justify-content: center;
    gap: 12px;
    margin: 24px 0;
}

.step-dot {
    width: 12px;
    height: 12px;
    border-radius: 50%;
    background: #e2e8f0;
    transition: all 0.3s ease;
}

.step-dot.active {
    background: linear-gradient(135deg, #f97316, #ea580c);
    box-shadow: 0 0 0 4px rgba(249, 115, 22, 0.2);
    transform: scale(1.2);
}

.step-dot.completed {
    background: #10b981;
}

/* === PAGE TITLE === */
.page-title {
    font-size: 1.75rem !important;
    font-weight: 700 !important;
    color: #1e293b !important;
    margin-bottom: 8px !important;
}

/* === STAT CARDS === */
.stat-card {
    background: white;
    border-radius: 16px;
    padding: 20px;
    border: 1px solid #e2e8f0;
    transition: all 0.3s ease;
}

.stat-card:hover {
    transform: translateY(-2px);
    box-shadow: var(--shadow-md);
}

.stat-card .value {
    font-size: 2rem;
    font-weight: 700;
    color: #1e293b;
}

.stat-card .label {
    font-size: 0.875rem;
    color: #64748b;
    font-weight: 500;
}

/* === TABLE STYLES === */
.table-header {
    display: flex;
    align-items: center;
    background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
    padding: 14px 20px;
    border-radius: 12px;
    margin-bottom: 8px;
    font-weight: 600;
    font-size: 13px;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

/* === BADGES === */
.badge {
    display: inline-flex;
    align-items: center;
    padding: 6px 12px;
    border-radius: 20px;
    font-size: 12px;
    font-weight: 600;
}

.badge-orange {
    background: linear-gradient(135deg, #fff7ed, #ffedd5);
    color: #c2410c;
    border: 1px solid #fed7aa;
}

.badge-green {
    background: linear-gradient(135deg, #ecfdf5, #d1fae5);
    color: #047857;
    border: 1px solid #a7f3d0;
}

.badge-blue {
    background: linear-gradient(135deg, #eff6ff, #dbeafe);
    color: #1d4ed8;
    border: 1px solid #bfdbfe;
}

.badge-red {
    background: linear-gradient(135deg, #fef2f2, #fee2e2);
    color: #b91c1c;
    border: 1px solid #fecaca;
}

.badge-gray {
    background: linear-gradient(135deg, #f8fafc, #f1f5f9);
    color: #475569;
    border: 1px solid #e2e8f0;
}

/* === TICKET ROW === */
.ticket-row {
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
    padding: 16px 20px;
    margin-bottom: 8px;
    transition: all 0.2s ease;
}

.ticket-row:hover {
    border-color: #f97316;
    box-shadow: 0 4px 12px rgba(249, 115, 22, 0.1);
}

/* === DETAIL CARD === */
.detail-card {
    background: white;
    border-radius: 16px;
    border: 1px solid #e2e8f0;
    overflow: hidden;
    margin-bottom: 16px;
}

.detail-card-header {
    background: linear-gradient(135deg, #1e293b 0%, #334155 100%);
    color: white;
    padding: 16px 20px;
    font-weight: 600;
    font-size: 16px;
}

.detail-card-content {
    padding: 20px;
}

/* === TABS === */
.stTabs [data-baseweb="tab-list"] {
    background: #f8fafc;
    border-radius: 12px;
    padding: 6px;
    gap: 4px;
}

.stTabs [data-baseweb="tab"] {
    border-radius: 8px;
    font-weight: 600;
    padding: 12px 20px;
}

.stTabs [aria-selected="true"] {
    background: white !important;
    box-shadow: var(--shadow-sm);
}

/* === SUCCESS SCREEN === */
.success-screen {
    text-align: center;
    padding: 60px 20px;
}

.success-icon {
    width: 100px;
    height: 100px;
    border-radius: 50%;
    background: linear-gradient(135deg, #10b981 0%, #059669 100%);
    display: flex;
    align-items: center;
    justify-content: center;
    margin: 0 auto 24px;
    box-shadow: 0 20px 40px rgba(16, 185, 129, 0.3);
}

.success-icon span {
    font-size: 50px;
    color: white;
}

/* === ANIMATIONS === */
@keyframes fadeInUp {
    from {
        opacity: 0;
        transform: translateY(20px);
    }
    to {
        opacity: 1;
        transform: translateY(0);
    }
}

@keyframes pulse {
    0%, 100% {
        transform: scale(1);
    }
    50% {
        transform: scale(1.05);
    }
}

.animate-fadeInUp {
    animation: fadeInUp 0.5s ease-out;
}

/* === RESPONSIVE === */
@media (max-width: 768px) {
    .header-premium h1 {
        font-size: 1.75rem;
    }
    
    .category-card {
        padding: 20px 16px;
    }
    
    .category-card .icon {
        font-size: 36px;
    }
}

/* === SCROLLBAR === */
::-webkit-scrollbar {
    width: 8px;
    height: 8px;
}

::-webkit-scrollbar-track {
    background: #f1f5f9;
    border-radius: 4px;
}

::-webkit-scrollbar-thumb {
    background: #cbd5e1;
    border-radius: 4px;
}

::-webkit-scrollbar-thumb:hover {
    background: #94a3b8;
}

/* === EXPANDER === */
.streamlit-expanderHeader {
    background: #f8fafc !important;
    border-radius: 12px !important;
    font-weight: 600 !important;
}

/* === FORM CARD === */
.form-card {
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 16px;
    padding: 24px;
    margin: 16px 0;
}

/* === STEP TITLE === */
.step-title {
    text-align: center;
    margin-bottom: 24px;
}

.step-title h2 {
    font-size: 1.5rem;
    font-weight: 700;
    color: #1e293b;
    margin-bottom: 8px;
}

.step-title p {
    color: #64748b;
    font-size: 1rem;
}

/* === STREAMLIT OVERRIDES === */
.stApp {
    background: linear-gradient(135deg, #fafbfc 0%, #f0f2f5 50%, #e8eaed 100%) !important;
}

#MainMenu, footer, header, .stDeployButton,
[data-testid="stToolbar"], [data-testid="stDecoration"],
.viewerBadge_container__r5tak { 
    display: none !important; 
}

.main .block-container {
    padding: var(--sp-6) var(--sp-6) !important;
    max-width: 1400px !important;
}

/* === TYPOGRAPHY === */
h1, h2, h3 {
    color: var(--neutral-900) !important;
    font-weight: 600 !important;
    letter-spacing: -0.02em !important;
}

.page-header {
    margin-bottom: var(--sp-6);
}

.page-title {
    font-size: var(--text-2xl);
    font-weight: 700;
    color: var(--neutral-900);
    margin: 0;
    letter-spacing: -0.025em;
}

.page-subtitle {
    font-size: var(--text-sm);
    color: var(--neutral-500);
    margin-top: var(--sp-1);
}

.section-header {
    display: flex;
    align-items: center;
    gap: var(--sp-2);
    font-size: var(--text-base);
    font-weight: 600;
    color: var(--neutral-800);
    margin-bottom: var(--sp-4);
    padding-bottom: var(--sp-3);
    border-bottom: 1px solid var(--neutral-200);
}

/* === CARDS - PREMIUM GLASSMORPHISM === */
.card {
    background: rgba(255,255,255,0.9);
    backdrop-filter: blur(20px);
    -webkit-backdrop-filter: blur(20px);
    border: 1px solid rgba(255,255,255,0.8);
    border-radius: 16px;
    padding: var(--sp-5);
    transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1);
    box-shadow: 0 4px 24px rgba(0,0,0,0.06);
}

.card:hover {
    border-color: rgba(249,115,22,0.3);
    box-shadow: 0 8px 32px rgba(0,0,0,0.1);
    transform: translateY(-2px);
}

.card-elevated {
    background: rgba(255,255,255,0.95);
    backdrop-filter: blur(20px);
    border: 1px solid rgba(255,255,255,0.9);
    border-radius: 20px;
    padding: var(--sp-6);
    box-shadow: 0 8px 32px rgba(0,0,0,0.08), 0 2px 8px rgba(0,0,0,0.04);
}

/* KPI Cards - Premium */
.kpi-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: var(--sp-4);
    margin-bottom: var(--sp-6);
}

.kpi-card {
    background: linear-gradient(145deg, rgba(255,255,255,0.95) 0%, rgba(248,250,252,0.9) 100%);
    backdrop-filter: blur(10px);
    border: 1px solid rgba(255,255,255,0.8);
    border-radius: 16px;
    padding: var(--sp-5);
    transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1);
    box-shadow: 0 4px 20px rgba(0,0,0,0.05);
    position: relative;
    overflow: hidden;
}

.kpi-card::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 3px;
    background: linear-gradient(90deg, #f97316, #ea580c);
    opacity: 0;
    transition: opacity 0.25s ease;
}

.kpi-card:hover {
    border-color: rgba(249,115,22,0.2);
    transform: translateY(-4px);
    box-shadow: 0 12px 40px rgba(249,115,22,0.12);
}

.kpi-card:hover::before {
    opacity: 1;
}

.kpi-label {
    font-size: var(--text-xs);
    font-weight: 600;
    color: var(--neutral-500);
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: var(--sp-2);
}

.kpi-value {
    font-size: var(--text-3xl);
    font-weight: 800;
    color: var(--neutral-900);
    line-height: 1;
    letter-spacing: -0.02em;
}

.kpi-value.brand { color: var(--brand-500); }
.kpi-value.success { color: var(--success); }
.kpi-value.warning { color: var(--warning); }
.kpi-value.info { color: var(--info); }

/* === STATUS BADGES - PREMIUM === */
.badge {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    padding: 6px 14px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    white-space: nowrap;
    letter-spacing: 0.02em;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    transition: all 0.2s ease;
}

.badge::before {
    content: '';
    width: 8px;
    height: 8px;
    border-radius: 50%;
    animation: pulse-dot 2s ease-in-out infinite;
}

@keyframes pulse-dot {
    0%, 100% { transform: scale(1); opacity: 1; }
    50% { transform: scale(1.2); opacity: 0.7; }
}

.status-diagnostic {
    background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
    color: #92400e;
    border: 1px solid rgba(245,158,11,0.3);
}
.status-diagnostic::before { background: #f59e0b; box-shadow: 0 0 8px rgba(245,158,11,0.5); }

.status-encours {
    background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%);
    color: #1e40af;
    border: 1px solid rgba(59,130,246,0.3);
}
.status-encours::before { background: #3b82f6; box-shadow: 0 0 8px rgba(59,130,246,0.5); }

.status-termine {
    background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%);
    color: #065f46;
    border: 1px solid rgba(16,185,129,0.3);
}
.status-termine::before { background: #10b981; box-shadow: 0 0 8px rgba(16,185,129,0.5); }

.status-rendu {
    background: linear-gradient(135deg, #10b981 0%, #059669 100%);
    color: white;
    border: 1px solid rgba(16,185,129,0.5);
    box-shadow: 0 4px 12px rgba(16,185,129,0.3);
}
.status-rendu::before { background: rgba(255,255,255,0.8); animation: none; }

.status-cloture {
    background: linear-gradient(135deg, #f1f5f9 0%, #e2e8f0 100%);
    color: #475569;
    border: 1px solid rgba(100,116,139,0.2);
}
.status-cloture::before { background: #94a3b8; animation: none; }

.status-attente-piece {
    background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
    color: #92400e;
    border: 1px solid rgba(245,158,11,0.3);
}
.status-attente-piece::before { background: #f59e0b; }

.status-attente-accord {
    background: linear-gradient(135deg, #fee2e2 0%, #fecaca 100%);
    color: #991b1b;
    border: 1px solid rgba(239,68,68,0.3);
}
.status-attente-accord::before { background: #ef4444; box-shadow: 0 0 8px rgba(239,68,68,0.5); }

/* === TICKET LIST - PREMIUM === */
.ticket-list {
    display: flex;
    flex-direction: column;
    gap: var(--sp-3);
}

.ticket-row {
    display: flex;
    align-items: center;
    gap: var(--sp-4);
    padding: var(--sp-4) var(--sp-5);
    background: rgba(255,255,255,0.9);
    backdrop-filter: blur(10px);
    border: 1px solid rgba(255,255,255,0.8);
    border-radius: 14px;
    transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1);
    cursor: pointer;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04);
}

.ticket-row:hover {
    border-color: rgba(249,115,22,0.4);
    background: linear-gradient(135deg, rgba(255,247,237,0.9) 0%, rgba(255,255,255,0.95) 100%);
    transform: translateX(6px);
    box-shadow: 0 4px 20px rgba(249,115,22,0.12);
}

.ticket-code {
    font-family: 'SF Mono', Monaco, 'Courier New', monospace !important;
    font-size: var(--text-sm);
    font-weight: 600;
    color: var(--brand-600);
    min-width: 100px;
}

.ticket-client {
    font-weight: 500;
    color: var(--neutral-900);
    flex: 1;
}

.ticket-device {
    font-size: var(--text-sm);
    color: var(--neutral-500);
    flex: 1;
}

.ticket-date {
    font-size: var(--text-xs);
    color: var(--neutral-400);
    min-width: 80px;
}

.ticket-alert {
    background: var(--error-light);
    color: var(--error-dark);
    padding: 2px 8px;
    border-radius: var(--r-sm);
    font-size: var(--text-xs);
    font-weight: 600;
}

/* === TICKET TABLE - MODERN DESIGN === */
.ticket-table-container {
    background: var(--neutral-0);
    border: 1px solid var(--neutral-200);
    border-radius: var(--r-xl);
    overflow: hidden;
    box-shadow: 0 4px 24px rgba(0,0,0,0.06);
}

.ticket-table-header {
    display: grid;
    grid-template-columns: 100px 1.4fr 1.2fr 120px 160px 90px 60px;
    gap: 12px;
    padding: 16px 24px;
    background: linear-gradient(180deg, #1e293b 0%, #334155 100%);
    border-bottom: none;
    align-items: center;
}

.ticket-table-header-cell {
    font-size: 11px;
    font-weight: 700;
    color: rgba(255,255,255,0.8);
    text-transform: uppercase;
    letter-spacing: 0.08em;
}

.ticket-table-row {
    display: grid;
    grid-template-columns: 100px 1.4fr 1.2fr 120px 160px 90px 60px;
    gap: 12px;
    padding: 18px 24px;
    border-bottom: 1px solid rgba(0,0,0,0.04);
    align-items: center;
    transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1);
    cursor: pointer;
    background: rgba(255,255,255,0.5);
}

.ticket-table-row:hover {
    background: linear-gradient(90deg, rgba(255,247,237,0.95) 0%, rgba(255,255,255,0.9) 100%);
    transform: scale(1.005);
    box-shadow: 0 4px 20px rgba(249,115,22,0.1);
    border-radius: 0;
}

.ticket-table-row:nth-child(even) {
    background: rgba(248,250,252,0.6);
}

.ticket-table-row:nth-child(even):hover {
    background: linear-gradient(90deg, rgba(255,247,237,0.95) 0%, rgba(255,255,255,0.9) 100%);
}

.ticket-table-row:last-child {
    border-bottom: none;
}

/* Cellules du tableau - Premium */
.tt-ticket {
    display: flex;
    flex-direction: column;
    gap: 4px;
}

.tt-ticket-code {
    font-family: 'SF Mono', Monaco, 'Courier New', monospace !important;
    font-size: 13px;
    font-weight: 700;
    color: #0f172a;
    background: linear-gradient(135deg, #f1f5f9 0%, #e2e8f0 100%);
    padding: 4px 10px;
    border-radius: 6px;
    display: inline-block;
}

.tt-ticket-date {
    font-size: 11px;
    color: #64748b;
    font-weight: 500;
}

.tt-client {
    display: flex;
    align-items: center;
    gap: 12px;
}

.tt-client-avatar {
    width: 40px;
    height: 40px;
    border-radius: 12px;
    background: linear-gradient(135deg, #fff7ed 0%, #fed7aa 100%);
    border: 2px solid rgba(249,115,22,0.2);
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 13px;
    font-weight: 700;
    color: #ea580c;
    flex-shrink: 0;
}

.tt-client-info {
    display: flex;
    flex-direction: column;
    min-width: 0;
}

.tt-client-name {
    font-size: 13px;
    font-weight: 500;
    color: var(--neutral-900);
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}

.tt-client-tel {
    font-size: 11px;
    color: var(--neutral-400);
    font-family: 'SF Mono', Monaco, monospace !important;
}

.tt-appareil {
    display: flex;
    align-items: center;
    gap: 10px;
}

.tt-appareil-icon {
    width: 32px;
    height: 32px;
    border-radius: var(--r-md);
    background: var(--neutral-100);
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 14px;
    flex-shrink: 0;
}

.tt-appareil-info {
    display: flex;
    flex-direction: column;
    min-width: 0;
}

.tt-appareil-marque {
    font-size: 13px;
    font-weight: 500;
    color: var(--neutral-800);
}

.tt-appareil-modele {
    font-size: 11px;
    color: var(--neutral-400);
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}

.tt-tech {
    display: flex;
    align-items: center;
    gap: 8px;
}

.tt-tech-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 4px 10px;
    border-radius: var(--r-full);
    font-size: 11px;
    font-weight: 500;
    color: white;
}

.tt-tech-unassigned {
    display: flex;
    align-items: center;
    gap: 6px;
    color: var(--neutral-400);
    font-size: 12px;
    font-style: italic;
}

.tt-tech-unassigned-icon {
    width: 28px;
    height: 28px;
    border: 2px dashed var(--neutral-300);
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 14px;
    color: var(--neutral-300);
}

.tt-status {
    display: flex;
    align-items: center;
}

.tt-contact {
    display: flex;
    align-items: center;
    gap: 4px;
}

.tt-contact-icon {
    width: 24px;
    height: 24px;
    border-radius: 6px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 12px;
    transition: all 0.15s ease;
}

.tt-contact-icon.active {
    background: var(--success-light);
    color: var(--success);
}

.tt-contact-icon.inactive {
    background: var(--neutral-100);
    color: var(--neutral-300);
}

.tt-action {
    display: flex;
    justify-content: flex-end;
}

.tt-action-btn {
    width: 36px;
    height: 36px;
    border-radius: var(--r-md);
    background: var(--neutral-100);
    border: 1px solid var(--neutral-200);
    display: flex;
    align-items: center;
    justify-content: center;
    color: var(--neutral-500);
    cursor: pointer;
    transition: all 0.15s ease;
}

.tt-action-btn:hover {
    background: var(--brand-500);
    border-color: var(--brand-500);
    color: white;
    transform: scale(1.05);
}

/* Empty state */
.ticket-empty-state {
    padding: 60px 20px;
    text-align: center;
    color: var(--neutral-400);
}

.ticket-empty-state-icon {
    font-size: 48px;
    margin-bottom: 16px;
    opacity: 0.5;
}

.ticket-empty-state-text {
    font-size: 14px;
}

/* Pagination */
.ticket-pagination {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 16px 20px;
    background: var(--neutral-50);
    border-top: 1px solid var(--neutral-200);
}

.ticket-pagination-info {
    font-size: 13px;
    color: var(--neutral-500);
}

.ticket-pagination-dots {
    display: flex;
    gap: 6px;
}

.ticket-pagination-dot {
    width: 8px;
    height: 8px;
    border-radius: 50%;
    background: var(--neutral-300);
    transition: all 0.15s ease;
}

.ticket-pagination-dot.active {
    background: var(--brand-500);
    transform: scale(1.2);
}

/* === TABLE HEADER === */
.table-header {
    display: flex;
    align-items: center;
    gap: var(--sp-4);
    padding: var(--sp-3) var(--sp-4);
    background: var(--neutral-100);
    border-radius: var(--r-md);
    margin-bottom: var(--sp-2);
    font-size: var(--text-xs);
    font-weight: 600;
    color: var(--neutral-500);
    text-transform: uppercase;
    letter-spacing: 0.05em;
}

/* === BUTTONS - PREMIUM DESIGN === */
.stButton > button {
    font-family: var(--font) !important;
    font-weight: 600 !important;
    font-size: 0.9rem !important;
    padding: 12px 20px !important;
    border-radius: 12px !important;
    transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1) !important;
    border: none !important;
    letter-spacing: 0.01em !important;
    white-space: pre-wrap !important;
    line-height: 1.4 !important;
}

.stButton > button[kind="primary"],
.stButton > button[data-testid="baseButton-primary"] {
    background: linear-gradient(135deg, #f97316 0%, #ea580c 50%, #dc2626 100%) !important;
    color: white !important;
    font-weight: 700 !important;
    box-shadow: 0 4px 15px rgba(249,115,22,0.3), 0 2px 4px rgba(0,0,0,0.1) !important;
    text-shadow: 0 1px 2px rgba(0,0,0,0.1) !important;
}

.stButton > button[kind="primary"]:hover,
.stButton > button[data-testid="baseButton-primary"]:hover {
    background: linear-gradient(135deg, #ea580c 0%, #dc2626 50%, #b91c1c 100%) !important;
    transform: translateY(-3px) scale(1.02) !important;
    box-shadow: 0 8px 25px rgba(249,115,22,0.4), 0 4px 8px rgba(0,0,0,0.15) !important;
}

.stButton > button[kind="secondary"],
.stButton > button[data-testid="baseButton-secondary"] {
    background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%) !important;
    color: #374151 !important;
    border: 1px solid #e2e8f0 !important;
    font-weight: 600 !important;
    box-shadow: 0 2px 4px rgba(0,0,0,0.04) !important;
}

.stButton > button[kind="secondary"]:hover,
.stButton > button[data-testid="baseButton-secondary"]:hover {
    background: linear-gradient(180deg, #fff7ed 0%, #ffedd5 100%) !important;
    border-color: #f97316 !important;
    color: #ea580c !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 4px 12px rgba(249,115,22,0.15) !important;
}

/* Boutons désactivés */
.stButton > button:disabled {
    background: #f1f5f9 !important;
    color: #94a3b8 !important;
    cursor: not-allowed !important;
    box-shadow: none !important;
}

/* Client Interface - Larger Buttons */
.main .block-container .stButton > button {
    padding: 14px 24px !important;
    font-size: 0.95rem !important;
    border-radius: 14px !important;
    min-height: 56px !important;
}

/* === FORM INPUTS - PREMIUM === */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea {
    border: 2px solid #e2e8f0 !important;
    border-radius: 12px !important;
    padding: 12px 16px !important;
    font-size: 0.9rem !important;
    background: linear-gradient(180deg, #ffffff 0%, #fafbfc 100%) !important;
    transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
    box-shadow: 0 2px 4px rgba(0,0,0,0.02) !important;
}

.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: #f97316 !important;
    box-shadow: 0 0 0 4px rgba(249,115,22,0.1), 0 4px 12px rgba(249,115,22,0.08) !important;
    outline: none !important;
    background: #ffffff !important;
}

.stTextInput > div > div > input:hover,
.stTextArea > div > div > textarea:hover {
    border-color: #cbd5e1 !important;
}

.stTextInput > div > div > input::placeholder {
    color: #94a3b8 !important;
    font-style: italic !important;
}

.stSelectbox > div > div {
    border: 2px solid #e2e8f0 !important;
    border-radius: 12px !important;
    background: linear-gradient(180deg, #ffffff 0%, #fafbfc 100%) !important;
    transition: all 0.25s ease !important;
}

.stSelectbox > div > div > div {
    padding: 10px 14px !important;
    font-size: 0.9rem !important;
    min-height: 44px !important;
}

.stSelectbox > div > div:hover {
    border-color: #cbd5e1 !important;
}

.stSelectbox > div > div:focus-within {
    border-color: #f97316 !important;
    box-shadow: 0 0 0 4px rgba(249,115,22,0.1) !important;
}

/* Number Input Premium */
.stNumberInput > div > div > input {
    border: 2px solid #e2e8f0 !important;
    border-radius: 12px !important;
    padding: 10px 14px !important;
    font-size: 0.9rem !important;
    font-weight: 600 !important;
    background: linear-gradient(180deg, #ffffff 0%, #fafbfc 100%) !important;
}

.stNumberInput > div > div > input:focus {
    border-color: #f97316 !important;
    box-shadow: 0 0 0 4px rgba(249,115,22,0.1) !important;
}

/* Labels Premium */
.stTextInput > label, .stTextArea > label,
.stSelectbox > label, .stNumberInput > label {
    font-size: 0.85rem !important;
    font-weight: 600 !important;
    color: #374151 !important;
    letter-spacing: 0.01em !important;
    margin-bottom: 6px !important;
}

/* Date Input Premium */
.stDateInput > div > div > input {
    border: 2px solid #e2e8f0 !important;
    border-radius: 12px !important;
    padding: 10px 14px !important;
    font-size: 0.9rem !important;
    background: linear-gradient(180deg, #ffffff 0%, #fafbfc 100%) !important;
}

.stDateInput > div > div > input:focus {
    border-color: #f97316 !important;
    box-shadow: 0 0 0 4px rgba(249,115,22,0.1) !important;
}

/* === TABS === */
.stTabs [data-baseweb="tab-list"] {
    background: var(--neutral-100) !important;
    padding: 4px !important;
    border-radius: var(--r-lg) !important;
    gap: 4px !important;
}

.stTabs [data-baseweb="tab"] {
    padding: 8px 16px !important;
    font-size: var(--text-sm) !important;
    font-weight: 500 !important;
    color: var(--neutral-600) !important;
    background: transparent !important;
    border-radius: var(--r-md) !important;
    border: none !important;
}

.stTabs [data-baseweb="tab"]:hover {
    color: var(--neutral-900) !important;
    background: var(--neutral-50) !important;
}

.stTabs [aria-selected="true"] {
    background: linear-gradient(180deg, #ffffff 0%, #fafbfc 100%) !important;
    color: #0f172a !important;
    box-shadow: 0 4px 12px rgba(0,0,0,0.08) !important;
    font-weight: 600 !important;
}

.stTabs [data-baseweb="tab-highlight"],
.stTabs [data-baseweb="tab-border"] {
    display: none !important;
}

/* === EXPANDER - PREMIUM === */
.streamlit-expanderHeader {
    font-size: 0.95rem !important;
    font-weight: 600 !important;
    color: #374151 !important;
    background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%) !important;
    border-radius: 12px !important;
    padding: 14px 18px !important;
    border: 1px solid #e2e8f0 !important;
    transition: all 0.2s ease !important;
}

.streamlit-expanderHeader:hover {
    background: linear-gradient(180deg, #fff7ed 0%, #ffedd5 100%) !important;
    border-color: rgba(249,115,22,0.3) !important;
}

/* Cacher seulement l'icône SVG, pas le texte */
.streamlit-expanderHeader svg {
    display: none !important;
}

/* Style pour le contenu */
.streamlit-expanderContent {
    border: 1px solid #e2e8f0 !important;
    border-top: none !important;
    border-radius: 0 0 12px 12px !important;
    padding: 18px !important;
    background: linear-gradient(180deg, #ffffff 0%, #fafbfc 100%) !important;
}

/* === PROGRESS - PREMIUM === */
.stProgress > div > div {
    background: linear-gradient(90deg, #e2e8f0 0%, #f1f5f9 100%) !important;
    border-radius: 10px !important;
    height: 10px !important;
    box-shadow: inset 0 2px 4px rgba(0,0,0,0.06) !important;
}

.stProgress > div > div > div {
    background: linear-gradient(90deg, #fb923c 0%, #f97316 50%, #ea580c 100%) !important;
    border-radius: 10px !important;
    box-shadow: 0 2px 8px rgba(249,115,22,0.3) !important;
}

/* === METRICS - PREMIUM === */
[data-testid="stMetricValue"] {
    font-size: 2rem !important;
    font-weight: 800 !important;
    color: #0f172a !important;
    letter-spacing: -0.02em !important;
}

[data-testid="stMetricLabel"] {
    font-size: 0.85rem !important;
    font-weight: 600 !important;
    color: #64748b !important;
    text-transform: uppercase !important;
    letter-spacing: 0.05em !important;
}

/* === ALERTS - PREMIUM === */
.stAlert {
    border-radius: 14px !important;
    border: none !important;
    padding: 16px 20px !important;
    box-shadow: 0 4px 12px rgba(0,0,0,0.06) !important;
}

.stSuccess { 
    background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%) !important; 
    border-left: 4px solid #10b981 !important;
}
.stWarning { 
    background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%) !important; 
    border-left: 4px solid #f59e0b !important;
}
.stError { 
    background: linear-gradient(135deg, #fee2e2 0%, #fecaca 100%) !important; 
    border-left: 4px solid #ef4444 !important;
}
.stInfo { 
    background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%) !important; 
    border-left: 4px solid #3b82f6 !important;
}

/* === DIVIDERS === */
hr {
    border: none !important;
    height: 1px !important;
    background: var(--neutral-200) !important;
    margin: var(--sp-5) 0 !important;
}

/* === FILTER BAR === */
.filter-bar {
    background: var(--neutral-0);
    border: 1px solid var(--neutral-200);
    border-radius: var(--r-lg);
    padding: var(--sp-4);
    margin-bottom: var(--sp-4);
}

/* === NAV HEADER === */
.nav-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: var(--sp-4) var(--sp-5);
    background: var(--neutral-0);
    border-bottom: 1px solid var(--neutral-200);
    margin: calc(-1 * var(--sp-6)) calc(-1 * var(--sp-6)) var(--sp-5);
}

.nav-logo {
    display: flex;
    align-items: center;
    gap: var(--sp-3);
}

.nav-logo-text {
    font-size: var(--text-xl);
    font-weight: 700;
    color: var(--brand-600);
    letter-spacing: -0.02em;
}

.nav-actions {
    display: flex;
    align-items: center;
    gap: var(--sp-2);
}

/* === DETAIL SECTIONS === */
.detail-grid {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: var(--sp-4);
}

.detail-card {
    background: var(--neutral-0);
    border: 1px solid var(--neutral-200);
    border-radius: var(--r-lg);
    padding: var(--sp-5);
}

.detail-card-header {
    font-size: 0.95rem;
    font-weight: 700;
    color: #1e293b;
    margin-bottom: var(--sp-4);
    padding: 12px 16px;
    background: linear-gradient(135deg, rgba(249,115,22,0.08) 0%, rgba(234,88,12,0.04) 100%);
    border-radius: 12px;
    border-left: 4px solid #f97316;
    display: flex;
    align-items: center;
    gap: 10px;
    letter-spacing: -0.01em;
}

.detail-row {
    display: flex;
    justify-content: space-between;
    padding: var(--sp-3) 0;
    border-bottom: 1px solid rgba(0,0,0,0.05);
}

.detail-row:last-child { border-bottom: none; }

.detail-label {
    font-size: var(--text-sm);
    color: #64748b;
    font-weight: 500;
}

.detail-value {
    font-size: var(--text-sm);
    font-weight: 600;
    color: #0f172a;
}

/* === PRICE BOX - PREMIUM === */
.price-box {
    background: linear-gradient(145deg, #fff7ed 0%, #ffedd5 50%, #fed7aa 100%);
    border: 2px solid rgba(249,115,22,0.3);
    border-radius: 16px;
    padding: var(--sp-6);
    box-shadow: 0 4px 20px rgba(249,115,22,0.1);
    position: relative;
    overflow: hidden;
}

.price-box::before {
    content: '€';
    position: absolute;
    top: -20px;
    right: -10px;
    font-size: 120px;
    font-weight: 900;
    color: rgba(249,115,22,0.08);
    pointer-events: none;
}

.price-total {
    font-size: var(--text-3xl);
    font-weight: 800;
    color: #ea580c;
    letter-spacing: -0.02em;
}

.price-label {
    font-size: var(--text-sm);
    color: var(--neutral-500);
}

/* === EMPTY STATE === */
.empty-state {
    text-align: center;
    padding: var(--sp-12) var(--sp-6);
}

.empty-icon {
    font-size: 48px;
    margin-bottom: var(--sp-4);
    opacity: 0.4;
}

.empty-title {
    font-size: var(--text-lg);
    font-weight: 600;
    color: var(--neutral-700);
    margin-bottom: var(--sp-2);
}

.empty-text {
    font-size: var(--text-sm);
    color: var(--neutral-500);
}

/* === RESPONSIVE === */
@media (max-width: 768px) {
    .main .block-container {
        padding: var(--sp-4) !important;
    }
    
    .detail-grid {
        grid-template-columns: 1fr;
    }
    
    .kpi-grid {
        grid-template-columns: repeat(2, 1fr);
    }
    
    .ticket-row {
        flex-wrap: wrap;
    }
    
    .nav-header {
        flex-direction: column;
        gap: var(--sp-3);
    }
}

/* === ANIMATIONS === */
@keyframes fadeIn {
    from { opacity: 0; transform: translateY(8px); }
    to { opacity: 1; transform: translateY(0); }
}

.fade-in {
    animation: fadeIn 0.2s ease-out;
}

/* === PRINT === */
@media print {
    .stButton, .nav-header, .stTabs [data-baseweb="tab-list"] {
        display: none !important;
    }
}

/* ==============================================
   CLIENT INTERFACE - PREMIUM DESIGN
   Inspiration: Apple Store, Linear, Stripe
   ============================================== */

/* Client Page Background */
.client-page {
    min-height: 100vh;
    background: linear-gradient(180deg, #ffffff 0%, #f8fafc 50%, #f1f5f9 100%);
}

/* Logo Header */
.client-hero {
    text-align: center;
    padding: 2.5rem 1rem 2rem;
    background: linear-gradient(180deg, #ffffff 0%, rgba(249,115,22,0.03) 100%);
}

.client-logo-container {
    width: 90px;
    height: 90px;
    margin: 0 auto 1rem;
    background: linear-gradient(135deg, #fff7ed 0%, #ffedd5 100%);
    border-radius: 24px;
    display: flex;
    align-items: center;
    justify-content: center;
    box-shadow: 0 8px 32px rgba(249,115,22,0.15), 0 2px 8px rgba(0,0,0,0.05);
    border: 1px solid rgba(249,115,22,0.1);
}

.client-logo-container img {
    width: 60px;
    height: 60px;
}

.client-brand {
    font-size: 2.2rem;
    font-weight: 800;
    background: linear-gradient(135deg, #f97316 0%, #ea580c 50%, #c2410c 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    letter-spacing: -1px;
    margin-bottom: 0.5rem;
}

.client-tagline {
    font-size: 1rem;
    color: #64748b;
    max-width: 400px;
    margin: 0 auto;
    line-height: 1.5;
}

.client-contact {
    margin-top: 1rem;
    display: flex;
    justify-content: center;
    gap: 1.5rem;
    font-size: 0.85rem;
    color: #94a3b8;
}

.client-contact-item {
    display: flex;
    align-items: center;
    gap: 0.4rem;
}

/* Progress Steps */
.progress-container {
    max-width: 500px;
    margin: 0 auto 2rem;
    padding: 0 1rem;
}

.progress-bar-wrapper {
    position: relative;
    height: 6px;
    background: #e2e8f0;
    border-radius: 100px;
    overflow: hidden;
    margin-bottom: 1rem;
}

.progress-bar-fill {
    height: 100%;
    background: linear-gradient(90deg, #f97316 0%, #fb923c 100%);
    border-radius: 100px;
    transition: width 0.5s cubic-bezier(0.4, 0, 0.2, 1);
    box-shadow: 0 0 12px rgba(249,115,22,0.4);
}

.progress-steps {
    display: flex;
    justify-content: space-between;
    position: relative;
}

.progress-step {
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 0.5rem;
}

.progress-step-dot {
    width: 32px;
    height: 32px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 0.8rem;
    font-weight: 600;
    transition: all 0.3s ease;
}

.progress-step-dot.completed {
    background: linear-gradient(135deg, #f97316 0%, #ea580c 100%);
    color: white;
    box-shadow: 0 4px 12px rgba(249,115,22,0.3);
}

.progress-step-dot.current {
    background: white;
    color: #f97316;
    border: 2px solid #f97316;
    box-shadow: 0 0 0 4px rgba(249,115,22,0.15);
}

.progress-step-dot.pending {
    background: #f1f5f9;
    color: #94a3b8;
    border: 2px solid #e2e8f0;
}

.progress-step-label {
    font-size: 0.7rem;
    color: #94a3b8;
    text-align: center;
    max-width: 60px;
}

.progress-step-label.active {
    color: #f97316;
    font-weight: 600;
}

/* Step Title */
.step-title {
    text-align: center;
    margin-bottom: 2rem;
}

.step-title h2 {
    font-size: 1.5rem;
    font-weight: 700;
    color: #1e293b;
    margin-bottom: 0.5rem;
}

.step-title p {
    font-size: 0.95rem;
    color: #64748b;
}

/* Device Cards */
.device-grid {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: 1rem;
    max-width: 500px;
    margin: 0 auto;
    padding: 0 1rem;
}

.device-card {
    background: white;
    border: 2px solid #e2e8f0;
    border-radius: 20px;
    padding: 1.5rem;
    text-align: center;
    cursor: pointer;
    transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1);
    position: relative;
    overflow: hidden;
}

.device-card::before {
    content: '';
    position: absolute;
    inset: 0;
    background: linear-gradient(135deg, rgba(249,115,22,0.05) 0%, transparent 100%);
    opacity: 0;
    transition: opacity 0.3s ease;
}

.device-card:hover {
    border-color: #fb923c;
    transform: translateY(-4px);
    box-shadow: 0 12px 40px rgba(249,115,22,0.15), 0 4px 12px rgba(0,0,0,0.05);
}

.device-card:hover::before {
    opacity: 1;
}

.device-card:hover .device-icon {
    transform: scale(1.1);
}

.device-icon {
    font-size: 2.5rem;
    margin-bottom: 0.75rem;
    transition: transform 0.3s ease;
    position: relative;
    z-index: 1;
}

.device-name {
    font-size: 1rem;
    font-weight: 600;
    color: #1e293b;
    position: relative;
    z-index: 1;
}

.device-card-special {
    background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
    border-style: dashed;
}

/* Brand Cards */
.brand-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 0.75rem;
    max-width: 500px;
    margin: 0 auto;
    padding: 0 1rem;
}

.brand-card {
    background: white;
    border: 2px solid #e2e8f0;
    border-radius: 16px;
    padding: 1.25rem 0.75rem;
    text-align: center;
    cursor: pointer;
    transition: all 0.2s ease;
}

.brand-card:hover {
    border-color: #f97316;
    background: linear-gradient(135deg, #fff7ed 0%, white 100%);
    transform: translateY(-2px);
    box-shadow: 0 8px 24px rgba(249,115,22,0.12);
}

.brand-logo {
    font-size: 1.8rem;
    margin-bottom: 0.5rem;
}

.brand-name {
    font-size: 0.85rem;
    font-weight: 600;
    color: #374151;
}

/* Model Grid */
.model-grid {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: 0.5rem;
    max-width: 500px;
    margin: 0 auto;
    padding: 0 1rem;
    max-height: 400px;
    overflow-y: auto;
}

.model-card {
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
    padding: 1rem;
    text-align: center;
    cursor: pointer;
    transition: all 0.15s ease;
    font-size: 0.9rem;
    color: #374151;
}

.model-card:hover {
    border-color: #f97316;
    background: #fff7ed;
}

/* Problem Cards */
.problem-grid {
    display: flex;
    flex-direction: column;
    gap: 0.5rem;
    max-width: 500px;
    margin: 0 auto;
    padding: 0 1rem;
}

.problem-card {
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 14px;
    padding: 1rem 1.25rem;
    display: flex;
    align-items: center;
    gap: 1rem;
    cursor: pointer;
    transition: all 0.2s ease;
}

.problem-card:hover {
    border-color: #f97316;
    background: linear-gradient(90deg, #fff7ed 0%, white 100%);
    transform: translateX(4px);
    box-shadow: 0 4px 16px rgba(249,115,22,0.1);
}

.problem-icon {
    width: 44px;
    height: 44px;
    border-radius: 12px;
    background: linear-gradient(135deg, #fff7ed 0%, #ffedd5 100%);
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 1.3rem;
    flex-shrink: 0;
}

.problem-text {
    font-size: 0.95rem;
    color: #374151;
    font-weight: 500;
    text-align: left;
}

/* Success Screen */
.success-screen {
    text-align: center;
    padding: 3rem 1.5rem;
    max-width: 450px;
    margin: 0 auto;
}

.success-icon {
    width: 100px;
    height: 100px;
    margin: 0 auto 1.5rem;
    background: linear-gradient(135deg, #22c55e 0%, #16a34a 100%);
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    box-shadow: 0 12px 40px rgba(34,197,94,0.3);
    animation: successPop 0.5s cubic-bezier(0.175, 0.885, 0.32, 1.275);
}

@keyframes successPop {
    0% { transform: scale(0); opacity: 0; }
    50% { transform: scale(1.1); }
    100% { transform: scale(1); opacity: 1; }
}

.success-icon svg {
    width: 50px;
    height: 50px;
    color: white;
}

.success-title {
    font-size: 1.8rem;
    font-weight: 700;
    color: #1e293b;
    margin-bottom: 0.5rem;
}

.success-subtitle {
    font-size: 1rem;
    color: #64748b;
    margin-bottom: 1.5rem;
}

.success-ticket-code {
    display: inline-block;
    background: linear-gradient(135deg, #1e293b 0%, #334155 100%);
    color: white;
    font-family: 'SF Mono', Monaco, monospace;
    font-size: 1.5rem;
    font-weight: 700;
    padding: 1rem 2rem;
    border-radius: 16px;
    letter-spacing: 2px;
    margin-bottom: 1.5rem;
    box-shadow: 0 8px 24px rgba(30,41,59,0.2);
}

.success-qr {
    margin: 1.5rem 0;
    padding: 1rem;
    background: white;
    border-radius: 16px;
    display: inline-block;
    box-shadow: 0 4px 16px rgba(0,0,0,0.08);
}

/* Back Button */
.back-button {
    display: inline-flex;
    align-items: center;
    gap: 0.5rem;
    color: #64748b;
    font-size: 0.9rem;
    font-weight: 500;
    padding: 0.5rem 1rem;
    border-radius: 10px;
    transition: all 0.2s ease;
    margin-bottom: 1rem;
}

.back-button:hover {
    background: #f1f5f9;
    color: #1e293b;
}

/* Security Section */
.security-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 1rem;
    max-width: 450px;
    margin: 0 auto 1.5rem;
}

.security-option {
    background: white;
    border: 2px solid #e2e8f0;
    border-radius: 16px;
    padding: 1.5rem 1rem;
    text-align: center;
    cursor: pointer;
    transition: all 0.2s ease;
}

.security-option:hover, .security-option.selected {
    border-color: #f97316;
    background: #fff7ed;
}

.security-option.selected {
    box-shadow: 0 0 0 3px rgba(249,115,22,0.2);
}

.security-icon {
    font-size: 2rem;
    margin-bottom: 0.5rem;
}

.security-label {
    font-size: 0.85rem;
    font-weight: 600;
    color: #374151;
}

/* Form Card */
.form-card {
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 20px;
    padding: 1.5rem;
    max-width: 500px;
    margin: 0 auto;
    box-shadow: 0 4px 24px rgba(0,0,0,0.05);
}

/* CGV */
.cgv-box {
    background: #f8fafc;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
    padding: 1rem;
    max-height: 200px;
    overflow-y: auto;
    font-size: 0.8rem;
    color: #64748b;
    line-height: 1.6;
}

/* === PREMIUM FIXES v4 === */
.stButton > button { white-space: nowrap !important; }
.stButton > button span { white-space: nowrap !important; }

/* Hide Streamlit chrome for "custom app" look */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
[data-testid="stToolbar"] {visibility: hidden !important;}
[data-testid="stStatusWidget"] {visibility: hidden !important;}
.stApp > footer {display: none !important;}
a[href*="streamlit.io"] {display:none !important;}

/* Home tiles (Accueil / Technicien / Client) */
.home-grid{ max-width: 1040px; margin: 0 auto; }
.home-tile .stButton > button{
    height: 96px !important;
    border-radius: 18px !important;
    border: 1px solid rgba(226,232,240,1) !important;
    background: rgba(255,255,255,0.92) !important;
    box-shadow: 0 18px 40px rgba(16,24,40,0.08) !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important;
    padding: 0 22px !important;
    gap: 16px !important;
    text-align: left !important;
    transition: transform .12s ease, box-shadow .12s ease, border-color .12s ease !important;
}
.home-tile .stButton > button:hover{
    transform: translateY(-1px) !important;
    box-shadow: 0 22px 55px rgba(16,24,40,0.10) !important;
    border-color: rgba(249,115,22,0.35) !important;
}

/* Small header buttons: keep readable */
[data-testid="column"] .stButton > button{
    min-height: 44px !important;
}

/* === BOUTONS UNIFORMES (Client step1, step2) === */
.uniform-buttons button,
.brand-grid button {
    height: 50px !important;
    min-height: 50px !important;
    font-size: 15px !important;
    font-weight: 500 !important;
    border-radius: 10px !important;
}

/* === GRILLE MARQUES === */
.brand-grid [data-testid="stVerticalBlock"] {
    gap: 8px !important;
}

/* === HOME BUTTONS === */
.home-buttons button {
    height: 50px !important;
    font-size: 15px !important;
    font-weight: 500 !important;
    border-radius: 10px !important;
}

</style>
""", unsafe_allow_html=True)
# =============================================================================
# DATABASE
# =============================================================================

# =============================================================================
# DB (SQLite local ou Postgres Supabase)
# =============================================================================

def is_postgres():
    """Retourne True si une config Postgres/Supabase est présente."""
    # 1) URL complète via env
    if os.getenv("SUPABASE_DB_URL"):
        return True
    # 2) URL complète via secrets
    try:
        if "SUPABASE_DB_URL" in st.secrets:
            return True
    except Exception:
        pass
    # 3) Format Streamlit connections (recommandé)
    try:
        return "connections" in st.secrets and "postgresql" in st.secrets["connections"]
    except Exception:
        return False


@st.cache_resource
def _get_pg_connection():
    """Connexion Postgres persistante (cache_resource)"""
    if psycopg2 is None:
        raise RuntimeError("psycopg2 n'est pas installé. Ajoute 'psycopg2-binary' dans requirements.txt")

    # URL complète
    url = os.getenv("SUPABASE_DB_URL")
    if not url:
        try:
            url = st.secrets.get("SUPABASE_DB_URL")
        except Exception:
            url = None

    if url:
        # Force SSL si absent (Supabase)
        if "sslmode=" not in url:
            sep = "&" if "?" in url else "?"
            url = f"{url}{sep}sslmode=require"
        return psycopg2.connect(url, connect_timeout=10)

    # Format Streamlit: [connections.postgresql]
    cfg = st.secrets["connections"]["postgresql"]
    host = cfg.get("host")
    port = int(cfg.get("port", 5432))
    database = cfg.get("database", "postgres")
    user = cfg.get("username") or cfg.get("user")
    password = cfg.get("password")
    sslmode = cfg.get("sslmode", "require")
    return psycopg2.connect(
        host=host, port=port, dbname=database, user=user, password=password,
        sslmode=sslmode, connect_timeout=10
    )


def _pg_connect():
    """Récupère la connexion Postgres persistante."""
    conn = _get_pg_connection()
    # Vérifier si la connexion est encore valide
    try:
        conn.cursor().execute("SELECT 1")
    except:
        # Connexion fermée, on clear le cache et on réessaie
        _get_pg_connection.clear()
        conn = _get_pg_connection()
    return conn


class _PgCursorWrapper:
    """Wrap un curseur psycopg2 pour:
    - accepter les placeholders SQLite (?) en les convertissant en %s
    - fournir lastrowid via RETURNING id sur INSERT clients/tickets
    """
    def __init__(self, cur):
        self._cur = cur
        self.lastrowid = None

    def execute(self, sql, params=None):
        if params is None:
            params = ()
        sql2 = sql.replace("?", "%s")
        
        # Convertir syntaxe SQLite -> PostgreSQL
        s = sql2.lstrip().lower()
        
        # INSERT OR REPLACE INTO params -> INSERT ... ON CONFLICT (cle) DO UPDATE
        if "insert or replace into params" in s:
            sql2 = "INSERT INTO params (cle, valeur) VALUES (%s, %s) ON CONFLICT (cle) DO UPDATE SET valeur = EXCLUDED.valeur"
        
        # INSERT OR IGNORE -> INSERT ... ON CONFLICT DO NOTHING
        if "insert or ignore" in s:
            sql2 = sql2.replace("INSERT OR IGNORE", "INSERT")
            sql2 = sql2.replace("insert or ignore", "INSERT")
            sql2 = sql2.rstrip().rstrip(";") + " ON CONFLICT DO NOTHING"

        s = sql2.lstrip().lower()
        if s.startswith("insert into clients") and "returning" not in s and "on conflict" not in s:
            sql2 = sql2.rstrip().rstrip(";") + " RETURNING id"
            self._cur.execute(sql2, params)
            row = self._cur.fetchone()
            self.lastrowid = row[0] if row else None
            return self

        if s.startswith("insert into tickets") and "returning" not in s and "on conflict" not in s:
            sql2 = sql2.rstrip().rstrip(";") + " RETURNING id"
            self._cur.execute(sql2, params)
            row = self._cur.fetchone()
            self.lastrowid = row[0] if row else None
            return self

        self._cur.execute(sql2, params)
        return self

    def executemany(self, sql, seq_of_params):
        sql2 = sql.replace("?", "%s")
        self._cur.executemany(sql2, seq_of_params)
        return self

    def fetchone(self):
        return self._cur.fetchone()

    def fetchall(self):
        return self._cur.fetchall()

    def __iter__(self):
        return iter(self._cur)

    def __getattr__(self, name):
        return getattr(self._cur, name)


class _PgConnProxy:
    def __init__(self, conn):
        self._conn = conn

    def cursor(self):
        cur = self._conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        return _PgCursorWrapper(cur)

    def commit(self):
        return self._conn.commit()

    def close(self):
        # Ne PAS fermer la connexion persistante
        pass

    def __getattr__(self, name):
        return getattr(self._conn, name)


def get_db():
    """Retourne une connexion DB.

    - SQLite si aucune config Supabase
    - Postgres (Supabase) si secrets/env présents
    """
    if is_postgres():
        return _PgConnProxy(_pg_connect())

    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    # Sur Supabase/Postgres, le schéma est créé via le SQL Editor.
    if is_postgres():
        return
    conn = get_db()
    c = conn.cursor()
    
    c.execute("""CREATE TABLE IF NOT EXISTS clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT, prenom TEXT, telephone TEXT UNIQUE, email TEXT,
        societe TEXT,
        carte_camby INTEGER DEFAULT 0,
        date_creation TEXT DEFAULT CURRENT_TIMESTAMP)""")
    
    c.execute("""CREATE TABLE IF NOT EXISTS tickets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ticket_code TEXT UNIQUE,
        client_id INTEGER,
        categorie TEXT, marque TEXT, modele TEXT, modele_autre TEXT,
        imei TEXT,
        panne TEXT, panne_detail TEXT,
        pin TEXT, pattern TEXT,
        notes_client TEXT, notes_internes TEXT,
        commentaire_client TEXT,
        reparation_supp TEXT, prix_supp REAL,
        devis_estime REAL, acompte REAL DEFAULT 0, tarif_final REAL,
        personne_charge TEXT,
        technicien_assigne TEXT,
        commande_piece INTEGER DEFAULT 0,
        date_recuperation TEXT,
        client_contacte INTEGER DEFAULT 0,
        client_accord INTEGER DEFAULT 0,
        paye INTEGER DEFAULT 0,
        msg_whatsapp INTEGER DEFAULT 0,
        msg_sms INTEGER DEFAULT 0,
        msg_email INTEGER DEFAULT 0,
        statut TEXT DEFAULT 'En attente de diagnostic',
        date_depot TEXT DEFAULT CURRENT_TIMESTAMP,
        date_maj TEXT DEFAULT CURRENT_TIMESTAMP,
        date_cloture TEXT)""")
    
    c.execute("""CREATE TABLE IF NOT EXISTS params (
        id INTEGER PRIMARY KEY, cle TEXT UNIQUE, valeur TEXT)""")
    
    c.execute("""CREATE TABLE IF NOT EXISTS catalog_marques (
        id INTEGER PRIMARY KEY, categorie TEXT, marque TEXT, UNIQUE(categorie, marque))""")
    
    c.execute("""CREATE TABLE IF NOT EXISTS catalog_modeles (
        id INTEGER PRIMARY KEY, categorie TEXT, marque TEXT, modele TEXT, 
        UNIQUE(categorie, marque, modele))""")
    
    # Table commandes de pièces
    c.execute("""CREATE TABLE IF NOT EXISTS commandes_pieces (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ticket_id INTEGER,
        description TEXT,
        fournisseur TEXT,
        reference TEXT,
        prix REAL,
        statut TEXT DEFAULT 'A commander',
        date_commande TEXT,
        date_reception TEXT,
        notes TEXT,
        date_creation TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (ticket_id) REFERENCES tickets(id))""")
    
    # Table membres équipe
    c.execute("""CREATE TABLE IF NOT EXISTS membres_equipe (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT,
        role TEXT,
        couleur TEXT,
        actif INTEGER DEFAULT 1)""")
    
    # Migrations diverses - commit après chaque pour PostgreSQL
    migrations = [
        "ALTER TABLE tickets ADD COLUMN commentaire_client TEXT",
        "ALTER TABLE tickets ADD COLUMN imei TEXT",
        "ALTER TABLE tickets ADD COLUMN reparation_supp TEXT",
        "ALTER TABLE tickets ADD COLUMN prix_supp REAL",
        "ALTER TABLE tickets ADD COLUMN commande_piece INTEGER DEFAULT 0",
        "ALTER TABLE tickets ADD COLUMN technicien_assigne TEXT",
        "ALTER TABLE tickets ADD COLUMN date_recuperation TEXT",
        "ALTER TABLE tickets ADD COLUMN client_contacte INTEGER DEFAULT 0",
        "ALTER TABLE tickets ADD COLUMN client_accord INTEGER DEFAULT 0",
        "ALTER TABLE tickets ADD COLUMN paye INTEGER DEFAULT 0",
        "ALTER TABLE tickets ADD COLUMN msg_whatsapp INTEGER DEFAULT 0",
        "ALTER TABLE tickets ADD COLUMN msg_sms INTEGER DEFAULT 0",
        "ALTER TABLE tickets ADD COLUMN msg_email INTEGER DEFAULT 0",
        "ALTER TABLE clients ADD COLUMN societe TEXT",
        "ALTER TABLE clients ADD COLUMN carte_camby INTEGER DEFAULT 0",
        "ALTER TABLE tickets ADD COLUMN type_ecran TEXT",
        "ALTER TABLE tickets ADD COLUMN date_cloture TEXT",
    ]
    for sql in migrations:
        try:
            c.execute(sql)
            conn.commit()  # Commit après chaque migration réussie
        except Exception as e:
            try:
                conn.rollback()  # Rollback si erreur (colonne existe déjà)
            except:
                pass
    
    # Initialiser les membres équipe par défaut
    c.execute("SELECT COUNT(*) FROM membres_equipe")
    if c.fetchone()[0] == 0:
        for m in MEMBRES_EQUIPE_DEFAUT:
            c.execute("INSERT INTO membres_equipe (nom, role, couleur) VALUES (?, ?, ?)", 
                     (m['nom'], m['role'], m['couleur']))
        conn.commit()
    
    # Params défaut
    params = {
        "PIN_ACCUEIL": "2626", "PIN_TECH": "2626",
        "TEL_BOUTIQUE": "04 79 60 89 22",
        "ADRESSE_BOUTIQUE": "79 Place Saint Léger, 73000 Chambéry",
        "NOM_BOUTIQUE": "Klikphone",
        "HORAIRES_BOUTIQUE": "Lundi-Samedi 10h-19h",
        "URL_SUIVI": "https://klikphone-sav.streamlit.app",
        "SMTP_HOST": "",
        "SMTP_PORT": "587",
        "SMTP_USER": "",
        "SMTP_PASS": "",
        "SMTP_FROM": "",
        "SMTP_FROM_NAME": "Klikphone"
    }
    for k, v in params.items():
        c.execute("INSERT OR IGNORE INTO params (cle, valeur) VALUES (?, ?)", (k, v))
    
    # Catalogue
    c.execute("SELECT COUNT(*) FROM catalog_marques")
    if c.fetchone()[0] == 0:
        for cat, marques in MARQUES.items():
            for m in marques:
                c.execute("INSERT OR IGNORE INTO catalog_marques (categorie, marque) VALUES (?, ?)", (cat, m))
        for (cat, marque), modeles_list in MODELES.items():
            for m in modeles_list:
                c.execute("INSERT OR IGNORE INTO catalog_modeles (categorie, marque, modele) VALUES (?, ?, ?)", (cat, marque, m))
    
    conn.commit()
    conn.close()

def get_param(k):
    """Récupère un paramètre (avec cache)"""
    cache_key = f"_cache_param_{k}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]
    
    conn = get_db()
    r = conn.cursor().execute("SELECT valeur FROM params WHERE cle=?", (k,)).fetchone()
    conn.close()
    result = r["valeur"] if r else ""
    st.session_state[cache_key] = result
    return result

def set_param(k, v):
    conn = get_db()
    conn.cursor().execute("INSERT OR REPLACE INTO params (cle, valeur) VALUES (?, ?)", (k, v))
    conn.commit()
    conn.close()
    # Invalider le cache
    cache_key = f"_cache_param_{k}"
    if cache_key in st.session_state:
        del st.session_state[cache_key]

def get_marques(cat):
    """Récupère les marques pour une catégorie (avec cache)"""
    # Cache en session pour éviter les requêtes répétées
    cache_key = f"_cache_marques_{cat}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]
    
    conn = get_db()
    r = [row["marque"] for row in conn.cursor().execute(
        "SELECT marque FROM catalog_marques WHERE categorie=? ORDER BY marque", (cat,)).fetchall()]
    conn.close()
    result = r if r else MARQUES.get(cat, ["Autre"])
    st.session_state[cache_key] = result
    return result

def get_modeles(cat, marque):
    """Récupère les modèles pour une catégorie/marque (avec cache)"""
    cache_key = f"_cache_modeles_{cat}_{marque}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]
    
    conn = get_db()
    r = [row["modele"] for row in conn.cursor().execute(
        "SELECT modele FROM catalog_modeles WHERE categorie=? AND marque=? ORDER BY modele", 
        (cat, marque)).fetchall()]
    conn.close()
    result = r if r else MODELES.get((cat, marque), ["Autre"])
    st.session_state[cache_key] = result
    return result

def ajouter_marque(cat, marque):
    try:
        conn = get_db()
        conn.cursor().execute("INSERT INTO catalog_marques (categorie, marque) VALUES (?, ?)", (cat, marque))
        conn.commit()
        conn.close()
        return True
    except: return False

def ajouter_modele(cat, marque, modele):
    try:
        conn = get_db()
        conn.cursor().execute("INSERT INTO catalog_modeles (categorie, marque, modele) VALUES (?, ?, ?)", (cat, marque, modele))
        conn.commit()
        conn.close()
        return True
    except: return False

# =============================================================================
# MÉTIER
# =============================================================================
def get_or_create_client(nom, tel, prenom="", email="", societe="", carte_camby=0):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM clients WHERE telephone=?", (tel,))
    r = c.fetchone()
    if r:
        cid = r["id"]
        c.execute("UPDATE clients SET nom=?, prenom=?, email=?, societe=?, carte_camby=? WHERE id=?", (nom, prenom, email, societe, carte_camby, cid))
    else:
        c.execute("INSERT INTO clients (nom, prenom, telephone, email, societe, carte_camby) VALUES (?,?,?,?,?,?)", (nom, prenom, tel, email, societe, carte_camby))
        cid = c.lastrowid
    conn.commit()
    conn.close()
    return cid

def check_client_exists(tel):
    """Vérifie si un client existe déjà par son téléphone"""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, nom, prenom FROM clients WHERE telephone=?", (tel,))
    r = c.fetchone()
    conn.close()
    return dict(r) if r else None

@st.cache_data(ttl=30)
def get_all_clients():
    """Récupère tous les clients - CACHED 30s"""
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT c.*, COUNT(t.id) as nb_tickets 
                 FROM clients c 
                 LEFT JOIN tickets t ON c.id = t.client_id 
                 GROUP BY c.id 
                 ORDER BY c.nom, c.prenom""")
    clients = [dict(row) for row in c.fetchall()]
    conn.close()
    return clients

def chercher_clients(query):
    """Recherche clients avec normalisation (sans accents, insensible à la casse)"""
    if not query or len(query) < 2:
        return []
    
    query_norm = normalize_search(query)
    clients = get_all_clients()
    
    resultats = []
    for c in clients:
        # Normaliser les champs du client
        nom_norm = normalize_search(c.get('nom', ''))
        prenom_norm = normalize_search(c.get('prenom', ''))
        tel_norm = normalize_search(c.get('telephone', ''))
        email_norm = normalize_search(c.get('email', ''))
        
        # Rechercher dans tous les champs
        if (query_norm in nom_norm or 
            query_norm in prenom_norm or 
            query_norm in tel_norm or
            query_norm in email_norm or
            query_norm in f"{nom_norm} {prenom_norm}" or
            query_norm in f"{prenom_norm} {nom_norm}"):
            resultats.append(c)
    
    return resultats

def get_client_by_id(client_id):
    """Récupère un client par son ID"""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM clients WHERE id=?", (client_id,))
    r = c.fetchone()
    conn.close()
    return dict(r) if r else None

def update_client(client_id, nom=None, prenom=None, telephone=None, email=None, societe=None):
    """Met à jour les informations d'un client"""
    conn = get_db()
    c = conn.cursor()
    updates = []
    params = []
    if nom is not None: updates.append("nom=?"); params.append(nom)
    if prenom is not None: updates.append("prenom=?"); params.append(prenom)
    if telephone is not None: updates.append("telephone=?"); params.append(telephone)
    if email is not None: updates.append("email=?"); params.append(email)
    if societe is not None: updates.append("societe=?"); params.append(societe)
    if updates:
        params.append(client_id)
        c.execute(f"UPDATE clients SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
    conn.close()

def supprimer_client(client_id):
    """Supprime un client et tous ses tickets associés"""
    conn = get_db()
    c = conn.cursor()
    # Supprimer d'abord les tickets du client
    c.execute("DELETE FROM tickets WHERE client_id=?", (client_id,))
    # Puis supprimer le client
    c.execute("DELETE FROM clients WHERE id=?", (client_id,))
    conn.commit()
    conn.close()
    return True

def supprimer_ticket(ticket_id):
    """Supprime une réparation (ticket) tout en conservant le client"""
    conn = get_db()
    c = conn.cursor()
    # Supprimer les commandes de pièces associées
    c.execute("DELETE FROM commandes_pieces WHERE ticket_id=?", (ticket_id,))
    # Supprimer le ticket
    c.execute("DELETE FROM tickets WHERE id=?", (ticket_id,))
    conn.commit()
    conn.close()
    # Invalider les caches
    try:
        clear_tickets_cache()
        clear_commandes_cache()
        clear_ticket_full_cache()
    except:
        pass
    return True

def search_clients(query):
    """Recherche des clients par nom, prénom, téléphone ou société - SANS accents ni majuscules"""
    if not query or len(query) < 2:
        return []
    return chercher_clients(query)

# Fonctions commandes de pièces
FOURNISSEURS = ["Utopya", "Piece2mobile", "Amazon", "Mobilax", "Autre"]

@st.cache_data(ttl=20)  # Cache 20 secondes pour les commandes
def _get_commandes_pieces_cached(ticket_id, statut):
    """Version cachée de get_commandes_pieces"""
    conn = get_db()
    c = conn.cursor()
    q = """SELECT cp.*, t.ticket_code, t.marque, t.modele, t.modele_autre, t.panne_detail, t.categorie,
           c.nom as client_nom, c.prenom as client_prenom, c.telephone as client_tel, c.email as client_email
           FROM commandes_pieces cp
           LEFT JOIN tickets t ON cp.ticket_id = t.id
           LEFT JOIN clients c ON t.client_id = c.id
           WHERE 1=1"""
    params = []
    if ticket_id:
        q += " AND cp.ticket_id = ?"
        params.append(ticket_id)
    if statut:
        q += " AND cp.statut = ?"
        params.append(statut)
    q += " ORDER BY cp.date_creation DESC"
    c.execute(q, params)
    commandes = [dict(row) for row in c.fetchall()]
    conn.close()
    return commandes

def get_commandes_pieces(ticket_id=None, statut=None):
    """Récupère les commandes de pièces avec cache"""
    return _get_commandes_pieces_cached(ticket_id, statut)

def clear_commandes_cache():
    """Invalide le cache des commandes"""
    try:
        _get_commandes_pieces_cached.clear()
    except:
        pass

def ajouter_commande_piece(ticket_id, description, fournisseur, reference="", prix=0, notes=""):
    """Ajoute une commande de pièce"""
    conn = get_db()
    c = conn.cursor()
    c.execute("""INSERT INTO commandes_pieces 
                 (ticket_id, description, fournisseur, reference, prix, notes, statut)
                 VALUES (?, ?, ?, ?, ?, ?, 'A commander')""",
              (ticket_id, description, fournisseur, reference, prix, notes))
    conn.commit()
    conn.close()
    clear_commandes_cache()

def update_commande_piece(commande_id, statut=None, date_commande=None, date_reception=None, notes=None):
    """Met à jour une commande de pièce"""
    conn = get_db()
    c = conn.cursor()
    updates = []
    params = []
    if statut: updates.append("statut=?"); params.append(statut)
    if date_commande: updates.append("date_commande=?"); params.append(date_commande)
    if date_reception: updates.append("date_reception=?"); params.append(date_reception)
    if notes is not None: updates.append("notes=?"); params.append(notes)
    if updates:
        params.append(commande_id)
        c.execute(f"UPDATE commandes_pieces SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
    conn.close()
    clear_commandes_cache()

def delete_commande_piece(commande_id):
    """Supprime une commande de pièce"""
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM commandes_pieces WHERE id=?", (commande_id,))
    conn.commit()
    conn.close()
    clear_commandes_cache()

# Fonctions membres équipe
@st.cache_data(ttl=60)  # Cache pendant 60 secondes
def get_membres_equipe():
    """Récupère tous les membres de l'équipe"""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM membres_equipe WHERE actif=1 ORDER BY nom")
    membres = [dict(row) for row in c.fetchall()]
    conn.close()
    return membres

def ajouter_membre_equipe(nom, role, couleur):
    """Ajoute un membre à l'équipe"""
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO membres_equipe (nom, role, couleur) VALUES (?, ?, ?)", (nom, role, couleur))
    conn.commit()
    conn.close()

def update_membre_equipe(membre_id, nom=None, role=None, couleur=None, actif=None):
    """Met à jour un membre de l'équipe"""
    conn = get_db()
    c = conn.cursor()
    updates = []
    params = []
    if nom: updates.append("nom=?"); params.append(nom)
    if role: updates.append("role=?"); params.append(role)
    if couleur: updates.append("couleur=?"); params.append(couleur)
    if actif is not None: updates.append("actif=?"); params.append(actif)
    if updates:
        params.append(membre_id)
        c.execute(f"UPDATE membres_equipe SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
    conn.close()

def supprimer_membre_equipe(membre_id):
    """Désactive un membre de l'équipe"""
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE membres_equipe SET actif=0 WHERE id=?", (membre_id,))
    conn.commit()
    conn.close()

def creer_ticket(client_id, cat, marque, modele, modele_autre, panne, panne_detail, pin, pattern, notes, imei="", commande_piece=0):
    conn = get_db()
    c = conn.cursor()
    c.execute("""INSERT INTO tickets 
        (client_id, categorie, marque, modele, modele_autre, imei, panne, panne_detail, pin, pattern, notes_client, commande_piece, statut) 
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,'En attente de diagnostic')""", 
        (client_id, cat, marque, modele, modele_autre, imei, panne, panne_detail, pin, pattern, notes, commande_piece))
    tid = c.lastrowid
    code = f"KP-{tid:06d}"
    c.execute("UPDATE tickets SET ticket_code=? WHERE id=?", (code, tid))
    conn.commit()
    conn.close()
    clear_tickets_cache()
    return code

def get_ticket(tid=None, code=None):
    conn = get_db()
    c = conn.cursor()
    if tid: c.execute("SELECT * FROM tickets WHERE id=?", (tid,))
    elif code: c.execute("SELECT * FROM tickets WHERE ticket_code=?", (code,))
    else: return None
    r = c.fetchone()
    conn.close()
    return dict(r) if r else None

@st.cache_data(ttl=10)  # Cache 10 secondes pour ticket individuel
def _get_ticket_full_cached(tid, code):
    """Version cachée de get_ticket_full"""
    conn = get_db()
    c = conn.cursor()
    q = """SELECT t.*, c.nom as client_nom, c.prenom as client_prenom, 
           c.telephone as client_tel, c.email as client_email,
           c.societe as client_societe, c.carte_camby as client_carte_camby
           FROM tickets t JOIN clients c ON t.client_id=c.id"""
    if tid: c.execute(q + " WHERE t.id=?", (tid,))
    elif code: c.execute(q + " WHERE t.ticket_code=?", (code,))
    else: 
        conn.close()
        return None
    r = c.fetchone()
    conn.close()
    return dict(r) if r else None

def get_ticket_full(tid=None, code=None):
    """Récupère un ticket complet avec cache"""
    return _get_ticket_full_cached(tid, code)

def clear_ticket_full_cache():
    """Invalide le cache du ticket"""
    try:
        _get_ticket_full_cached.clear()
    except:
        pass

def update_ticket(tid, **kw):
    if not kw: return
    conn = get_db()
    c = conn.cursor()
    fields = ", ".join([f"{k}=?" for k in kw.keys()])
    vals = list(kw.values()) + [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), tid]
    c.execute(f"UPDATE tickets SET {fields}, date_maj=? WHERE id=?", vals)
    conn.commit()
    conn.close()
    # Invalider les caches
    clear_tickets_cache()
    clear_ticket_full_cache()

def changer_statut(tid, statut):
    conn = get_db()
    c = conn.cursor()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if statut == "Clôturé":
        c.execute("UPDATE tickets SET statut=?, date_maj=?, date_cloture=? WHERE id=?", (statut, now, now, tid))
    else:
        c.execute("UPDATE tickets SET statut=?, date_maj=? WHERE id=?", (statut, now, tid))
    conn.commit()
    conn.close()
    # Invalider les caches
    clear_tickets_cache()
    clear_ticket_full_cache()

@st.cache_data(ttl=15)  # Cache 15 secondes pour les tickets
def _chercher_tickets_cached(statut, tel, code, nom):
    """Version cachée de chercher_tickets"""
    conn = get_db()
    c = conn.cursor()
    q = """SELECT t.*, c.nom as client_nom, c.prenom as client_prenom, c.telephone as client_tel,
           c.carte_camby as client_carte_camby
           FROM tickets t JOIN clients c ON t.client_id=c.id WHERE 1=1"""
    p = []
    if statut: q += " AND t.statut=?"; p.append(statut)
    if tel: q += " AND c.telephone LIKE ?"; p.append(f"%{tel}%")
    if code: q += " AND t.ticket_code LIKE ?"; p.append(f"%{code}%")
    if nom: q += " AND (c.nom LIKE ? OR c.prenom LIKE ?)"; p.extend([f"%{nom}%", f"%{nom}%"])
    q += " ORDER BY t.date_depot DESC"
    c.execute(q, p)
    r = [dict(row) for row in c.fetchall()]
    conn.close()
    return r

def chercher_tickets(statut=None, tel=None, code=None, nom=None):
    """Recherche tickets avec cache"""
    return _chercher_tickets_cached(statut, tel, code, nom)

def clear_tickets_cache():
    """Invalide le cache des tickets"""
    try:
        _chercher_tickets_cached.clear()
    except:
        pass

def ajouter_note(tid, note):
    t = get_ticket(tid=tid)
    if not t: return
    ts = datetime.now().strftime("%d/%m %H:%M")
    n = f"[{ts}] {note}"
    notes = t.get("notes_internes") or ""
    notes = notes + "\n" + n if notes else n
    update_ticket(tid, notes_internes=notes)

# =============================================================================
# UTILS
# =============================================================================
def fmt_date(d):
    if not d: return "N/A"
    try:
        # Gérer les microsecondes si présentes
        d_str = str(d)[:19]  # Couper les microsecondes
        return datetime.strptime(d_str, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
    except:
        try:
            # Essayer format date seule
            return datetime.strptime(str(d)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except:
            return str(d)[:16] if d else "N/A"

def fmt_prix(p):
    return f"{p:.2f} €" if p else "N/A"

def qr_url(data):
    return f"https://api.qrserver.com/v1/create-qr-code/?size=150x150&data={urllib.parse.quote(data)}"

def wa_link(tel, msg):
    t = "".join(filter(str.isdigit, tel))
    if t.startswith("0"): t = "33" + t[1:]
    return f"https://wa.me/{t}?text={urllib.parse.quote(msg)}"

def get_status_class(statut):
    if "diagnostic" in statut.lower(): return "status-diagnostic"
    elif "cours" in statut.lower(): return "status-encours"
    elif "terminée" in statut.lower(): return "status-termine"
    elif "rendu" in statut.lower(): return "status-rendu"
    else: return "status-clôturé"

def sms_link(tel, msg):
    """Genere un lien SMS"""
    t = "".join(filter(str.isdigit, tel))
    if t.startswith("0"): t = "+33" + t[1:]
    return f"sms:{t}?body={urllib.parse.quote(msg)}"

def email_link(email, sujet, msg):
    """Genere un lien mailto"""
    return f"mailto:{email}?subject={urllib.parse.quote(sujet)}&body={urllib.parse.quote(msg)}"

def envoyer_email(destinataire, sujet, message, html_content=None):
    """Envoie un email via SMTP avec option HTML - Encodage UTF-8"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.header import Header
    from email.utils import formataddr
    
    # Récupérer les paramètres SMTP
    smtp_host = get_param("SMTP_HOST")
    smtp_port = get_param("SMTP_PORT")
    smtp_user = get_param("SMTP_USER")
    smtp_pass = get_param("SMTP_PASS")
    smtp_from = get_param("SMTP_FROM")
    smtp_from_name = get_param("SMTP_FROM_NAME") or "Klikphone"
    
    if not smtp_host or not smtp_user or not smtp_pass:
        return False, "Configuration SMTP incomplète. Allez dans Config > Email."
    
    try:
        # Créer le message avec encodage UTF-8
        msg = MIMEMultipart('alternative')
        msg['From'] = formataddr((str(Header(smtp_from_name, 'utf-8')), smtp_from or smtp_user))
        msg['To'] = destinataire
        msg['Subject'] = Header(sujet, 'utf-8')
        
        # Corps du message en texte (UTF-8)
        msg.attach(MIMEText(message, 'plain', 'utf-8'))
        
        # Corps en HTML si fourni (retirer le bouton imprimer)
        if html_content:
            # Retirer le bouton imprimer du HTML envoyé par email
            html_clean = html_content.replace('onclick="window.print()"', 'style="display:none"')
            html_clean = html_clean.replace('IMPRIMER', '')
            msg.attach(MIMEText(html_clean, 'html', 'utf-8'))
        
        # Connexion et envoi - Encoder en bytes UTF-8
        server = smtplib.SMTP(smtp_host, int(smtp_port or 587))
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_from or smtp_user, destinataire, msg.as_bytes())
        server.quit()
        
        return True, "Email envoyé avec succès!"
    except Exception as e:
        return False, f"Erreur d'envoi: {str(e)}"

def envoyer_email_avec_pdf(destinataire, sujet, message, pdf_bytes, filename="document.pdf"):
    """Envoie un email avec une pièce jointe PDF - Encodage UTF-8"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    from email.header import Header
    from email.utils import formataddr
    
    smtp_host = get_param("SMTP_HOST")
    smtp_port = get_param("SMTP_PORT")
    smtp_user = get_param("SMTP_USER")
    smtp_pass = get_param("SMTP_PASS")
    smtp_from = get_param("SMTP_FROM")
    smtp_from_name = get_param("SMTP_FROM_NAME") or "Klikphone"
    
    if not smtp_host or not smtp_user or not smtp_pass:
        return False, "Configuration SMTP incomplète."
    
    try:
        msg = MIMEMultipart()
        msg['From'] = formataddr((str(Header(smtp_from_name, 'utf-8')), smtp_from or smtp_user))
        msg['To'] = destinataire
        msg['Subject'] = Header(sujet, 'utf-8')
        
        msg.attach(MIMEText(message, 'plain', 'utf-8'))
        
        # Ajouter le PDF en pièce jointe
        pdf_part = MIMEApplication(pdf_bytes, _subtype='pdf')
        pdf_part.add_header('Content-Disposition', 'attachment', filename=filename)
        msg.attach(pdf_part)
        
        server = smtplib.SMTP(smtp_host, int(smtp_port or 587))
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_from or smtp_user, destinataire, msg.as_bytes())
        server.quit()
        
        return True, "Email avec PDF envoyé!"
    except Exception as e:
        return False, f"Erreur: {str(e)}"

def html_to_pdf(html_content):
    """Convertit HTML en PDF (utilise une approche simple avec base64)"""
    try:
        # Nettoyer le HTML - retirer les boutons d'impression
        html_clean = html_content.replace('onclick="window.print()"', 'style="display:none !important"')
        html_clean = html_clean.replace('>IMPRIMER', ' style="display:none !important">IMPRIMER')
        
        # Essayer avec weasyprint si disponible
        try:
            from weasyprint import HTML
            pdf_bytes = HTML(string=html_clean).write_pdf()
            return pdf_bytes
        except ImportError:
            pass
        
        # Essayer avec pdfkit si disponible
        try:
            import pdfkit
            pdf_bytes = pdfkit.from_string(html_clean, False)
            return pdf_bytes
        except ImportError:
            pass
        
        # Fallback: retourner None et envoyer en HTML
        return None
    except Exception as e:
        return None

def envoyer_vers_caisse(ticket, payment_override=None):
    """Envoie un ticket de réparation vers caisse.enregistreuse.fr"""
    try:
        import requests
        import streamlit as st

        apikey = get_param("CAISSE_APIKEY")
        shopid = get_param("CAISSE_SHOPID")
        if not apikey or not shopid:
            return False, "Configuration API manquante (APIKEY ou SHOPID)"

        # Lire depuis la config (avec valeurs par défaut)
        caisse_id = (get_param("CAISSE_ID") or "49343").strip()
        user_id = (get_param("CAISSE_USER_ID") or "42867").strip()

        if not caisse_id.isdigit():
            return False, f"CAISSE_ID invalide: '{caisse_id}'"
        if not user_id.isdigit():
            return False, f"CAISSE_USER_ID invalide: '{user_id}'"

        delivery_method = (get_param("CAISSE_DELIVERY_METHOD") or "4").strip()
        if not delivery_method.isdigit():
            delivery_method = "4"

        payment_mode = str(payment_override) if payment_override is not None else "-1"
        try:
            int(payment_mode)
        except:
            return False, f"payment invalide: {payment_mode}"

        devis = float(ticket.get("devis_estime") or 0)
        prix_supp = float(ticket.get("prix_supp") or 0)
        total = devis + prix_supp
        if total <= 0:
            return False, "Montant total à 0 : rien à envoyer"

        modele_txt = f"{ticket.get('marque','')} {ticket.get('modele','')}".strip()
        panne_txt = (ticket.get("panne") or "").strip()
        if ticket.get("panne_detail"):
            panne_txt += f" ({ticket['panne_detail']})"

        description = f"Reparation {modele_txt} - {panne_txt}".strip()
        if ticket.get("type_ecran"):
            description += f" [{ticket['type_ecran']}]"
        description = description.replace("_", " ")

        # URL avec idcaisse dans le querystring (comme leur exemple curl)
        api_url = f"https://caisse.enregistreuse.fr/workers/webapp.php?idboutique={shopid}&key={apikey}&idUser={user_id}&idcaisse={caisse_id}&payment={payment_mode}&deliveryMethod={delivery_method}"
        
        st.info(f"🔗 URL: ...idUser={user_id}&idcaisse={caisse_id}&payment={payment_mode}")

        # Payload POST : seulement les données variables (client, items)
        payload = [
            ("publicComment", f"Ticket: {ticket.get('ticket_code','')}"),
        ]

        # Client optionnel
        if ticket.get("client_nom") or ticket.get("client_prenom"):
            payload += [
                ("client[nom]", ticket.get("client_nom", "")),
                ("client[prenom]", ticket.get("client_prenom", "")),
                ("client[telephone]", ticket.get("client_tel", "")),
            ]
            if ticket.get("client_email"):
                payload.append(("client[email]", ticket.get("client_email")))

        # Items
        if ticket.get("reparation_supp") and prix_supp > 0:
            payload.append(("itemsList[]", f"Free_{devis:.2f}_{description}"))
            rep_supp = (ticket.get("reparation_supp") or "Reparation supplementaire").replace("_", " ")
            payload.append(("itemsList[]", f"Free_{prix_supp:.2f}_{rep_supp}"))
        else:
            payload.append(("itemsList[]", f"Free_{total:.2f}_{description}"))

        st.success(f"📤 Envoi avec idcaisse={caisse_id} dans l'URL")
        
        # Envoyer
        res = requests.post(api_url, data=payload, timeout=15)

        if res.status_code != 200:
            return False, f"HTTP {res.status_code}: {res.text[:300]}"

        try:
            data = res.json()
            if data.get("result") == "OK":
                return True, f"✅ Vente créée ! (idcaisse={caisse_id}, idUser={user_id})"
            return False, f"Erreur API: {data.get('errorMessage', data)}"
        except:
            txt = (res.text or "").strip()
            if txt.isdigit() or "OK" in txt.upper():
                return True, f"✅ Vente créée ! ID: {txt}"
            return False, f"Réponse: {txt[:300]}"

    except Exception as e:
        return False, f"Erreur: {e}"

def export_clients_excel():
    """Exporte la liste des clients en Excel"""
    try:
        import io
        clients = get_all_clients()
        
        # Créer le fichier Excel avec openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Clients Klikphone"
            
            # Style header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = ["Nom", "Prénom", "Société", "Téléphone", "Email", "Date création", "Nb tickets"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Données
            for row_idx, client in enumerate(clients, 2):
                ws.cell(row=row_idx, column=1, value=client.get('nom', '')).border = thin_border
                ws.cell(row=row_idx, column=2, value=client.get('prenom', '')).border = thin_border
                ws.cell(row=row_idx, column=3, value=client.get('societe', '') or '').border = thin_border
                ws.cell(row=row_idx, column=4, value=client.get('telephone', '')).border = thin_border
                ws.cell(row=row_idx, column=5, value=client.get('email', '')).border = thin_border
                ws.cell(row=row_idx, column=6, value=client.get('date_creation', '')[:10] if client.get('date_creation') else '').border = thin_border
                ws.cell(row=row_idx, column=7, value=client.get('nb_tickets', 0)).border = thin_border
            
            # Ajuster largeur colonnes
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            
            # Sauvegarder dans un buffer
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue(), "clients_klikphone.xlsx"
            
        except ImportError:
            # Fallback: créer un CSV
            import csv
            output = io.StringIO()
            writer = csv.writer(output, delimiter=';')
            writer.writerow(["Nom", "Prénom", "Société", "Téléphone", "Email", "Date création", "Nb tickets"])
            for client in clients:
                writer.writerow([
                    client.get('nom', ''),
                    client.get('prenom', ''),
                    client.get('societe', '') or '',
                    client.get('telephone', ''),
                    client.get('email', ''),
                    client.get('date_creation', '')[:10] if client.get('date_creation') else '',
                    client.get('nb_tickets', 0)
                ])
            return output.getvalue().encode('utf-8-sig'), "clients_klikphone.csv"
            
    except Exception as e:
        return None, str(e)

def envoyer_sms_api(tel, message):
    """Placeholder pour API SMS (à configurer selon fournisseur)"""
    return False, "API SMS non configurée. Utilisez le lien SMS."

def get_messages_predefs(t):
    """Retourne les messages prédéfinis pour un ticket"""
    prénom = t.get('client_prenom', '') or 'Client'
    nom = t.get('client_nom', '')
    marque = t.get('marque', '') or 'votre appareil'
    modèle = t.get('modele', '')
    code = t.get('ticket_code', '')
    devis = float(t.get('devis_estime') or 0)
    tarif = float(t.get('tarif_final') or 0)
    prix_supp = float(t.get('prix_supp') or 0)
    reparation_supp = t.get('reparation_supp', '')
    
    # Paramètres dynamiques de la boutique
    adresse = get_param("ADRESSE_BOUTIQUE") or "79 Place Saint Léger, 73000 Chambéry"
    horaires = get_param("HORAIRES_BOUTIQUE") or "Lundi-Samedi 10h-19h"
    tel_boutique = get_param("TEL_BOUTIQUE") or "04 79 60 89 22"
    nom_boutique = get_param("NOM_BOUTIQUE") or "Klikphone"
    
    # Calcul du montant TOTAL (inclut réparation supplémentaire)
    total_devis = devis + prix_supp
    total_tarif = tarif + prix_supp if tarif > 0 else 0
    
    # Formater le montant final
    if total_tarif > 0:
        montant = f"{total_tarif:.2f} €"
    elif total_devis > 0:
        montant = f"{total_devis:.2f} €"
    else:
        montant = "Nous consulter"
    
    # Formater le devis (inclut prix_supp)
    devis_txt = f"{total_devis:.2f} €" if total_devis > 0 else "Nous consulter"
    
    # Détail si réparation supplémentaire
    detail_supp = ""
    if reparation_supp and prix_supp > 0:
        detail_supp = f"\n(dont {reparation_supp}: {prix_supp:.2f} €)"
    
    messages = {
        "-- Choisir un message --": "",
        
        "Appareil reçu": f"""Bonjour {prénom},

Nous vous informons que votre appareil {marque} {modèle} a bien été réceptionné à la boutique {nom_boutique}.

Numéro de suivi : {code}

Nous allons procéder au diagnostic et reviendrons vers vous dans les plus brefs délais.

Cordialement,
L'équipe {nom_boutique}
{tel_boutique}""",

        "Diagnostic en cours": f"""Bonjour {prénom},

Nous vous informons que le diagnostic de votre appareil {marque} {modèle} est actuellement en cours à la boutique {nom_boutique}.

Numéro de suivi : {code}

Nous reviendrons vers vous rapidement avec le résultat du diagnostic.

Cordialement,
L'équipe {nom_boutique}
{tel_boutique}""",

        "Devis à valider": f"""Bonjour {prénom},

Le diagnostic de votre appareil {marque} {modèle} est terminé.

Devis de réparation : {devis_txt}{detail_supp}

Merci de nous confirmer votre accord pour procéder à la réparation en répondant à ce message ou en nous appelant.

Cordialement,
L'équipe {nom_boutique}
{tel_boutique}""",

        "En cours de réparation": f"""Bonjour {prénom},

Nous vous informons que votre appareil {marque} {modèle} est actuellement en cours de réparation à la boutique {nom_boutique}.

Numéro de suivi : {code}

Nous vous prévenons dès que la réparation sera terminée.

Cordialement,
L'équipe {nom_boutique}
{tel_boutique}""",

        "Attente de pièce": f"""Bonjour {prénom},

Nous vous informons que nous sommes en attente d'une pièce pour la réparation de votre appareil {marque} {modèle}.

Délai estimé : 2 à 5 jours ouvrables.

Nous vous prévenons dès réception de la pièce.

Cordialement,
L'équipe {nom_boutique}
{tel_boutique}""",

        "Appareil prêt": f"""Bonjour {prénom},

Nous vous informons que votre appareil {marque} {modèle} est prêt à être récupéré à la boutique {nom_boutique}.

Montant à régler : {montant}{detail_supp}

Adresse : {adresse}
Horaires : {horaires}

N'oubliez pas votre pièce d'identité.

Cordialement,
L'équipe {nom_boutique}
{tel_boutique}""",

        "Relance récupération": f"""Bonjour {prénom},

Nous vous rappelons que votre appareil {marque} {modèle} est prêt et vous attend à la boutique {nom_boutique} depuis plusieurs jours.

Merci de venir le récupérer dans les meilleurs délais.

Adresse : {adresse}
Horaires : {horaires}

Cordialement,
L'équipe {nom_boutique}
{tel_boutique}""",

        "Non réparable": f"""Bonjour {prénom},

Après diagnostic approfondi, nous avons le regret de vous informer que votre appareil {marque} {modèle} n'est malheureusement pas réparable.

Vous pouvez venir le récupérer à la boutique. Aucun frais ne vous sera facturé pour le diagnostic.

Nous restons à votre disposition pour toute question.

Cordialement,
L'équipe {nom_boutique}
{tel_boutique}""",

        "Rappel RDV": f"""Bonjour {prénom},

Nous vous rappelons votre rendez-vous à la boutique {nom_boutique} pour votre appareil {marque} {modèle}.

Adresse : {adresse}

À bientôt !

L'équipe {nom_boutique}
{tel_boutique}""",

        "Personnalisé": ""
    }
    return messages

# =============================================================================
# TICKETS HTML
# =============================================================================
def ticket_client_html(t, for_email=False):
    """Ticket client - NOIR SUR BLANC pour impression"""
    panne = t.get("panne", "")
    if t.get("panne_detail"): panne += f" ({t['panne_detail']})"
    modele_txt = t.get("modele", "")
    if t.get("modele_autre"): modele_txt += f" ({t['modele_autre']})"
    
    type_ecran = (t.get('type_ecran') or '').strip()
    comment_public = (t.get("commentaire_client") or "").strip()
    
    # Label dynamique selon la panne
    panne_base = t.get("panne", "")
    if panne_base == "Écran casse":
        precision_label = "Type écran"
        precision_icon = "📱"
    else:
        precision_label = "Précision"
        precision_icon = "📝"
    
    # Tarifs avec réparation supplémentaire
    devis = t.get('devis_estime') or 0
    acompte = t.get('acompte') or 0
    rep_supp = t.get('reparation_supp') or ''
    prix_supp = t.get('prix_supp') or 0
    tarif = t.get('tarif_final') or (devis + prix_supp)
    reste = max(0, tarif - acompte)
    
    # URL de suivi
    url_suivi = get_param("URL_SUIVI") or "https://klikphone-sav.streamlit.app"
    ticket_code = t.get('ticket_code', '')
    url_suivi_ticket = f"{url_suivi}?ticket={ticket_code}"
    qr_url_val = f"https://api.qrserver.com/v1/create-qr-code/?size=200x200&data={urllib.parse.quote(url_suivi_ticket)}"
    
    # HTML conditionnel
    type_ecran_html = f'<div class="info-row"><span class="label">{precision_label}</span><span class="value">{type_ecran}</span></div>' if type_ecran else ""
    rep_supp_html = f'<div class="tarif-row"><span>{rep_supp}</span><span>{prix_supp:.2f} €</span></div>' if rep_supp and prix_supp else ""
    comment_html = f'<div class="notes-box"><div class="notes-title">💬 Commentaire</div>{comment_public}</div>' if comment_public else ""
    
    # Version EMAIL (colorée)
    if for_email:
        type_ecran_email = f'<div style="color:#3b82f6;margin-top:5px;">{precision_icon} {precision_label}: {type_ecran}</div>' if type_ecran else ""
        rep_supp_email = f'<div class="tarif-row"><span>{rep_supp}</span><span>{prix_supp:.2f} €</span></div>' if rep_supp and prix_supp else ""
        comment_section = f"""
        <div style="background:#fff7ed;border-radius:8px;padding:12px;margin:15px 0;">
            <div style="font-weight:600;color:#ea580c;font-size:12px;margin-bottom:5px;">💬 Commentaire</div>
            <div style="color:#1e293b;">{comment_public}</div>
        </div>
        """ if comment_public else ""
        
        return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
body {{ font-family: Arial, sans-serif; font-size: 14px; margin: 0; padding: 20px; background: #f5f5f5; }}
.ticket {{ max-width: 400px; margin: 0 auto; background: #fff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.1); }}
.header {{ background: linear-gradient(135deg, #f97316, #ea580c); color: white; padding: 20px; text-align: center; }}
.header h1 {{ margin: 0; font-size: 24px; }}
.header p {{ margin: 5px 0 0; opacity: 0.9; font-size: 12px; }}
.ticket-num {{ background: #1e293b; color: white; text-align: center; padding: 15px; font-size: 20px; font-weight: bold; letter-spacing: 2px; }}
.content {{ padding: 20px; }}
.section {{ margin-bottom: 15px; padding-bottom: 15px; border-bottom: 1px solid #eee; }}
.section:last-child {{ border-bottom: none; }}
.section-title {{ font-weight: bold; color: #f97316; font-size: 12px; text-transform: uppercase; margin-bottom: 8px; }}
.tarif-box {{ background: #f8fafc; border-radius: 8px; padding: 15px; margin: 15px 0; }}
.tarif-row {{ display: flex; justify-content: space-between; margin: 5px 0; }}
.tarif-row.total {{ font-weight: bold; font-size: 16px; border-top: 1px solid #ddd; padding-top: 10px; margin-top: 10px; }}
.qr-section {{ text-align: center; padding: 15px; background: #f8fafc; }}
.qr-section img {{ width: 100px; height: 100px; }}
.footer {{ background: #1e293b; color: white; padding: 15px; text-align: center; font-size: 12px; }}
</style>
</head>
<body>
<div class="ticket">
<div class="header">
<h1>KLIKPHONE</h1>
<p>Spécialiste Apple - 79 Place Saint Léger, Chambéry</p>
</div>
<div class="ticket-num">TICKET N° {t['ticket_code']}</div>
<div class="content">
<div class="section">
<div class="section-title">👤 Client</div>
<div style="font-weight:600;font-size:16px;">{t.get('client_nom','')} {t.get('client_prenom','')}</div>
<div style="color:#64748b;">Tél: {t.get('client_tel','')}</div>
</div>
<div class="section">
<div class="section-title">📱 Appareil</div>
<div style="font-weight:600;">{t.get('marque','')} {modele_txt}</div>
<div style="color:#64748b;margin-top:5px;">Motif: {panne}</div>
{type_ecran_email}
</div>
<div class="tarif-box">
<div class="tarif-row"><span>Devis estimé</span><span>{devis:.2f} €</span></div>
{rep_supp_email}
<div class="tarif-row"><span>Acompte versé</span><span>- {acompte:.2f} €</span></div>
<div class="tarif-row total"><span>Reste à payer</span><span>{reste:.2f} €</span></div>
</div>
{comment_section}
</div>
<div class="qr-section">
<img src="{qr_url_val}" alt="QR Code">
<p style="color:#64748b;font-size:11px;margin-top:8px;">Scannez pour suivre votre réparation</p>
</div>
<div class="footer">
<p>Date de dépôt: {fmt_date(t.get('date_depot',''))}</p>
<p style="margin-top:8px;">Merci de votre confiance !</p>
</div>
</div>
</body>
</html>"""
    
    # Version IMPRESSION - 80mm - NOIR SUR BLANC
    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <title>Ticket Client - {t['ticket_code']}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: Arial, sans-serif; background: #fff; padding: 10px; color: #000; }}
        .ticket {{ background: #fff; width: 80mm; max-width: 80mm; margin: 0 auto; border: 2px solid #000; }}
        .header {{ text-align: center; border-bottom: 2px solid #000; padding: 8px; }}
        .header img {{ width: 14mm; height: 14mm; margin-bottom: 2mm; }}
        .header h1 {{ font-size: 14px; font-weight: 900; margin: 0; color: #000; }}
        .header .contact {{ font-size: 9px; color: #333; margin-top: 2px; }}
        .header .phone {{ font-size: 11px; font-weight: 800; color: #000; }}
        .ticket-num {{ background: #000; color: #fff; text-align: center; font-size: 14px; font-weight: 900; padding: 8px; letter-spacing: 1px; }}
        .section {{ margin: 6px 8px; padding-bottom: 6px; border-bottom: 1px dashed #999; }}
        .section:last-child {{ border-bottom: none; }}
        .section-title {{ font-size: 10px; font-weight: 800; color: #000; text-transform: uppercase; margin-bottom: 4px; border-bottom: 1px solid #000; padding-bottom: 2px; }}
        .info-row {{ display: flex; justify-content: space-between; margin: 3px 0; font-size: 11px; color: #000; }}
        .info-row .label {{ color: #333; }}
        .info-row .value {{ font-weight: 700; color: #000; text-align: right; max-width: 55%; }}
        .big-name {{ font-size: 13px; font-weight: 800; margin-bottom: 2px; color: #000; }}
        .tarif-box {{ background: #f5f5f5; border: 2px solid #000; padding: 8px; margin: 6px 8px; }}
        .tarif-row {{ display: flex; justify-content: space-between; margin: 3px 0; font-size: 11px; color: #000; }}
        .tarif-row.total {{ font-size: 14px; font-weight: 900; border-top: 2px solid #000; padding-top: 6px; margin-top: 6px; color: #000; }}
        .qr-box {{ text-align: center; padding: 8px; margin: 6px 8px; background: #f9f9f9; border: 1px solid #ccc; }}
        .qr-box img {{ width: 22mm; height: 22mm; }}
        .qr-box p {{ font-size: 8px; margin-top: 2px; color: #333; }}
        .notes-box {{ background: #f9f9f9; border-left: 3px solid #000; padding: 6px; margin: 6px 8px; font-size: 9px; color: #000; }}
        .notes-title {{ font-weight: 800; font-size: 8px; margin-bottom: 2px; text-transform: uppercase; }}
        .disclaimer {{ font-size: 7px; color: #333; border-top: 1px dashed #999; padding: 6px 8px; margin-top: 4px; line-height: 1.3; }}
        .footer {{ text-align: center; font-weight: 800; font-size: 11px; padding: 8px; border-top: 1px solid #000; color: #000; }}
        .print-btn {{ display: block; width: calc(100% - 16px); margin: 8px; background: #000; color: #fff; padding: 10px; border: none; font-size: 12px; font-weight: 700; cursor: pointer; }}
        @media print {{
            body {{ padding: 0; }}
            .ticket {{ border: 1px solid #000; }}
            .print-btn {{ display: none !important; }}
        }}
    </style>
</head>
<body>
<div class="ticket">
    <div class="header">
        <img src="data:image/png;base64,{LOGO_B64}" alt="Logo">
        <h1>KLIKPHONE</h1>
        <div class="contact">
            <div class="phone">04 79 60 89 22</div>
            <div>79 Pl. Saint Léger, 73000 Chambéry</div>
        </div>
    </div>
    <div class="ticket-num">N° {t['ticket_code']}</div>
    <div class="section">
        <div class="section-title">👤 Client</div>
        <div class="big-name">{t.get('client_nom','')} {t.get('client_prenom','')}</div>
        <div class="info-row">
            <span class="label">Téléphone</span>
            <span class="value">{t.get('client_tel','')}</span>
        </div>
    </div>
    <div class="section">
        <div class="section-title">📱 Appareil</div>
        <div class="big-name">{t.get('marque','')} {modele_txt}</div>
        <div class="info-row">
            <span class="label">Motif</span>
            <span class="value">{panne}</span>
        </div>
        {type_ecran_html}
        <div class="info-row">
            <span class="label">Date dépôt</span>
            <span class="value">{fmt_date(t.get('date_depot',''))}</span>
        </div>
    </div>
    <div class="tarif-box">
        <div class="tarif-row"><span>Devis estimé</span><span>{devis:.2f} €</span></div>
        {rep_supp_html}
        <div class="tarif-row"><span>Acompte versé</span><span>- {acompte:.2f} €</span></div>
        <div class="tarif-row total"><span>RESTE À PAYER</span><span>{reste:.2f} €</span></div>
    </div>
    {comment_html}
    <div class="qr-box">
        <img src="{qr_url_val}" alt="QR Code">
        <p>Scannez pour suivre votre réparation</p>
    </div>
    <div class="disclaimer">
        • Klikphone ne consulte pas vos données personnelles<br>
        • Pensez à sauvegarder vos données avant dépôt<br>
        • Garantie 3 mois sur les réparations
    </div>
    <div class="footer">Merci de votre confiance !</div>
    <button class="print-btn" onclick="window.print()">🖨️ IMPRIMER</button>
</div>
</body>
</html>"""


def ticket_staff_html(t):
    """Génère un ticket STAFF HTML pour impression 80mm - NOIR SUR BLANC"""
    modele = t.get('modele_autre') or f"{t.get('modele', '')}"
    panne = t.get('panne_detail') if t.get('panne_detail') else t.get('panne', '')
    type_ecran = (t.get('type_ecran') or '').strip()
    
    # Label dynamique selon la panne
    panne_base = t.get("panne", "")
    if panne_base == "Écran casse":
        precision_label = "Type écran"
    else:
        precision_label = "Précision"
    
    notes_publiques = (t.get('commentaire_client') or '').strip()
    notes_privees = (t.get('notes_internes') or '').strip()
    notes_client = (t.get('notes_client') or '').strip()
    
    # Calcul reste à payer AVEC réparation supplémentaire
    devis = t.get('devis_estime') or 0
    acompte = t.get('acompte') or 0
    rep_supp = t.get('reparation_supp') or ''
    prix_supp = t.get('prix_supp') or 0
    tarif_final = t.get('tarif_final') or (devis + prix_supp)
    reste = max(0, tarif_final - acompte)
    
    # Ligne réparation supplémentaire
    rep_supp_html = f'<div class="tarif-row"><span>{rep_supp}</span><span>{prix_supp:.2f} €</span></div>' if rep_supp and prix_supp else ""
    
    # Précision dans section appareil
    type_ecran_html = f'<div class="info-row"><span class="label">{precision_label}</span><span class="value">{type_ecran}</span></div>' if type_ecran else ""
    
    # Notes
    notes_client_html = f'<div class="notes-box"><div class="notes-title">📋 Note client</div>{notes_client}</div>' if notes_client else ""
    notes_pub_html = f'<div class="notes-box"><div class="notes-title">💬 Note publique</div>{notes_publiques}</div>' if notes_publiques else ""
    notes_priv_html = f'<div class="notes-box private"><div class="notes-title">🔒 Note privée</div>{notes_privees}</div>' if notes_privees else ""
    
    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <title>Ticket Staff - {t['ticket_code']}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: Arial, sans-serif; background: #fff; padding: 10px; color: #000; }}
        .ticket {{ background: #fff; width: 80mm; max-width: 80mm; margin: 0 auto; border: 2px solid #000; }}
        .header {{ background: #f5f5f5; border-bottom: 2px solid #000; padding: 10px; text-align: center; }}
        .header img {{ width: 35px; height: 35px; margin-bottom: 4px; }}
        .header h1 {{ font-size: 14px; font-weight: 900; margin: 0; color: #000; }}
        .header p {{ font-size: 9px; color: #333; margin-top: 2px; }}
        .ticket-num {{ background: #000; color: #fff; text-align: center; padding: 8px; font-size: 16px; font-weight: 900; letter-spacing: 2px; }}
        .status {{ background: #f5f5f5; border-bottom: 1px solid #000; text-align: center; padding: 6px; font-size: 11px; font-weight: 700; color: #000; }}
        .content {{ padding: 8px; }}
        .section {{ margin-bottom: 8px; padding-bottom: 8px; border-bottom: 1px dashed #999; }}
        .section:last-child {{ border-bottom: none; }}
        .section-title {{ font-weight: 800; color: #000; font-size: 10px; text-transform: uppercase; margin-bottom: 4px; border-bottom: 1px solid #000; padding-bottom: 2px; }}
        .info-row {{ display: flex; justify-content: space-between; margin: 3px 0; font-size: 11px; color: #000; }}
        .info-row .label {{ color: #333; }}
        .info-row .value {{ font-weight: 700; color: #000; text-align: right; max-width: 55%; }}
        .big-name {{ font-size: 13px; font-weight: 800; margin-bottom: 3px; color: #000; }}
        .security-box {{ background: #fff; border: 3px solid #000; padding: 8px; margin: 8px 0; text-align: center; }}
        .security-box .title {{ font-weight: 900; font-size: 10px; color: #000; margin-bottom: 4px; }}
        .security-box .codes {{ font-size: 14px; font-weight: 900; font-family: monospace; color: #000; }}
        .tarif-box {{ background: #f5f5f5; border: 2px solid #000; padding: 10px; margin: 8px 0; }}
        .tarif-row {{ display: flex; justify-content: space-between; margin: 3px 0; font-size: 11px; color: #000; }}
        .tarif-row.total {{ font-size: 16px; font-weight: 900; border-top: 2px solid #000; padding-top: 6px; margin-top: 6px; color: #000; }}
        .notes-box {{ background: #f9f9f9; border-left: 3px solid #000; padding: 6px; margin: 5px 0; font-size: 9px; color: #000; }}
        .notes-box.private {{ border-color: #666; background: #f0f0f0; }}
        .notes-title {{ font-weight: 800; font-size: 8px; margin-bottom: 3px; text-transform: uppercase; color: #000; }}
        .footer {{ background: #f5f5f5; padding: 8px; text-align: center; font-size: 10px; color: #000; border-top: 1px solid #000; }}
        .print-btn {{ display: block; width: calc(100% - 16px); margin: 8px; background: #000; color: #fff; padding: 10px; border: none; font-size: 12px; font-weight: 700; cursor: pointer; }}
        @media print {{
            body {{ padding: 0; }}
            .ticket {{ border: 1px solid #000; }}
            .print-btn {{ display: none !important; }}
        }}
    </style>
</head>
<body>
<div class="ticket">
    <div class="header">
        <img src="data:image/png;base64,{LOGO_B64}" alt="Logo">
        <h1>TICKET STAFF</h1>
        <p>Document interne - Ne pas remettre au client</p>
    </div>
    <div class="ticket-num">{t['ticket_code']}</div>
    <div class="status">📍 {t.get('statut','En attente')}</div>
    <div class="content">
        <div class="section">
            <div class="section-title">👤 Client</div>
            <div class="big-name">{t.get('client_nom','')} {t.get('client_prenom','')}</div>
            <div class="info-row">
                <span class="label">Téléphone</span>
                <span class="value" style="font-family:monospace;">{t.get('client_tel','')}</span>
            </div>
        </div>
        <div class="section">
            <div class="section-title">📱 Appareil</div>
            <div class="big-name">{t.get('marque','')} {modele}</div>
            <div class="info-row">
                <span class="label">Réparation</span>
                <span class="value">{panne}</span>
            </div>
            {type_ecran_html}
        </div>
        <div class="security-box">
            <div class="title">🔐 CODES DE SÉCURITÉ</div>
            <div class="codes">PIN: {t.get('pin') or '----'} | Schéma: {t.get('pattern') or '----'}</div>
        </div>
        <div class="tarif-box">
            <div class="tarif-row"><span>Devis initial</span><span>{devis:.2f} €</span></div>
            {rep_supp_html}
            <div class="tarif-row"><span>Acompte versé</span><span>- {acompte:.2f} €</span></div>
            <div class="tarif-row total"><span>RESTE À PAYER</span><span>{reste:.2f} €</span></div>
        </div>
        <div class="section">
            <div class="section-title">📅 Dates</div>
            <div class="info-row"><span class="label">Dépôt</span><span class="value">{fmt_date(t.get('date_depot',''))}</span></div>
            <div class="info-row"><span class="label">Récupération</span><span class="value">{t.get('date_recuperation') or 'À définir'}</span></div>
        </div>
        {notes_client_html}
        {notes_pub_html}
        {notes_priv_html}
    </div>
    <div class="footer"><strong>Technicien:</strong> {t.get('technicien_assigne') or 'Non assigné'}</div>
    <button class="print-btn" onclick="window.print()">🖨️ IMPRIMER</button>
</div>
</body>
</html>"""


def ticket_devis_facture_html(t, doc_type="devis", for_email=False):
    """Génère un ticket DEVIS ou RÉCAPITULATIF - version impression ou email"""
    modele_txt = t.get("modele", "")
    if t.get("modele_autre"): modele_txt += f" ({t['modele_autre']})"
    
    panne = t.get('panne', '')
    if panne == "Autre" and t.get('panne_detail'):
        panne = t.get('panne_detail')
    elif t.get('panne_detail'):
        panne += f" ({t['panne_detail']})"
    

    # Commentaire public (imprimé sur le document)
    comment_public = (t.get('commentaire_client') or '').strip()
    comment_section_doc = f"""<div class="section"><div class="section-title">Commentaire</div><p style="margin:0;">{comment_public}</p></div>""" if comment_public else ""

    # Tarifs
    devis_val = t.get('devis_estime') or 0
    acompte_val = t.get('acompte') or 0
    rep_supp = t.get('reparation_supp') or ""
    prix_supp = t.get('prix_supp') or 0
    
    # Calculs TVA (prix TTC)
    total_ttc = devis_val + prix_supp
    total_ht = total_ttc / 1.20
    tva = total_ttc - total_ht
    reste = max(0, total_ttc - acompte_val)
    
    # Type de document
    is_facture = doc_type == "facture"
    doc_title = "REÇU" if is_facture else "DEVIS"
    doc_num = f"R-{t['ticket_code']}" if is_facture else f"D-{t['ticket_code']}"
    
    from datetime import datetime
    date_doc = datetime.now().strftime("%d/%m/%Y")
    
    # VERSION EMAIL (colorée, design)
    if for_email:
        doc_color = "#16a34a" if is_facture else "#3b82f6"
        rep_supp_line = ""
        if rep_supp:
            rep_supp_line = f"<tr><td style='padding:10px;border-bottom:1px solid #e5e7eb;'>Réparation supp.<br><small>{rep_supp}</small></td><td style='padding:10px;text-align:right;border-bottom:1px solid #e5e7eb;'>{prix_supp:.2f} €</td></tr>"
        
        return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
body {{ font-family: Arial, sans-serif; font-size: 14px; margin: 0; padding: 20px; background: #f5f5f5; }}
.document {{ max-width: 450px; margin: 0 auto; background: #fff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.1); }}
.header {{ background: linear-gradient(135deg, {doc_color}, {doc_color}dd); color: white; padding: 25px; text-align: center; }}
.header h1 {{ margin: 0; font-size: 28px; font-weight: 800; }}
.header .doc-num {{ font-size: 16px; opacity: 0.9; margin-top: 8px; }}
.header .date {{ font-size: 12px; opacity: 0.8; margin-top: 5px; }}
.company-info {{ background: #f8f9fa; padding: 15px; text-align: center; font-size: 12px; color: #666; border-bottom: 1px solid #e5e7eb; }}
.section {{ padding: 15px 20px; border-bottom: 1px solid #e5e7eb; }}
.section-title {{ font-size: 11px; text-transform: uppercase; color: #999; margin-bottom: 8px; letter-spacing: 1px; }}
.items-table {{ width: 100%; border-collapse: collapse; }}
.items-table th {{ background: #f3f4f6; padding: 10px; text-align: left; font-size: 11px; text-transform: uppercase; color: #666; }}
.items-table th:last-child {{ text-align: right; }}
.totals {{ background: #f8f9fa; padding: 15px 20px; }}
.total-line {{ display: flex; justify-content: space-between; padding: 5px 0; font-size: 13px; }}
.total-line.main {{ font-size: 18px; font-weight: 700; border-top: 2px solid {doc_color}; padding-top: 12px; margin-top: 8px; }}
.total-line.reste {{ font-size: 20px; font-weight: 800; color: #dc2626; border-top: 2px dashed #dc2626; padding-top: 12px; margin-top: 10px; }}
.footer {{ background: #1e293b; color: white; padding: 15px; text-align: center; font-size: 11px; }}
</style>
</head>
<body>
<div class="document">
<div class="header">
<h1>{doc_title}</h1>
<div class="doc-num">{doc_num}</div>
<div class="date">Date: {date_doc}</div>
</div>
<div class="company-info">
<strong>KLIKPHONE</strong> - Spécialiste Apple<br>
79 Place Saint Léger, 73000 Chambéry | Tél: 04 79 60 89 22
</div>
<div class="section">
<div class="section-title">Client</div>
<p><strong>{t.get('client_nom','')} {t.get('client_prenom','')}</strong></p>
<p>Tél: {t.get('client_tel','')}</p>
</div>
<div class="section">
<div class="section-title">Appareil</div>
<p><strong>{t.get('marque','')} {modele_txt}</strong></p>
</div>
<div class="section">
<div class="section-title">Prestations</div>
<table class="items-table">
<thead><tr><th>Description</th><th>Prix TTC</th></tr></thead>
<tbody>
<tr><td style="padding:10px;border-bottom:1px solid #e5e7eb;">{panne}</td><td style="padding:10px;text-align:right;border-bottom:1px solid #e5e7eb;">{devis_val:.2f} €</td></tr>
{rep_supp_line}
</tbody>
</table>
</div>
{comment_section_doc}
<div class="totals">
<div class="total-line main"><span>Total TTC</span><span>{total_ttc:.2f} €</span></div>
<div class="total-line" style="color:#666;font-size:11px;"><span>dont HT: {total_ht:.2f} € | TVA: {tva:.2f} €</span></div>
<div class="total-line"><span>Acompte versé</span><span style="color:#16a34a;">- {acompte_val:.2f} €</span></div>
<div class="total-line reste"><span>RESTE À PAYER</span><span>{reste:.2f} €</span></div>
</div>
<div class="footer">
<p>{"Ce devis est valable 30 jours." if not is_facture else "Merci pour votre confiance !"}</p>
<p style="margin-top:5px;color:#fbbf24;">{"" if not is_facture else "⚠️ Ce ticket ne fait pas office de facture."}</p>
</div>
</div>
</body>
</html>"""
    
    # VERSION IMPRESSION - 80mm x 200mm
    rep_supp_line = f'<div class="info-line"><strong>+ {rep_supp}:</strong> {prix_supp:.2f} €</div>' if rep_supp else ""
    
    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{doc_title} - Klikphone</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Inter', sans-serif;
            background-color: #f0f2f5;
            margin: 0;
            padding: 1rem;
            display: flex;
            justify-content: center;
        }}
        .ticket {{
            background: #fff;
            width: 80mm;
            min-height: 150mm;
            max-width: 80mm;
            padding: 4mm;
            border: 1px solid #ccc;
            border-radius: 4px;
            font-size: 11px;
            line-height: 1.4;
            color: #000;
        }}
        .logo-center {{
            text-align: center;
            margin-bottom: 3mm;
        }}
        .logo-center img {{
            width: 18mm;
            height: 18mm;
        }}
        h1 {{
            text-align: center;
            font-size: 16px;
            font-weight: 900;
            margin-bottom: 2mm;
            color: #000;
        }}
        .doc-info {{
            text-align: center;
            font-size: 10px;
            margin-bottom: 3mm;
            padding-bottom: 2mm;
            border-bottom: 1px dashed #999;
        }}
        .doc-num {{
            font-size: 12px;
            font-weight: 800;
        }}
        .contact {{
            text-align: center;
            font-size: 9px;
            margin-bottom: 3mm;
            color: #000;
        }}
        h2 {{
            font-size: 11px;
            font-weight: 800;
            border-bottom: 1px solid #000;
            padding-bottom: 1mm;
            margin: 3mm 0 2mm 0;
            text-align: center;
            color: #000;
        }}
        .info-line {{
            font-size: 10px;
            margin: 1.5mm 0;
            color: #000;
        }}
        .info-line strong {{
            font-weight: 700;
        }}
        .total-box {{
            border: 2px solid #000;
            padding: 3mm;
            margin: 4mm 0;
            text-align: center;
        }}
        .total-box .label {{
            font-size: 10px;
            margin-bottom: 1mm;
        }}
        .total-box .amount {{
            font-size: 16px;
            font-weight: 900;
        }}
        .disclaimer {{
            font-size: 7px;
            font-style: italic;
            border-top: 1px dashed #999;
            padding-top: 2mm;
            margin-top: 3mm;
            text-align: center;
            color: #333;
        }}
        .print-btn {{
            display: block;
            width: 100%;
            margin-top: 4mm;
            background: #000;
            color: #fff;
            padding: 3mm;
            border: none;
            border-radius: 2mm;
            font-size: 11px;
            font-weight: 700;
            cursor: pointer;
        }}
        .print-btn:hover {{ background: #333; }}

        @media print {{
            body {{
                background: #fff;
                padding: 0;
                margin: 0;
            }}
            .ticket {{
                border: none;
                border-radius: 0;
                padding: 3mm;
                min-height: auto;
            }}
            .print-btn {{
                display: none !important;
            }}
        }}
    </style>
</head>
<body>
<div class="ticket">
    <div class="logo-center">
        <img src="data:image/png;base64,{LOGO_B64}" alt="Logo">
    </div>
    <h1>{doc_title}</h1>
    <div class="doc-info">
        <div class="doc-num">{doc_num}</div>
        <div>Date: {date_doc}</div>
    </div>
    <div class="contact">
        KLIKPHONE - 79 Pl. Saint Léger, 73000 Chambéry<br>
        Tél: 04 79 60 89 22
    </div>

    <h2>Client</h2>
    <div class="info-line"><strong>Nom:</strong> {t.get('client_nom','')} {t.get('client_prenom','')}</div>
    <div class="info-line"><strong>Téléphone:</strong> {t.get('client_tel','')}</div>

    <h2>Appareil</h2>
    <div class="info-line"><strong>Modèle:</strong> {t.get('marque','')} {modele_txt}</div>

    <h2>Prestations</h2>
    <div class="info-line"><strong>{panne}:</strong> {devis_val:.2f} €</div>
    {rep_supp_line}

    <h2>Récapitulatif</h2>
    <div class="info-line"><strong>Total TTC:</strong> {total_ttc:.2f} €</div>
    <div class="info-line" style="font-size:8px;">HT: {total_ht:.2f} € | TVA (20%): {tva:.2f} €</div>
    <div class="info-line"><strong>Acompte versé:</strong> - {acompte_val:.2f} €</div>

    <div class="total-box">
        <div class="label">RESTE À PAYER</div>
        <div class="amount">{reste:.2f} €</div>
    </div>

    <div class="disclaimer">
        {"Ce devis est valable 30 jours. Les prix sont susceptibles de modification après diagnostic." if not is_facture else "Merci pour votre confiance ! Ce ticket ne fait pas office de facture."}
    </div>

    <button class="print-btn" onclick="window.print()">🖨️ IMPRIMER</button>
</div>
</body>
</html>"""

def ticket_combined_html(t):
    """Génère les deux tickets - 80mm x 200mm STRICT chacun"""
    panne = t.get("panne", "")
    if t.get("panne_detail"): panne += f" ({t['panne_detail']})"
    modele_txt = t.get("modele", "")
    if t.get("modele_autre"): modele_txt += f" ({t['modele_autre']})"
    
    # URL de suivi
    url_suivi = get_param("URL_SUIVI") or "https://klikphone-sav.streamlit.app"
    ticket_code = t.get('ticket_code', '')
    url_suivi_ticket = f"{url_suivi}?ticket={ticket_code}"
    qr_url_val = f"https://api.qrserver.com/v1/create-qr-code/?size=200x200&data={urllib.parse.quote(url_suivi_ticket)}"
    
    # Notes
    notes = t.get('notes_internes') or 'N/A'
    if len(notes) > 60: notes = notes[:60] + "..."
    
    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Tickets Klikphone</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Inter', sans-serif;
            background-color: #f0f2f5;
            margin: 0;
            padding: 1rem;
            display: flex;
            flex-direction: column;
            align-items: center;
        }}
        .print-btn {{
            display: block;
            width: 80mm;
            margin: 0 auto 1rem auto;
            background-color: #000;
            color: white;
            padding: 2mm;
            border-radius: 1mm;
            font-weight: 700;
            font-size: 10px;
            cursor: pointer;
            border: none;
            text-align: center;
        }}
        .print-btn:hover {{ background-color: #333; }}
        
        .ticket {{
            background: #fff;
            width: 80mm;
            max-width: 80mm;
            height: 200mm;
            max-height: 200mm;
            overflow: hidden;
            margin: 0 auto 1rem auto;
            padding: 3mm;
            border: 1px solid #ccc;
            border-radius: 2px;
            font-size: 9px;
            line-height: 1.3;
            color: #000;
        }}
        .logo-center {{
            text-align: center;
            margin-bottom: 1mm;
        }}
        .logo-center img {{
            width: 12mm;
            height: 12mm;
            display: block;
            margin: 0 auto;
        }}
        h1 {{
            text-align: center;
            font-size: 11px;
            font-weight: 900;
            margin-bottom: 1mm;
            color: #000;
        }}
        .contact {{
            text-align: center;
            font-size: 8px;
            margin-bottom: 2mm;
            color: #000;
        }}
        .contact .phone {{
            font-size: 10px;
            font-weight: 800;
        }}
        h2 {{
            font-size: 9px;
            font-weight: 800;
            border-bottom: 1px solid #000;
            padding-bottom: 0.5mm;
            margin: 2mm 0 1mm 0;
            text-align: center;
            color: #000;
        }}
        .info-line {{
            font-size: 8px;
            margin: 0.8mm 0;
            color: #000;
        }}
        .info-line strong {{
            font-weight: 700;
        }}
        .qr-box {{
            text-align: center;
            padding: 2mm 0;
            margin: 2mm 0;
            background: #f5f5f5;
            border-radius: 1mm;
        }}
        .qr-box img {{
            width: 28mm;
            height: 28mm;
            display: block;
            margin: 0 auto;
        }}
        .qr-box.big img {{
            width: 32mm;
            height: 32mm;
        }}
        .qr-box p {{
            font-size: 7px;
            margin-top: 1mm;
            font-weight: 600;
            color: #000;
            text-align: center;
        }}
        .ticket-num {{
            text-align: center;
            font-size: 11px;
            font-weight: 900;
            padding: 1.5mm;
            border: 2px solid #000;
            margin: 2mm 0;
            color: #000;
        }}
        .status {{
            text-align: center;
            font-weight: 700;
            font-size: 9px;
            margin-bottom: 2mm;
            color: #000;
        }}
        .security-box {{
            border: 2px solid #000;
            padding: 1.5mm;
            margin: 2mm 0;
            text-align: center;
            background: #f9f9f9;
        }}
        .security-box .title {{
            font-weight: 800;
            font-size: 8px;
            margin-bottom: 1mm;
        }}
        .security-box .codes {{
            font-size: 10px;
            font-weight: 900;
        }}
        .disclaimer {{
            font-size: 6px;
            font-style: italic;
            border-top: 1px dashed #999;
            padding-top: 1.5mm;
            margin-top: 2mm;
            line-height: 1.2;
            color: #333;
        }}
        .footer {{
            text-align: center;
            font-weight: 700;
            font-size: 9px;
            margin-top: 2mm;
            padding-top: 1.5mm;
            border-top: 1px solid #000;
            color: #000;
        }}
        .separator {{
            text-align: center;
            padding: 3mm 0;
            color: #999;
            font-size: 9px;
            width: 80mm;
        }}

        @media print {{
            @page {{
                size: 80mm 200mm;
                margin: 0;
            }}
            body {{ 
                background: #fff; 
                padding: 0; 
                margin: 0;
                width: 80mm;
            }}
            .print-btn {{ display: none !important; }}
            .separator {{ display: none !important; }}
            .ticket {{
                border: none;
                border-radius: 0;
                padding: 3mm;
                margin: 0;
                width: 80mm;
                height: 200mm;
                max-height: 200mm;
                overflow: hidden;
                page-break-after: always;
                page-break-inside: avoid;
            }}
            .ticket:last-child {{ page-break-after: auto; }}
        }}
    </style>
</head>
<body>

<button class="print-btn" onclick="window.print()">🖨️ IMPRIMER LES 2 TICKETS</button>

<!-- ========== TICKET CLIENT ========== -->
<div class="ticket">
    <div class="logo-center">
        <img src="data:image/png;base64,{LOGO_B64}" alt="Logo">
    </div>
    <h1>Ticket de Réparation Client</h1>
    <div class="contact">
        <p class="phone">04 79 60 89 22</p>
        <p>79 Pl. Saint Léger, 73000 Chambéry</p>
    </div>

    <h2>Client</h2>
    <div class="info-line"><strong>Nom:</strong> {t.get('client_nom','')} {t.get('client_prenom','')}</div>
    <div class="info-line"><strong>Tél:</strong> {t.get('client_tel','')}</div>

    <h2>Demande</h2>
    <div class="info-line"><strong>N°:</strong> {t['ticket_code']}</div>
    <div class="info-line"><strong>Appareil:</strong> {t.get('marque','')} {modele_txt}</div>
    <div class="info-line"><strong>Motif:</strong> {panne}</div>
    <div class="info-line"><strong>Date:</strong> {fmt_date(t.get('date_depot',''))}</div>

    <h2>Montant</h2>
    <div class="info-line"><strong>Devis:</strong> {(t.get('devis_estime') or 0):.2f} € | <strong>Acompte:</strong> {(t.get('acompte') or 0):.2f} €</div>

    <div class="qr-box">
        <img src="{qr_url_val}" alt="QR Code">
        <p>Scannez pour suivre votre réparation</p>
    </div>

    <div class="disclaimer">
        • Klikphone ne consulte pas vos données • Sauvegardez vos données avant dépôt • Aucune responsabilité post-réparation
    </div>

    <div class="footer">Merci de votre confiance !</div>
</div>

<div class="separator">✂️ ─────────── Découper ici ─────────── ✂️</div>

<!-- ========== TICKET STAFF ========== -->
<div class="ticket">
    <div class="logo-center">
        <img src="data:image/png;base64,{LOGO_B64}" alt="Logo">
    </div>
    <h1>Ticket Staff</h1>

    <div class="qr-box big">
        <img src="{qr_url_val}" alt="QR Code">
        <p>Scanner pour accéder au dossier</p>
    </div>

    <div class="ticket-num">N° {t['ticket_code']}</div>
    <div class="status">STATUT: {t.get('statut','')}</div>

    <h2>Client</h2>
    <div class="info-line"><strong>Nom:</strong> {t.get('client_nom','')} {t.get('client_prenom','')}</div>
    <div class="info-line"><strong>Tél:</strong> {t.get('client_tel','')}</div>

    <h2>Appareil</h2>
    <div class="info-line"><strong>Modèle:</strong> {t.get('marque','')} {modele_txt}</div>
    <div class="info-line"><strong>Panne:</strong> {panne}</div>

    <div class="security-box">
        <div class="title">🔐 CODES</div>
        <div class="codes">PIN: {t.get('pin') or '----'} | Schéma: {t.get('pattern') or '----'}</div>
    </div>

    <h2>Tarifs</h2>
    <div class="info-line"><strong>Devis:</strong> {fmt_prix(t.get('devis_estime'))} | <strong>Acompte:</strong> {fmt_prix(t.get('acompte'))} | <strong>Final:</strong> {fmt_prix(t.get('tarif_final'))}</div>

    <h2>Dates</h2>
    <div class="info-line"><strong>Dépôt:</strong> {fmt_date(t.get('date_depot',''))} | <strong>Récup:</strong> {fmt_date(t.get('date_recup',''))}</div>

    <h2>Notes</h2>
    <div class="info-line">{notes}</div>

    <div class="footer">Motif: {panne}</div>
</div>

</body>
</html>"""

# =============================================================================
# INTERFACE CLIENT - STYLE PORTAIL KLIKPHONE
# =============================================================================
def reset_client():
    st.session_state.step = 1
    st.session_state.data = {}
    st.session_state.pattern = []
    st.session_state.done = None
    if "success_timestamp" in st.session_state:
        del st.session_state.success_timestamp

def ui_client():
    import time
    
    if "step" not in st.session_state: reset_client()
    
    # Écran de succès PREMIUM
    if st.session_state.done:
        code = st.session_state.done
        t = get_ticket_full(code=code)
        url = get_param("URL_SUIVI") or "https://klikphone-sav.streamlit.app"
        
        # Initialiser le timestamp si pas encore fait
        if "success_timestamp" not in st.session_state:
            st.session_state.success_timestamp = time.time()
        
        # Calculer le temps ecoule
        elapsed = time.time() - st.session_state.success_timestamp
        remaining = max(0, 20 - int(elapsed))
        
        # Si 20 secondes ecoulees, reset automatique
        if remaining <= 0:
            reset_client()
            st.rerun()
        
        # Écran de succès
        st.markdown(f"""
        <div style="min-height:80vh;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:2rem;">
            <div style="width:100px;height:100px;border-radius:50%;background:linear-gradient(135deg,#22c55e 0%,#16a34a 100%);display:flex;align-items:center;justify-content:center;margin-bottom:1.5rem;box-shadow:0 15px 40px rgba(34,197,94,0.3);">
                <span style="font-size:50px;color:white;">✓</span>
            </div>
            <h1 style="font-size:2rem;font-weight:700;color:#1e293b;margin-bottom:0.5rem;">
                Demande enregistrée !
            </h1>
            <p style="font-size:1rem;color:#64748b;margin-bottom:1.5rem;">
                Votre appareil est entre de bonnes mains
            </p>
            <p style="font-size:0.85rem;color:#94a3b8;text-transform:uppercase;letter-spacing:2px;margin-bottom:0.5rem;">
                Votre numéro de ticket
            </p>
            <div style="display:inline-block;background:linear-gradient(135deg,#1e293b 0%,#334155 100%);color:white;font-family:monospace;font-size:1.8rem;font-weight:700;padding:1rem 2rem;border-radius:12px;letter-spacing:2px;margin-bottom:1.5rem;">
                {code}
            </div>
            <div style="background:white;padding:1rem;border-radius:16px;box-shadow:0 4px 20px rgba(0,0,0,0.08);margin-bottom:1rem;">
                <img src="{qr_url(f'{url}?ticket={code}')}" style="width:120px;height:120px;"/>
                <p style="font-size:0.75rem;color:#94a3b8;margin-top:0.5rem;">Scannez pour suivre</p>
            </div>
            <div style="background:#fff7ed;border:1px solid #fed7aa;border-radius:12px;padding:1rem 1.5rem;max-width:350px;">
                <p style="color:#9a3412;font-size:0.85rem;margin:0;">
                    📱 Conservez ce numéro pour suivre votre réparation
                </p>
            </div>
            <p style="color:#94a3b8;font-size:0.85rem;margin-top:1.5rem;">
                ⏱️ Retour automatique dans <strong style="color:#1e293b;">{remaining}s</strong>
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Bouton prochain client
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("👤 PROCHAIN CLIENT", type="primary", use_container_width=True, key="next_client"):
                reset_client()
                st.rerun()
        
        # Bouton pour voir le ticket (remplace l'expander)
        if t:
            if "show_ticket_depot" not in st.session_state:
                st.session_state.show_ticket_depot = False
            
            if st.button("🎫 Voir le ticket de dépôt" + (" ▼" if st.session_state.show_ticket_depot else " ▶"), 
                        key="toggle_ticket_depot", use_container_width=True, type="secondary"):
                st.session_state.show_ticket_depot = not st.session_state.show_ticket_depot
                st.rerun()
            
            if st.session_state.show_ticket_depot:
                st.components.v1.html(ticket_client_html(t), height=750, scrolling=True)
        
        # Forcer un rerun pour le compteur
        time.sleep(1)
        st.rerun()
        
        return
    
    # === INTERFACE CLIENT PREMIUM ===
    step = st.session_state.step
    
    # Header avec logo
    st.markdown(f"""
    <div style="text-align:center;padding:2rem 1rem 1.5rem;background:linear-gradient(180deg,#ffffff 0%,rgba(249,115,22,0.03) 100%);">
        <div style="width:80px;height:80px;margin:0 auto 1rem;background:linear-gradient(135deg,#fff7ed 0%,#ffedd5 100%);border-radius:20px;display:flex;align-items:center;justify-content:center;box-shadow:0 8px 24px rgba(249,115,22,0.15);">
            <img src="data:image/png;base64,{LOGO_B64}" style="width:50px;height:50px;" alt="Klikphone">
        </div>
        <h1 style="font-size:2rem;font-weight:800;background:linear-gradient(135deg,#f97316 0%,#ea580c 100%);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:0.25rem;letter-spacing:-1px;">
            KLIKPHONE
        </h1>
        <p style="color:#64748b;font-size:0.95rem;margin-bottom:0.5rem;">
            Spécialiste Apple & Multimarque
        </p>
        <p style="color:#94a3b8;font-size:0.8rem;">
            📍 79 Place Saint Léger, Chambéry • 📞 04 79 60 89 22
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Barre de progression native Streamlit (plus fiable)
    st.progress(step / 6)
    
    # Indicateur d'étape
    step_names = ["", "Type d'appareil", "Marque", "Modèle", "Problème", "Sécurité", "Coordonnées"]
    st.markdown(f"""
    <p style="text-align:center;color:#64748b;font-size:0.9rem;margin-bottom:1.5rem;">
        Étape <strong style="color:#f97316;">{step}</strong>/6 : {step_names[step]}
    </p>
    """, unsafe_allow_html=True)
    
    if step == 1: client_step1()
    elif step == 2: client_step2()
    elif step == 3: client_step3()
    elif step == 4: client_step4()
    elif step == 5: client_step5()
    elif step == 6: client_step6()

def client_step1():
    """Étape 1: Choix du type d'appareil - Design simple"""
    
    # Titre
    st.markdown("""
    <div style="text-align: center; margin-bottom: 1.5rem;">
        <h2 style="font-size: 1.5rem; font-weight: 700; color: #1e293b; margin-bottom: 0.5rem;">
            Quel appareil déposez-vous ?
        </h2>
        <p style="font-size: 0.95rem; color: #64748b;">Sélectionnez le type d'appareil</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Wrapper avec classe CSS pour boutons uniformes (défini dans load_css)
    st.markdown('<div class="uniform-buttons">', unsafe_allow_html=True)
    
    # Grille 2x2 avec icônes modernes
    col1, col2 = st.columns(2)
    with col1:
        if st.button("📱  Smartphone", key="cat_smartphone", use_container_width=True):
            st.session_state.data["cat"] = "Smartphone"
            st.session_state.data["is_commande"] = False
            st.session_state.step = 2
            st.rerun()
    with col2:
        if st.button("🔲  Tablette", key="cat_tablette", use_container_width=True):
            st.session_state.data["cat"] = "Tablette"
            st.session_state.data["is_commande"] = False
            st.session_state.step = 2
            st.rerun()
    
    col3, col4 = st.columns(2)
    with col3:
        if st.button("💻  PC Portable", key="cat_pc", use_container_width=True):
            st.session_state.data["cat"] = "PC Portable"
            st.session_state.data["is_commande"] = False
            st.session_state.step = 2
            st.rerun()
    with col4:
        if st.button("🎮  Console", key="cat_console", use_container_width=True):
            st.session_state.data["cat"] = "Console"
            st.session_state.data["is_commande"] = False
            st.session_state.step = 2
            st.rerun()
    
    # Séparateur
    st.markdown("<div style='text-align: center; color: #94a3b8; margin: 1rem 0;'>─ ou ─</div>", unsafe_allow_html=True)
    
    # Options spéciales
    col5, col6 = st.columns(2)
    with col5:
        if st.button("📦  Commander pièce", key="cat_commande", use_container_width=True):
            st.session_state.data["cat"] = "Commande"
            st.session_state.data["is_commande"] = True
            st.session_state.step = 2
            st.rerun()
    with col6:
        if st.button("⚙️  Autre appareil", key="cat_autre", use_container_width=True):
            st.session_state.data["cat"] = "Autre"
            st.session_state.data["is_commande"] = False
            st.session_state.step = 4
            st.rerun()
    
    st.markdown('</div>', unsafe_allow_html=True)  # Ferme uniform-buttons

def client_step2():
    """Étape 2: Choix de la marque - Design Premium avec logos"""
    cat = st.session_state.data.get("cat", "")
    
    # Si c'est une commande, demander directement les infos
    if cat == "Commande":
        st.markdown("""
        <div class="step-title">
            <h2>📦 Que souhaitez-vous commander ?</h2>
            <p>Décrivez la pièce ou l'accessoire souhaité</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Bouton retour
        if st.button("← Retour", key="back2", type="secondary"):
            st.session_state.step = 1
            st.rerun()
        
        commande_detail = st.text_area(
            "Description de votre commande", 
            placeholder="Ex: Écran iPhone 12 Pro noir, Coque Samsung Galaxy S21, Chargeur MacBook Pro...",
            height=150, 
            key="commande_detail"
        )
        
        if st.button("Continuer →", type="primary", use_container_width=True):
            if commande_detail:
                st.session_state.data["marque"] = "Commande"
                st.session_state.data["modèle"] = "Commande"
                st.session_state.data["modele_autre"] = commande_detail
                st.session_state.data["panne"] = "Commande"
                st.session_state.data["panne_detail"] = commande_detail
                st.session_state.step = 5
                st.rerun()
            else:
                st.warning("⚠️ Veuillez décrire votre commande")
        return
    
    # Titre avec l'appareil choisi
    device_icons = {"Smartphone": "📱", "Tablette": "📟", "PC Portable": "💻", "Console": "🎮"}
    icon = device_icons.get(cat, "📱")
    
    st.markdown(f"""
    <div class="step-title" style="margin-bottom: 2rem;">
        <h2 style="font-size: 1.75rem; font-weight: 700; color: #1e293b; margin-bottom: 8px;">
            {icon} Quelle est la marque ?
        </h2>
        <p style="color: #64748b; font-size: 1rem;">Sélectionnez la marque de votre {cat.lower()}</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Bouton retour stylé
    if st.button("← Retour", key="back2", type="secondary"):
        st.session_state.step = 1
        st.rerun()
    
    marques = get_marques(cat)
    
    # Style premium: logo + texte alignés (sans hack overlay)
    def _slugify_brand(s: str) -> str:
        return re.sub(r'[^a-z0-9]+', '-', s.lower()).strip('-')

    brand_css = """<style>
    .brand-grid{
        max-width: 980px;
        margin: 0 auto;
        padding: 0 1rem;
    }
    /* Boutons marques — look SaaS moderne */
    .brand-grid .stButton>button{
        height: 64px !important;
        min-height: 64px !important;
        max-height: 64px !important;
        font-size: 16px !important;
        font-weight: 600 !important;
        border-radius: 14px !important;
        border: 1px solid rgba(226,232,240,1) !important;
        background: rgba(255,255,255,0.92) !important;
        box-shadow: 0 10px 24px rgba(16,24,40,0.06) !important;
        transition: transform .12s ease, box-shadow .12s ease, border-color .12s ease, background .12s ease !important;
        display: flex !important;
        align-items: center !important;
        justify-content: flex-start !important;
        padding: 0 18px !important;
        text-align: left !important;
    }
    .brand-grid .stButton>button:hover{
        transform: translateY(-1px) !important;
        box-shadow: 0 16px 36px rgba(16,24,40,0.10) !important;
        border-color: rgba(255,106,0,0.35) !important;
        background: rgba(255,255,255,1) !important;
    }
    .brand-grid .stButton>button:active{
        transform: translateY(0px) !important;
        box-shadow: 0 10px 24px rgba(16,24,40,0.08) !important;
    }
    /* Logo à gauche du texte (aligné comme page Modèles) */
    .brand-item button::before{
        content: "";
        width: 22px;
        height: 22px;
        margin-right: 12px;
        background-size: contain;
        background-repeat: no-repeat;
        background-position: center;
        display: inline-block;
        flex: 0 0 22px;
        border-radius: 6px;
    }
    .brand-autre button::before{
        content: "🔧";
        width: auto;
        height: auto;
        margin-right: 12px;
        background: none;
        display: inline-block;
        font-size: 18px;
        line-height: 1;
    }
    """

    # Règles par marque (SVG officiels via SimpleIcons) — injectées en CSS-only
    for _b, _url in BRAND_LOGOS.items():
        if not _url or _b == "Autre":
            continue
        brand_css += f".brand-{_slugify_brand(_b)} button::before{{background-image:url('{_url}');}}\n"

    brand_css += """</style>"""
    st.markdown(brand_css, unsafe_allow_html=True)

    st.markdown('<div class="brand-grid">', unsafe_allow_html=True)

    # Afficher les marques en grille 2 colonnes
    cols = st.columns(2)
    for i, m in enumerate(marques):
        with cols[i % 2]:
            # Wrapper pour cibler le bouton via CSS (logo à gauche)
            slug = _slugify_brand(m)
            wrapper_cls = f"brand-item brand-{slug}" if m != "Autre" else "brand-item brand-autre"
            st.markdown(f'<div class="{wrapper_cls}">', unsafe_allow_html=True)

            label = m if m != "Autre" else "Autre"
            if st.button(label, key=f"brand_{m}", use_container_width=True):
                st.session_state.data["marque"] = m
                st.session_state.step = 3
                st.rerun()

            st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

def client_step3():
    """Étape 3: Choix du modèle - Simplifié"""
    cat = st.session_state.data.get("cat", "")
    marque = st.session_state.data.get("marque", "")
    
    # Logo de la marque
    logo_url = BRAND_LOGOS.get(marque, "")
    
    # Titre simple avec logo
    if logo_url and marque != "Autre":
        st.markdown(f"""
        <div style="text-align: center; margin-bottom: 1.5rem;">
            <img src="{logo_url}" style="width: 32px; height: 32px; object-fit: contain; margin-bottom: 8px;">
            <h2 style="font-size: 1.3rem; font-weight: 700; color: #1e293b; margin: 0;">Quel modèle {marque} ?</h2>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div style="text-align: center; margin-bottom: 1.5rem;">
            <h2 style="font-size: 1.3rem; font-weight: 700; color: #1e293b; margin: 0;">🔧 Quel modèle ?</h2>
        </div>
        """, unsafe_allow_html=True)
    
    # Bouton retour
    if st.button("← Retour", key="back3", type="secondary"):
        st.session_state.step = 2
        st.rerun()
    
    # Si "Autre" marque, demander directement le modèle
    if marque == "Autre":
        modele_autre = st.text_input(
            "Précisez la marque et le modèle", 
            placeholder="Ex: Huawei P30 Pro, OnePlus 9...",
            key="input_modele_autre"
        )
        
        if st.button("Continuer →", type="primary", use_container_width=True):
            if modele_autre:
                st.session_state.data["modèle"] = "Autre"
                st.session_state.data["modele_autre"] = modele_autre
                st.session_state.step = 4
                st.rerun()
            else:
                st.warning("⚠️ Veuillez préciser le modèle")
    else:
        # Récupérer les modèles
        modeles_db = get_modeles(cat, marque)
        modeles_list = [m for m in modeles_db if m != "Autre"]
        modeles_list.append("Autre")
        
        # Liste déroulante simple
        mod = st.selectbox(
            "Sélectionnez votre modèle",
            ["-- Choisir le modèle --"] + modeles_list,
            key="select_modele"
        )
        
        # Si "Autre" est sélectionné
        modele_autre = ""
        if mod == "Autre":
            modele_autre = st.text_input(
                "Précisez le modèle",
                placeholder="Ex: iPhone 14 Pro Max",
                key="input_autre"
            )
        
        if st.button("Continuer →", type="primary", use_container_width=True):
            if mod == "-- Choisir le modèle --":
                st.warning("⚠️ Veuillez sélectionner un modèle")
            elif mod == "Autre" and not modele_autre:
                st.warning("⚠️ Veuillez préciser le modèle")
            else:
                st.session_state.data["modèle"] = mod
                st.session_state.data["modele_autre"] = modele_autre
                st.session_state.step = 4
                st.rerun()

def client_step4():
    """Étape 4: Description du problème"""
    cat = st.session_state.data.get("cat", "")
    
    # Si catégorie "Autre", demander d'abord l'appareil
    if cat == "Autre":
        st.markdown("""
        <div class="step-title">
            <h2>❓ Décrivez votre appareil</h2>
            <p>Quel appareil et quel problème ?</p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("← Retour", key="back4", type="secondary"):
            st.session_state.step = 1
            st.rerun()
        
        appareil = st.text_input(
            "Votre appareil",
            placeholder="Ex: Montre connectée Garmin, Drone DJI, Enceinte Bose..."
        )
        probleme = st.text_area(
            "Décrivez le problème",
            placeholder="Décrivez précisément le problème rencontré...",
            height=120
        )
        
        if st.button("Continuer →", type="primary", use_container_width=True):
            if appareil and probleme:
                st.session_state.data["marque"] = "Autre"
                st.session_state.data["modèle"] = "Autre"
                st.session_state.data["modele_autre"] = appareil
                st.session_state.data["panne"] = "Autre"
                st.session_state.data["panne_detail"] = probleme
                st.session_state.step = 5
                st.rerun()
            else:
                st.warning("⚠️ Veuillez remplir tous les champs")
        return
    
    # Récupérer les infos de l'appareil
    marque = st.session_state.data.get("marque", "")
    modele = st.session_state.data.get("modèle", "")
    modele_autre = st.session_state.data.get("modele_autre", "")
    appareil_txt = modele_autre if modele_autre else f"{marque} {modele}"
    
    st.markdown(f"""
    <div class="step-title">
        <h2>🔧 Quel est le problème ?</h2>
        <p>Sélectionnez le problème rencontré sur votre {appareil_txt}</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("← Retour", key="back4", type="secondary"):
        st.session_state.step = 3
        st.rerun()
    
    # Icônes pour chaque panne
    panne_icons = {
        "Écran casse": "📱💔",
        "Batterie": "🔋",
        "Connecteur de charge": "🔌",
        "Camera avant": "🤳",
        "Camera arriere": "📸",
        "Bouton volume": "🔊",
        "Bouton power": "⏻",
        "Haut-parleur (je n'entends pas les gens ou la musique)": "🔈",
        "Microphone (les gens ne m'entendent pas)": "🎤",
        "Vitre arriere": "🪟",
        "Désoxydation": "💧",
        "Problème logiciel": "⚙️",
        "Diagnostic": "🔍",
        "Autre": "❓"
    }
    
    # Afficher les pannes en liste
    for p in PANNES:
        icon = panne_icons.get(p, "🔧")
        # Raccourcir le texte pour l'affichage
        display_text = p.split("(")[0].strip() if "(" in p else p
        
        if st.button(f"{icon}  {display_text}", key=f"panne_{p}", use_container_width=True):
            st.session_state.data["panne"] = p
            if p == "Autre" or p == "Diagnostic":
                st.session_state.data["show_detail"] = True
                st.rerun()
            else:
                st.session_state.step = 5
                st.rerun()
    
    # Afficher zone de détail si nécessaire
    if st.session_state.data.get("show_detail"):
        st.markdown("---")
        detail = st.text_area(
            "Décrivez le problème en détail",
            placeholder="Expliquez précisément ce qui ne fonctionne pas...",
            height=100
        )
        
        if st.button("Continuer →", type="primary", use_container_width=True):
            st.session_state.data["panne_detail"] = detail
            st.session_state.step = 5
            st.rerun()

def client_step5():
    """Étape 5: Code de déverrouillage"""
    st.markdown("""
    <div class="step-title">
        <h2>🔐 Code de déverrouillage</h2>
        <p>Pour permettre au technicien de tester votre appareil</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("← Retour", key="back5", type="secondary"):
        st.session_state.step = 4
        st.rerun()
    
    st.markdown("""
    <div style="background:#fef3c7;border:1px solid #fcd34d;border-radius:12px;padding:1rem;margin-bottom:1.5rem;">
        <p style="margin:0;color:#92400e;font-size:0.9rem;">
            🔒 <strong>Confidentialité garantie</strong><br>
            Votre code ne sera utilisé que pour tester l'appareil après réparation.
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Options de sécurité
    st.markdown("**Comment votre appareil est-il verrouillé ?**")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        pin_btn = st.button("🔢\n\nCode PIN", key="sec_pin", use_container_width=True)
    with col2:
        schema_btn = st.button("⬡\n\nSchéma", key="sec_schema", use_container_width=True)
    with col3:
        none_btn = st.button("🔓\n\nAucun", key="sec_none", use_container_width=True)
    
    # Gestion du choix
    if "security_choice" not in st.session_state:
        st.session_state.security_choice = None
    
    if pin_btn:
        st.session_state.security_choice = "pin"
        st.rerun()
    elif schema_btn:
        st.session_state.security_choice = "schema"
        st.rerun()
    elif none_btn:
        st.session_state.step = 6
        st.rerun()
    
    # Afficher l'interface selon le choix
    if st.session_state.security_choice == "pin":
        st.markdown("---")
        pin = st.text_input(
            "Entrez votre code PIN",
            type="password",
            placeholder="••••••",
            max_chars=10
        )
        
        if st.button("Continuer →", type="primary", use_container_width=True):
            st.session_state.data["pin"] = pin
            st.session_state.step = 6
            st.rerun()
    
    elif st.session_state.security_choice == "schema":
        st.markdown("---")
        st.markdown("**Cliquez sur les points dans l'ordre de votre schéma :**")
        
        if "pattern" not in st.session_state:
            st.session_state.pattern = []
        
        # Grille 3x3 pour le schéma
        cols = st.columns([1, 2, 1])
        with cols[1]:
            for row in range(3):
                row_cols = st.columns(3)
                for col in range(3):
                    n = row * 3 + col + 1
                    with row_cols[col]:
                        if n in st.session_state.pattern:
                            pos = st.session_state.pattern.index(n) + 1
                            st.button(f"●{pos}", key=f"pt_{n}", disabled=True, use_container_width=True)
                        else:
                            if st.button("○", key=f"pt_{n}", use_container_width=True):
                                st.session_state.pattern.append(n)
                                st.rerun()
        
        if st.session_state.pattern:
            st.markdown(f"""
            <div style="text-align:center;margin:1rem 0;padding:0.75rem;background:#f1f5f9;border-radius:8px;">
                <strong>Séquence :</strong> {' → '.join(map(str, st.session_state.pattern))}
            </div>
            """, unsafe_allow_html=True)
            
            c1, c2 = st.columns(2)
            with c1:
                if st.button("🗑️ Effacer", use_container_width=True, type="secondary"):
                    st.session_state.pattern = []
                    st.rerun()
            with c2:
                if st.button("Valider →", type="primary", use_container_width=True):
                    st.session_state.data["pattern"] = "-".join(map(str, st.session_state.pattern))
                    st.session_state.step = 6
                    st.rerun()

def client_step6():
    """Étape 6: Coordonnées du client"""
    import time
    
    # Vérifier si un client existe déjà (popup avec retour automatique)
    if st.session_state.get("client_exists_popup"):
        client_info = st.session_state.client_exists_popup
        
        # Timer pour retour automatique
        if "client_exists_timestamp" not in st.session_state:
            st.session_state.client_exists_timestamp = time.time()
        
        elapsed = time.time() - st.session_state.client_exists_timestamp
        remaining = max(0, 10 - int(elapsed))
        
        # Si 10 secondes écoulées, reset automatique
        if remaining <= 0:
            st.session_state.client_exists_popup = None
            st.session_state.client_exists_timestamp = None
            reset_client()
            st.rerun()
        
        # Affichage du popup
        st.markdown(f"""
        <div style="min-height:70vh;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:2rem;">
            <div style="background:linear-gradient(135deg,#fef3c7 0%,#fde68a 100%);border:3px solid #f59e0b;border-radius:24px;padding:3rem 2rem;max-width:500px;box-shadow:0 20px 60px rgba(245,158,11,0.2);">
                <div style="font-size:5rem;margin-bottom:1rem;">👋</div>
                <h1 style="color:#92400e;margin-bottom:1rem;font-size:2rem;font-weight:700;">
                    Vous êtes déjà client !
                </h1>
                <p style="color:#78350f;font-size:1.2rem;margin-bottom:2rem;line-height:1.6;">
                    Bonjour <strong>{client_info.get('prenom', '')} {client_info.get('nom', '')}</strong>,<br>
                    votre numéro est déjà enregistré.
                </p>
                <div style="background:white;border-radius:16px;padding:1.5rem;margin-bottom:2rem;">
                    <p style="color:#1e293b;font-size:1.1rem;margin:0;">
                        📍 <strong>Merci de vous diriger vers l'accueil</strong><br>
                        <span style="color:#64748b;">Un conseiller va s'occuper de vous</span>
                    </p>
                </div>
                <p style="color:#92400e;font-size:0.9rem;">
                    Retour automatique dans <strong style="font-size:1.2rem;">{remaining}</strong> secondes
                </p>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Bouton pour nouveau client
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("🏠 NOUVEAU CLIENT", type="primary", use_container_width=True):
                st.session_state.client_exists_popup = None
                st.session_state.client_exists_timestamp = None
                reset_client()
                st.rerun()
        
        # Forcer rerun pour le compteur
        time.sleep(1)
        st.rerun()
        return
    
    st.markdown("""
    <div style="text-align:center;margin-bottom:1.5rem;">
        <h2 style="font-size:1.5rem;font-weight:700;color:#1e293b;">👤 Vos coordonnées</h2>
        <p style="color:#64748b;">Pour vous contacter quand votre appareil sera prêt</p>
    </div>
    """, unsafe_allow_html=True)
    
    if st.button("← Retour", key="back6", type="secondary"):
        st.session_state.step = 5
        st.rerun()
    
    col1, col2 = st.columns(2)
    with col1:
        prenom = st.text_input("Prénom *", placeholder="Jean")
        telephone = st.text_input("Téléphone *", placeholder="06 12 34 56 78")
    with col2:
        nom = st.text_input("Nom *", placeholder="Dupont")
        email = st.text_input("Email", placeholder="jean.dupont@email.com")
    
    societe = st.text_input(
        "Société (facultatif)",
        placeholder="Nom de l'entreprise si professionnel"
    )
    
    notes = st.text_area(
        "Remarques",
        placeholder="Accessoires laissés, précisions sur le problème, disponibilités...",
        height=80
    )
    
    st.markdown("---")
    
    # Options supplémentaires
    col_opt1, col_opt2 = st.columns(2)
    with col_opt1:
        carte_camby = st.checkbox(
            "🎫 J'ai la carte Camby",
            help="Cochez si vous possédez la carte de fidélité Camby"
        )
    with col_opt2:
        # Si c'est une commande pièce depuis le menu, pré-cocher
        is_commande = st.session_state.data.get("is_commande", False)
        commande_piece = st.checkbox(
            "⚙️ Pièce à commander",
            value=is_commande,  # Pré-coché si vient de "Commander pièce"
            help="Cochez si une pièce doit être commandée"
        )
    
    st.markdown("---")
    
    # CGV
    st.markdown("""
    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:1rem;font-size:0.8rem;color:#64748b;max-height:150px;overflow-y:auto;margin-bottom:1rem;">
        <strong>CONDITIONS GÉNÉRALES</strong><br>
        • Klikphone ne consulte pas les données de votre appareil.<br>
        • Pensez à sauvegarder avant dépôt.<br>
        • Garantie 3 mois sur les réparations.<br>
        • Récupération sous 30 jours après notification.
    </div>
    """, unsafe_allow_html=True)
    
    consent = st.checkbox("✅ J'accepte les conditions générales")
    
    # Bouton d'envoi
    if st.button("🚀 ENVOYER MA DEMANDE", type="primary", use_container_width=True):
        if not nom or not prenom or not telephone:
            st.error("❌ Le nom, prénom et téléphone sont obligatoires")
        elif not consent:
            st.error("❌ Veuillez accepter les conditions générales")
        else:
            # Vérifier si le client existe déjà
            existing_client = check_client_exists(telephone)
            if existing_client:
                # Afficher le popup
                st.session_state.client_exists_popup = existing_client
                st.rerun()
            else:
                # Nouveau client - créer le ticket
                d = st.session_state.data
                cid = get_or_create_client(nom, telephone, prenom, email, societe, 1 if carte_camby else 0)
                
                # Forcer commande_piece si is_commande
                final_commande_piece = commande_piece or is_commande
                
                code = creer_ticket(cid, d.get("cat",""), d.get("marque",""), d.get("modèle",""),
                                   d.get("modele_autre",""), d.get("panne",""), d.get("panne_detail",""),
                                   d.get("pin",""), d.get("pattern",""), notes, "", 1 if final_commande_piece else 0)
                
                # Si commande pièce, créer une entrée dans commandes_pieces
                if final_commande_piece:
                    t = get_ticket(code=code)
                    if t:
                        # Pour les commandes pièces, utiliser la description saisie
                        if is_commande and d.get('panne_detail'):
                            description_cmd = d.get('panne_detail')
                        else:
                            modele_txt = f"{d.get('marque','')} {d.get('modèle','')}"
                            if d.get('modele_autre'): modele_txt += f" ({d['modele_autre']})"
                            panne_txt = d.get('panne', '') or d.get('panne_detail', '') or 'Pièce à préciser'
                            description_cmd = f"Pièce pour {panne_txt} - {modele_txt}"
                        ajouter_commande_piece(t['id'], description_cmd, "A définir", "", 0, "Commande créée depuis totem client")
                
                st.session_state.done = code
                st.rerun()

# =============================================================================
# INTERFACE STAFF (ACCUEIL) - STYLE PORTAIL STAFF
# =============================================================================
def ui_accueil():
    # === HEADER NAV ===
    st.markdown(f"""
    <div class="nav-header">
        <div class="nav-logo">
            <img src="data:image/png;base64,{LOGO_B64}" style="width:36px;height:36px;">
            <span class="nav-logo-text">KLIKPHONE</span>
            <span style="color:var(--neutral-400);font-size:var(--text-sm);margin-left:8px;">SAV Manager</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Boutons navigation
    col_spacer, col_tech, col_logout = st.columns([6, 2, 2])
    with col_tech:
        if st.button("🔧 Tech", key="goto_tech", type="secondary", use_container_width=True):
            st.session_state.mode = "tech"
            st.rerun()
    with col_logout:
        if st.button("Sortir", key="logout_acc", type="secondary", use_container_width=True):
            st.session_state.mode = None
            st.session_state.auth = False
            st.rerun()
    
    # === KPI CARDS CLIQUABLES ===
    all_tickets = chercher_tickets()
    commandes_attente = get_commandes_pieces(statut="A commander")
    nb_total = len(all_tickets)
    nb_attente = len([t for t in all_tickets if t.get('statut') == "En attente de diagnostic"])
    nb_encours = len([t for t in all_tickets if t.get('statut') == "En cours de réparation"])
    nb_attente_piece = len([t for t in all_tickets if t.get('statut') == "En attente de pièce"])
    nb_attente_accord = len([t for t in all_tickets if t.get('statut') == "En attente d'accord client"])
    nb_commandes = len(commandes_attente)
    
    # Filtre actif
    filtre_actif = st.session_state.get("filtre_kpi", None)
    
    # KPI cliquables avec design amélioré
    col_k1, col_k2, col_k3, col_k4, col_k5 = st.columns(5)
    with col_k1:
        selected = filtre_actif is None
        if st.button(f"📋 **{nb_total}**\nTotal", key="kpi_total", use_container_width=True, type="primary" if selected else "secondary"):
            st.session_state.filtre_kpi = None
            st.rerun()
    with col_k2:
        selected = filtre_actif == "En attente de diagnostic"
        if st.button(f"⏳ **{nb_attente}**\nDiagnostic", key="kpi_attente", use_container_width=True, type="primary" if selected else "secondary"):
            st.session_state.filtre_kpi = "En attente de diagnostic"
            st.rerun()
    with col_k3:
        selected = filtre_actif == "En cours de réparation"
        if st.button(f"🔧 **{nb_encours}**\nRéparation", key="kpi_encours", use_container_width=True, type="primary" if selected else "secondary"):
            st.session_state.filtre_kpi = "En cours de réparation"
            st.rerun()
    with col_k4:
        selected = filtre_actif == "En attente d'accord client"
        color = "primary" if (selected or nb_attente_accord > 0) else "secondary"
        if st.button(f"⚠️ **{nb_attente_accord}**\nAccord", key="kpi_accord", use_container_width=True, type=color):
            st.session_state.filtre_kpi = "En attente d'accord client"
            st.rerun()
    with col_k5:
        selected = filtre_actif == "En attente de pièce"
        color = "primary" if (selected or nb_attente_piece > 0) else "secondary"
        if st.button(f"📦 **{nb_attente_piece}**\nPièces", key="kpi_pieces", use_container_width=True, type=color):
            st.session_state.filtre_kpi = "En attente de pièce"
            st.rerun()

# === TABS ===
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["📋 Demandes", "➕ Nouvelle", "👥 Clients", "📦 Commandes", "📄 Attestation", "⚙️ Config"])
    
    with tab1:
        staff_liste_demandes()
    with tab2:
        staff_nouvelle_demande()
    with tab3:
        staff_gestion_clients()
    with tab4:
        staff_commandes_pieces()
    with tab5:
        staff_attestation()
    with tab6:
        staff_config()
    
    # Footer
    st.markdown("""
    <div style="text-align:center;padding:20px 0;margin-top:40px;border-top:1px solid var(--neutral-200);color:var(--neutral-400);font-size:12px;">
        Créé par <strong>TkConcept26</strong>
    </div>
    """, unsafe_allow_html=True)

def staff_liste_demandes():
    # Si un ticket est sélectionné, afficher directement le traitement
    if st.session_state.get("edit_id"):
        staff_traiter_demande(st.session_state.edit_id)
        return
    
    # Appliquer filtre KPI si défini
    filtre_kpi = st.session_state.get("filtre_kpi", None)
    if filtre_kpi:
        st.info(f"🔍 Filtre actif: **{filtre_kpi}** - [Voir tous](javascript:void(0))")
        if st.button("❌ Effacer le filtre", key="clear_kpi_filter"):
            st.session_state.filtre_kpi = None
            st.rerun()
    
    # === FILTER BAR ===
    col1, col2, col3, col4, col5 = st.columns([2, 1.5, 1.5, 1.5, 1.5])
    with col1:
        # Si filtre KPI actif, pré-sélectionner le statut
        default_idx = 0
        if filtre_kpi and filtre_kpi in STATUTS:
            default_idx = STATUTS.index(filtre_kpi) + 1
        f_statut = st.selectbox("Statut", ["Tous"] + STATUTS, index=default_idx, key="f_statut", label_visibility="collapsed")
    with col2:
        f_code = st.text_input("N° Ticket", key="f_code", placeholder="🔍 KP-...", label_visibility="collapsed")
    with col3:
        f_tel = st.text_input("Téléphone", key="f_tel", placeholder="📞 06...", label_visibility="collapsed")
    with col4:
        f_nom = st.text_input("Nom", key="f_nom", placeholder="👤 Nom client", label_visibility="collapsed")
    with col5:
        # Filtre par technicien
        membres = get_membres_equipe()
        tech_options = ["👥 Tous"] + [m['nom'] for m in membres]
        f_tech = st.selectbox("Tech", tech_options, key="f_tech", label_visibility="collapsed")
    
    # Recherche avec les filtres - utiliser normalisation
    statut_filtre = filtre_kpi if filtre_kpi and filtre_kpi in STATUTS else (f_statut if f_statut != "Tous" else None)
    
    all_tickets = chercher_tickets(statut=statut_filtre)
    
    # Filtrer par code
    if f_code and f_code.strip():
        code_norm = normalize_search(f_code.strip())
        all_tickets = [t for t in all_tickets if code_norm in normalize_search(t.get('ticket_code', ''))]
    
    # Filtrer par téléphone  
    if f_tel and f_tel.strip():
        tel_norm = normalize_search(f_tel.strip())
        all_tickets = [t for t in all_tickets if tel_norm in normalize_search(t.get('client_tel', ''))]
    
    # Filtrer par nom (avec normalisation sans accents)
    if f_nom and f_nom.strip():
        nom_norm = normalize_search(f_nom.strip())
        all_tickets = [t for t in all_tickets if 
                      nom_norm in normalize_search(t.get('client_nom', '')) or 
                      nom_norm in normalize_search(t.get('client_prenom', ''))]
    
    # Filtrer par technicien si sélectionné
    if f_tech != "👥 Tous":
        all_tickets = [t for t in all_tickets if t.get('technicien_assigne') and f_tech in t.get('technicien_assigne', '')]
    
    # Séparer tickets actifs et archivés
    tickets_actifs = [t for t in all_tickets if t.get('statut') != 'Clôturé']
    tickets_archives = [t for t in all_tickets if t.get('statut') == 'Clôturé']
    
    # Afficher le compteur
    st.markdown(f"""
    <div style="display:flex;gap:16px;margin-bottom:12px;">
        <span style="background:#dcfce7;color:#166534;padding:4px 12px;border-radius:20px;font-size:13px;font-weight:600;">
            📋 {len(tickets_actifs)} actif(s)
        </span>
        <span style="background:#f3f4f6;color:#6b7280;padding:4px 12px;border-radius:20px;font-size:13px;">
            📦 {len(tickets_archives)} archivé(s)
        </span>
    </div>
    """, unsafe_allow_html=True)
    
    # Utiliser les tickets actifs pour l'affichage principal
    tickets = tickets_actifs
    
    # Pagination
    ITEMS_PER_PAGE = 8
    total_pages = max(1, (len(tickets) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE)
    
    if "accueil_page" not in st.session_state:
        st.session_state.accueil_page = 1
    
    current_page = st.session_state.accueil_page
    start_idx = (current_page - 1) * ITEMS_PER_PAGE
    end_idx = start_idx + ITEMS_PER_PAGE
    tickets_page = tickets[start_idx:end_idx]
    
    # === NOUVEAU DESIGN TABLEAU v3 - Streamlit columns alignées ===
    
    # Fonction helper pour obtenir les initiales
    def get_initials(nom, prenom):
        n = nom[0].upper() if nom else ""
        p = prenom[0].upper() if prenom else ""
        return f"{n}{p}" if n or p else "?"
    
    # Fonction helper pour icône appareil
    def get_device_icon(categorie):
        icons = {
            "Smartphone": "📱",
            "Tablette": "📟",
            "PC Portable": "💻",
            "Console": "🎮",
            "Commande": "📦"
        }
        return icons.get(categorie, "📱")

    # Proportions des colonnes (utilisées partout pour l'alignement)
    col_props = [1.0, 1.6, 1.8, 1.1, 1.6, 0.9, 0.5]
    
    # Header du tableau
    st.markdown("""
    <style>
    .ticket-table-wrapper {
        background: white;
        border: 1px solid #e5e5e5;
        border-radius: 16px;
        overflow: hidden;
        box-shadow: 0 1px 3px rgba(0,0,0,0.08);
        margin-bottom: 8px;
    }
    .ticket-table-header-row {
        background: linear-gradient(180deg, #f8f8f8 0%, #f3f3f3 100%);
        padding: 14px 16px;
        border-bottom: 1px solid #e5e5e5;
    }
    .ticket-row-wrapper {
        padding: 12px 16px;
        border-bottom: 1px solid #f0f0f0;
        transition: background 0.15s ease;
    }
    .ticket-row-wrapper:hover {
        background: #fefefe;
    }
    .ticket-row-wrapper:last-child {
        border-bottom: none;
    }
    </style>
    <div class="ticket-table-wrapper">
    <div class="ticket-table-header-row">
    """, unsafe_allow_html=True)
    
    # Header avec st.columns
    header_cols = st.columns(col_props)
    headers = ["Ticket", "Client", "Appareil", "Tech", "Statut", "Contact", ""]
    header_styles = "font-size:11px;font-weight:600;color:#737373;text-transform:uppercase;letter-spacing:0.5px;"
    
    for i, col in enumerate(header_cols):
        with col:
            st.markdown(f'<div style="{header_styles}">{headers[i]}</div>', unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)  # Fermer header row
    
    # Liste des tickets
    if not tickets_page:
        st.markdown("""
        <div style="padding:60px 20px;text-align:center;">
            <div style="font-size:48px;margin-bottom:16px;opacity:0.5;">📭</div>
            <div style="color:#a3a3a3;font-size:14px;">Aucun ticket trouvé.<br>Modifiez vos filtres ou créez un nouveau ticket.</div>
        </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        for idx, t in enumerate(tickets_page):
            status_class = get_status_class(t.get('statut', ''))
            
            # Données client
            client_nom = t.get('client_nom', '')
            client_prenom = t.get('client_prenom', '')
            client_full = f"{client_nom} {client_prenom}".strip()
            if len(client_full) > 16:
                client_full = client_full[:14] + "..."
            client_tel = t.get('client_tel', '')
            initials = get_initials(client_nom, client_prenom)
            
            # Données appareil - MODELE VISIBLE
            marque = t.get('marque', '')
            modele = t.get('modele', '')
            categorie = t.get('categorie', 'Smartphone')
            
            # Pour les commandes pièces, afficher la description
            if categorie == 'Commande':
                modele = t.get('modele_autre') or t.get('panne_detail') or 'Commande pièce'
                marque = '📦'
            elif t.get('modele_autre'): 
                modele = t['modele_autre']
            
            # Combiner marque + modèle pour affichage clair
            appareil_full = f"{marque} {modele}".strip()
            if len(appareil_full) > 22:
                appareil_display = appareil_full[:20] + "..."
            else:
                appareil_display = appareil_full
            device_icon = get_device_icon(categorie)
            
            # Technicien
            tech = t.get('technicien_assigne', '')
            tech_html = ""
            if tech:
                tech_color = "#9CA3AF"
                tech_name = tech
                for m in get_membres_equipe():
                    if m['nom'] in tech:
                        tech_name = m['nom']
                        tech_color = m['couleur']
                        break
                tech_html = f'<span style="background:{tech_color};color:white;padding:4px 10px;border-radius:16px;font-size:10px;font-weight:500;white-space:nowrap;">{tech_name}</span>'
            else:
                tech_html = '<span style="color:#a3a3a3;font-size:11px;font-style:italic;">—</span>'
            
            # Statut
            statut = t.get('statut', 'En attente')
            statut_short = statut
            if len(statut) > 18:
                statut_short = statut[:16] + "..."
            
            # Contact - INDICATEURS PLUS VISIBLES
            wa_bg = "#22c55e" if t.get('msg_whatsapp') else "#e5e5e5"
            wa_color = "white" if t.get('msg_whatsapp') else "#a3a3a3"
            sms_bg = "#3b82f6" if t.get('msg_sms') else "#e5e5e5"
            sms_color = "white" if t.get('msg_sms') else "#a3a3a3"
            email_bg = "#f59e0b" if t.get('msg_email') else "#e5e5e5"
            email_color = "white" if t.get('msg_email') else "#a3a3a3"
            
            # Date formatée
            date_depot = t.get('date_depot', '')
            date_formatted = "-"
            if date_depot:
                try:
                    from datetime import datetime
                    dt = datetime.fromisoformat(date_depot.replace('Z', '+00:00')) if 'T' in date_depot else datetime.strptime(date_depot[:10], '%Y-%m-%d')
                    date_formatted = dt.strftime('%d/%m')
                except:
                    pass
            
            # Wrapper de ligne
            st.markdown('<div class="ticket-row-wrapper">', unsafe_allow_html=True)
            
            # Colonnes de la ligne
            row_cols = st.columns(col_props)
            
            # Ticket
            with row_cols[0]:
                st.markdown(f'''
                <div>
                    <div style="font-family:'SF Mono',Monaco,monospace;font-size:12px;font-weight:600;color:#171717;">{t['ticket_code']}</div>
                    <div style="font-size:10px;color:#a3a3a3;">{date_formatted}</div>
                </div>
                ''', unsafe_allow_html=True)
            
            # Client
            with row_cols[1]:
                st.markdown(f'''
                <div style="display:flex;align-items:center;gap:8px;">
                    <div style="width:32px;height:32px;border-radius:50%;background:linear-gradient(135deg,#ffedd5,#fed7aa);display:flex;align-items:center;justify-content:center;font-size:10px;font-weight:600;color:#ea580c;flex-shrink:0;">{initials}</div>
                    <div style="min-width:0;overflow:hidden;">
                        <div style="font-size:12px;font-weight:500;color:#171717;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{client_full}</div>
                        <div style="font-size:10px;color:#a3a3a3;font-family:monospace;">{client_tel}</div>
                    </div>
                </div>
                ''', unsafe_allow_html=True)
            
            # Appareil - MODELE BIEN VISIBLE
            with row_cols[2]:
                st.markdown(f'''
                <div style="display:flex;align-items:center;gap:8px;">
                    <div style="width:28px;height:28px;border-radius:6px;background:#f5f5f5;display:flex;align-items:center;justify-content:center;font-size:13px;flex-shrink:0;">{device_icon}</div>
                    <div style="min-width:0;overflow:hidden;">
                        <div style="font-size:12px;font-weight:600;color:#171717;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;" title="{appareil_full}">{appareil_display}</div>
                        <div style="font-size:10px;color:#737373;">{categorie}</div>
                    </div>
                </div>
                ''', unsafe_allow_html=True)
            
            # Technicien
            with row_cols[3]:
                st.markdown(tech_html, unsafe_allow_html=True)
            
            # Statut
            with row_cols[4]:
                st.markdown(f'<span class="badge {status_class}" style="font-size:10px;" title="{statut}">{statut_short}</span>', unsafe_allow_html=True)
            
            # Contact
            with row_cols[5]:
                st.markdown(f'''
                <div style="display:flex;align-items:center;gap:3px;">
                    <div style="width:22px;height:22px;border-radius:5px;background:{wa_bg};display:flex;align-items:center;justify-content:center;font-size:10px;color:{wa_color};" title="WhatsApp {'✓' if t.get('msg_whatsapp') else ''}">📱</div>
                    <div style="width:22px;height:22px;border-radius:5px;background:{sms_bg};display:flex;align-items:center;justify-content:center;font-size:10px;color:{sms_color};" title="SMS {'✓' if t.get('msg_sms') else ''}">💬</div>
                    <div style="width:22px;height:22px;border-radius:5px;background:{email_bg};display:flex;align-items:center;justify-content:center;font-size:10px;color:{email_color};" title="Email {'✓' if t.get('msg_email') else ''}">✉️</div>
                </div>
                ''', unsafe_allow_html=True)
            
            # Action - Bouton Streamlit
            with row_cols[6]:
                if st.button("→", key=f"open_{t['id']}", help="Ouvrir", use_container_width=True):
                    st.session_state.edit_id = t['id']
                    st.rerun()
            
            st.markdown('</div>', unsafe_allow_html=True)  # Fermer row wrapper
        
        # Footer pagination
        pagination_dots = ""
        for i in range(min(total_pages, 10)):
            active_style = "background:#f97316;transform:scale(1.2);" if i + 1 == current_page else "background:#d4d4d4;"
            pagination_dots += f'<div style="width:8px;height:8px;border-radius:50%;{active_style}"></div>'
        
        st.markdown(f'''
        <div style="display:flex;align-items:center;justify-content:space-between;padding:12px 16px;background:#fafafa;border-top:1px solid #e5e5e5;">
            <div style="font-size:12px;color:#737373;">{len(tickets)} ticket(s) • Page {current_page}/{total_pages}</div>
            <div style="display:flex;gap:6px;">{pagination_dots}</div>
        </div>
        </div>
        ''', unsafe_allow_html=True)  # Fermer le wrapper principal
    
    # Navigation pagination (boutons Streamlit)
    if total_pages > 1:
        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
        col_prev, col_spacer, col_next = st.columns([1, 4, 1])
        with col_prev:
            if current_page > 1:
                if st.button("← Précédent", key="accueil_prev", type="secondary", use_container_width=True):
                    st.session_state.accueil_page = current_page - 1
                    st.rerun()
        with col_next:
            if current_page < total_pages:
                if st.button("Suivant →", key="accueil_next", type="secondary", use_container_width=True):
                    st.session_state.accueil_page = current_page + 1
                    st.rerun()
    
    # === SECTION TICKETS ARCHIVÉS ===
    if tickets_archives:
        st.markdown("<div style='height:24px;'></div>", unsafe_allow_html=True)
        with st.expander(f"📦 Tickets archivés ({len(tickets_archives)})", expanded=False):
            for t in tickets_archives[:20]:  # Limiter à 20 pour performance
                col_a1, col_a2, col_a3, col_a4, col_a5 = st.columns([1, 1.5, 1.5, 1, 0.5])
                with col_a1:
                    st.markdown(f"**{t['ticket_code']}**")
                with col_a2:
                    st.write(f"{t.get('client_nom', '')} {t.get('client_prenom', '')}")
                with col_a3:
                    modele = t.get('modele_autre') or f"{t.get('marque', '')} {t.get('modele', '')}"
                    st.write(modele[:25])
                with col_a4:
                    st.write(fmt_date(t.get('date_cloture', t.get('date_maj', '')))[:10])
                with col_a5:
                    if st.button("👁️", key=f"view_arch_{t['id']}", help="Voir"):
                        st.session_state.edit_id = t['id']
                        st.rerun()
                st.markdown("<hr style='margin:4px 0;border:none;border-top:1px solid #e5e7eb;'>", unsafe_allow_html=True)
            
            if len(tickets_archives) > 20:
                st.info(f"📋 {len(tickets_archives) - 20} autres tickets archivés...")

def staff_traiter_demande(tid):
    t = get_ticket_full(tid=tid)
    if not t:
        st.error("Demande non trouvée")
        return
    
    # === HEADER ===
    status_class = get_status_class(t.get('statut', ''))
    
    if t.get('categorie') == 'Commande':
        modele_txt = "📦 Commande pièce"
        if t.get('modele_autre'):
            modele_txt = t.get('modele_autre')
    else:
        modele_txt = f"{t.get('marque','')} {t.get('modele','')}"
        if t.get('modele_autre'): modele_txt += f" ({t['modele_autre']})"
    
    # Vérifier si on est en mode suppression
    if st.session_state.get(f"delete_ticket_{tid}"):
        st.markdown(f"""
        <div style="background:#fef2f2;border:2px solid #ef4444;border-radius:12px;padding:1.5rem;margin-bottom:1rem;">
            <h3 style="color:#dc2626;margin:0 0 1rem 0;">⚠️ Supprimer cette réparation ?</h3>
            <p><strong>{t['ticket_code']}</strong> - {t.get('client_nom','')} {t.get('client_prenom','')}</p>
            <p>📱 {modele_txt}</p>
            <p style="color:#dc2626;font-size:0.9rem;">Le client sera conservé, seule la réparation sera supprimée.</p>
        </div>
        """, unsafe_allow_html=True)
        
        pin_delete = st.text_input("🔐 Code PIN pour confirmer", type="password", key=f"pin_del_ticket_{tid}")
        
        col_del1, col_del2 = st.columns(2)
        with col_del1:
            if st.button("🗑️ Confirmer la suppression", type="primary", use_container_width=True, key=f"confirm_del_{tid}"):
                if pin_delete == "2626":
                    supprimer_ticket(tid)
                    st.success("✅ Réparation supprimée!")
                    st.session_state.edit_id = None
                    st.session_state[f"delete_ticket_{tid}"] = False
                    st.rerun()
                else:
                    st.error("❌ Code PIN incorrect!")
        with col_del2:
            if st.button("❌ Annuler", use_container_width=True, key=f"cancel_del_{tid}"):
                st.session_state[f"delete_ticket_{tid}"] = False
                st.rerun()
        return
    
    col_back, col_del, col_space = st.columns([1, 1, 5])
    with col_back:
        if st.button("← Retour", key="close_process", type="secondary", use_container_width=True):
            st.session_state.edit_id = None
            st.rerun()
    with col_del:
        if st.button("🗑️ Supprimer", key=f"del_ticket_btn_{tid}", type="secondary", use_container_width=True):
            st.session_state[f"delete_ticket_{tid}"] = True
            st.rerun()
    
    st.markdown(f"""
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:20px;">
        <div>
            <span style="font-size:1.5rem;font-weight:700;color:#1e293b;">{t['ticket_code']}</span>
            <span style="margin-left:16px;font-size:1rem;color:#64748b;">{t.get('client_nom','')} {t.get('client_prenom','')}</span>
        </div>
        <span class="badge {status_class}" style="font-size:0.9rem;padding:8px 16px;">{t.get('statut','')}</span>
    </div>
    """, unsafe_allow_html=True)
    
    # === COLONNES PRINCIPALES ===
    col1, col2 = st.columns([1, 1], gap="large")
    
    # =================================================================
    # COLONNE GAUCHE: Infos Client + Tickets
    # =================================================================
    with col1:
        st.markdown("""<div class="detail-card-header">👤 Client & Appareil</div>""", unsafe_allow_html=True)
        panne = t.get('panne', '')
        if t.get('panne_detail'): panne += f" ({t['panne_detail']})"
        
        # Label dynamique pour la précision
        panne_base = t.get('panne', '')
        if panne_base == "Écran casse":
            precision_label = "📱 Type écran"
        else:
            precision_label = "📝 Précision"
        
        camby_html = ' <span style="background:#22c55e;color:white;padding:2px 8px;border-radius:10px;font-size:11px;">🎫 CAMBY</span>' if t.get('client_carte_camby') else ''
        
        # Construire le HTML ligne par ligne
        html_lines = []
        html_lines.append('<div class="detail-card">')
        html_lines.append(f'<div class="detail-row"><span class="detail-label">Client</span><span class="detail-value">{t.get("client_nom","")} {t.get("client_prenom","")}{camby_html}</span></div>')
        html_lines.append(f'<div class="detail-row"><span class="detail-label">Téléphone</span><span class="detail-value" style="font-family:monospace;">{t.get("client_tel","")}</span></div>')
        html_lines.append(f'<div class="detail-row"><span class="detail-label">Appareil</span><span class="detail-value">{modele_txt}</span></div>')
        html_lines.append(f'<div class="detail-row"><span class="detail-label">Motif</span><span class="detail-value">{panne}</span></div>')
        
        # Précision si présente
        if t.get('type_ecran'):
            html_lines.append(f'<div class="detail-row" style="background:#dbeafe;border-radius:6px;padding:4px 8px;"><span class="detail-label" style="color:#1d4ed8;">{precision_label}</span><span class="detail-value" style="color:#1d4ed8;font-weight:600;">{t.get("type_ecran")}</span></div>')
        
        html_lines.append(f'<div class="detail-row"><span class="detail-label">Dépôt</span><span class="detail-value">{fmt_date(t.get("date_depot",""))}</span></div>')
        
        # Date récupération si présente
        if t.get('date_recuperation'):
            html_lines.append(f'<div class="detail-row" style="background:#dcfce7;border-radius:6px;padding:4px 8px;"><span class="detail-label" style="color:#166534;">📅 Récupération</span><span class="detail-value" style="color:#166534;font-weight:600;">{t.get("date_recuperation")}</span></div>')
        
        html_lines.append('</div>')
        
        st.markdown(''.join(html_lines), unsafe_allow_html=True)
        
        # Bouton modifier client
        if st.button("✏️ Modifier client/appareil", key=f"edit_client_btn_{tid}", type="secondary", use_container_width=True):
            st.session_state[f"show_edit_client_{tid}"] = not st.session_state.get(f"show_edit_client_{tid}", False)
            st.rerun()
        
        if st.session_state.get(f"show_edit_client_{tid}"):
            with st.container():
                client_id = t.get('client_id')
                client = get_client_by_id(client_id) if client_id else None
                if client:
                    c1, c2 = st.columns(2)
                    with c1:
                        new_nom = st.text_input("Nom", value=client.get('nom', ''), key=f"edit_nom_{tid}")
                        new_tel = st.text_input("Téléphone", value=client.get('telephone', ''), key=f"edit_tel_{tid}")
                    with c2:
                        new_prenom = st.text_input("Prénom", value=client.get('prenom', ''), key=f"edit_prenom_{tid}")
                        new_email = st.text_input("Email", value=client.get('email', '') or '', key=f"edit_email_{tid}")
                    
                    c3, c4 = st.columns(2)
                    with c3:
                        new_cat = st.selectbox("Catégorie", CATEGORIES, index=CATEGORIES.index(t.get('categorie', CATEGORIES[0])) if t.get('categorie') in CATEGORIES else 0, key=f"edit_cat_{tid}")
                        marques_dispo = get_marques(new_cat)
                        new_marque = st.selectbox("Marque", marques_dispo, index=marques_dispo.index(t.get('marque')) if t.get('marque') in marques_dispo else 0, key=f"edit_marque_{tid}")
                    with c4:
                        modeles_dispo = get_modeles(new_cat, new_marque)
                        new_modele = st.selectbox("Modèle", modeles_dispo, index=modeles_dispo.index(t.get('modele')) if t.get('modele') in modeles_dispo else 0, key=f"edit_modele_{tid}")
                        new_modele_autre = st.text_input("Modèle (autre)", value=t.get('modele_autre', ''), key=f"edit_modele_autre_{tid}")
                    
                    if st.button("💾 Enregistrer les modifications client", key=f"save_client_{tid}", type="primary", use_container_width=True):
                        update_client(client_id, nom=new_nom, prenom=new_prenom, telephone=new_tel, email=new_email)
                        update_ticket(tid, categorie=new_cat, marque=new_marque, modele=new_modele, modele_autre=new_modele_autre)
                        st.session_state[f"show_edit_client_{tid}"] = False
                        st.success("✅ Client mis à jour!")
                        st.rerun()
        
        # Codes sécurité
        if t.get('pin') or t.get('pattern'):
            st.markdown(f"""
            <div style="background:#fef3c7;border:1px solid #f59e0b;border-radius:8px;padding:12px;margin:12px 0;">
                <div style="font-weight:600;color:#92400e;margin-bottom:6px;">🔐 Codes de sécurité</div>
                <div style="font-size:0.9rem;">PIN: <strong>{t.get('pin') or '—'}</strong> | Schéma: <strong>{t.get('pattern') or '—'}</strong></div>
            </div>
            """, unsafe_allow_html=True)
        
        # --- TICKETS ---
        st.markdown('<div style="height:16px;"></div>', unsafe_allow_html=True)
        st.markdown('<div class="detail-card-header">🎫 Tickets & Documents</div>', unsafe_allow_html=True)
        
        col_t1, col_t2, col_t3, col_t4 = st.columns(4)
        with col_t1:
            if st.button("🎫 Client", key=f"show_client_{tid}", use_container_width=True):
                st.session_state[f"show_ticket_{tid}"] = "client" if st.session_state.get(f"show_ticket_{tid}") != "client" else None
                st.rerun()
        with col_t2:
            if st.button("📋 Staff", key=f"show_staff_{tid}", use_container_width=True):
                st.session_state[f"show_ticket_{tid}"] = "staff" if st.session_state.get(f"show_ticket_{tid}") != "staff" else None
                st.rerun()
        with col_t3:
            if st.button("📝 Devis", key=f"show_devis_{tid}", use_container_width=True):
                st.session_state[f"show_ticket_{tid}"] = "devis" if st.session_state.get(f"show_ticket_{tid}") != "devis" else None
                st.rerun()
        with col_t4:
            if st.button("🧾 Reçu", key=f"show_facture_{tid}", use_container_width=True):
                st.session_state[f"show_ticket_{tid}"] = "facture" if st.session_state.get(f"show_ticket_{tid}") != "facture" else None
                st.rerun()
        
        ticket_type = st.session_state.get(f"show_ticket_{tid}")
        if ticket_type:
            if ticket_type == "client":
                st.components.v1.html(ticket_client_html(t), height=700, scrolling=True)
            elif ticket_type == "staff":
                st.components.v1.html(ticket_staff_html(t), height=750, scrolling=True)
            elif ticket_type == "devis":
                st.components.v1.html(ticket_devis_facture_html(t, "devis"), height=700, scrolling=True)
            elif ticket_type == "facture":
                st.components.v1.html(ticket_devis_facture_html(t, "facture"), height=700, scrolling=True)
    
    # =================================================================
    # COLONNE DROITE: Gestion (avec auto-save)
    # =================================================================
    with col2:
        st.markdown("""<div class="detail-card-header">⚙️ Gestion de la réparation</div>""", unsafe_allow_html=True)
        
        # Type de réparation
        panne_actuelle = t.get('panne', PANNES[0])
        idx_panne = PANNES.index(panne_actuelle) if panne_actuelle in PANNES else 0
        new_panne = st.selectbox("Type de réparation", PANNES, index=idx_panne, key=f"rep_type_{tid}")
        
        # Auto-save panne
        if new_panne != panne_actuelle:
            update_ticket(tid, panne=new_panne)
            st.rerun()
        
        panne_detail = t.get('panne_detail') or ""
        if new_panne in ["Autre", "Diagnostic"]:
            new_panne_detail = st.text_input("Précisez", value=panne_detail, key=f"panne_detail_{tid}")
            if new_panne_detail != panne_detail:
                update_ticket(tid, panne_detail=new_panne_detail)
        
        # Précision sur la réparation (pour toutes les pannes)
        type_ecran_actuel = t.get('type_ecran') or ""
        if new_panne == "Écran casse":
            label_precision = "📱 Type d'écran"
            placeholder_precision = "Ex: Original, OLED, Incell, Premium..."
        else:
            label_precision = "📝 Précision sur la réparation"
            placeholder_precision = "Ex: Marque pièce, spécificité, remarque..."
        
        new_type_ecran = st.text_input(label_precision, value=type_ecran_actuel, placeholder=placeholder_precision, key=f"type_ecran_{tid}")
        # Note: type_ecran sera sauvegardé avec les notes pour éviter erreur si colonne absente
        
        # Technicien
        membres = get_membres_equipe()
        membres_options = ["-- Non assigné --"] + [f"{m['nom']} ({m['role']})" for m in membres]
        tech_actuel = t.get('technicien_assigne') or ""
        tech_idx = 0
        for i, opt in enumerate(membres_options):
            if tech_actuel and tech_actuel in opt:
                tech_idx = i
                break
        technicien = st.selectbox("👨‍🔧 Technicien", membres_options, index=tech_idx, key=f"technicien_{tid}")
        tech_name = technicien if technicien != "-- Non assigné --" else ""
        
        # Auto-save technicien
        if tech_name != tech_actuel:
            update_ticket(tid, technicien_assigne=tech_name)
            st.rerun()
        
        # Date récupération simplifiée
        st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)
        date_recup_actuelle = t.get('date_recuperation') or ""
        
        if st.button("📅 " + (f"Modifier récup: {date_recup_actuelle}" if date_recup_actuelle else "Définir date récupération"), key=f"toggle_date_{tid}", use_container_width=True):
            st.session_state[f"show_date_{tid}"] = not st.session_state.get(f"show_date_{tid}", False)
            st.rerun()
        
        if st.session_state.get(f"show_date_{tid}"):
            from datetime import datetime, timedelta, date as date_type
            default_date = datetime.now().date() + timedelta(days=1)
            default_hour = 18
            
            if date_recup_actuelle:
                try:
                    parts = date_recup_actuelle.replace("à", "").replace("h", "").split()
                    if len(parts) >= 2:
                        day_month = parts[0].split("/")
                        if len(day_month) == 2:
                            default_date = date_type(datetime.now().year, int(day_month[1]), int(day_month[0]))
                        default_hour = int(parts[1]) if len(parts) > 1 else 18
                except:
                    pass
            
            col_d1, col_d2, col_d3 = st.columns([1.5, 1, 1])
            with col_d1:
                date_picked = st.date_input("Date", value=default_date, min_value=datetime.now().date(), key=f"date_cal_{tid}", label_visibility="collapsed")
            with col_d2:
                heures = ["09h", "10h", "11h", "12h", "14h", "15h", "16h", "17h", "18h", "19h"]
                heure_defaut = f"{default_hour:02d}h" if f"{default_hour:02d}h" in heures else "18h"
                heure_idx = heures.index(heure_defaut) if heure_defaut in heures else 8
                heure_picked = st.selectbox("Heure", heures, index=heure_idx, key=f"heure_sel_{tid}", label_visibility="collapsed")
            with col_d3:
                if st.button("✓ OK", key=f"save_date_{tid}", type="primary", use_container_width=True):
                    new_date = f"{date_picked.strftime('%d/%m')} à {heure_picked}"
                    update_ticket(tid, date_recuperation=new_date)
                    st.session_state[f"show_date_{tid}"] = False
                    st.rerun()
        
        # Tarification
        st.markdown('<div style="height:12px;"></div>', unsafe_allow_html=True)
        st.markdown("""<div class="detail-card-header">💰 Tarification</div>""", unsafe_allow_html=True)
        
        col_a, col_b = st.columns(2)
        with col_a:
            devis_actuel = float(t.get('devis_estime') or 0)
            devis = st.number_input("Devis TTC (€)", value=devis_actuel, min_value=0.0, step=5.0, key=f"devis_{tid}")
        with col_b:
            acompte_actuel = float(t.get('acompte') or 0)
            acompte = st.number_input("Acompte (€)", value=acompte_actuel, min_value=0.0, step=5.0, key=f"acompte_{tid}")
        
        # Auto-save devis/acompte
        if devis != devis_actuel or acompte != acompte_actuel:
            update_ticket(tid, devis_estime=devis, acompte=acompte)
        
        # Réparation supplémentaire
        st.markdown('<div style="padding:8px;background:#f8fafc;border-radius:8px;border:1px dashed #94a3b8;margin-top:8px;">', unsafe_allow_html=True)
        st.markdown('<span style="font-size:11px;color:#475569;font-weight:600;">➕ Réparation supplémentaire</span>', unsafe_allow_html=True)
        col_rep1, col_rep2 = st.columns([2, 1])
        with col_rep1:
            rep_supp_actuel = t.get('reparation_supp') or ""
            rep_supp_val = st.text_input("Desc.", value=rep_supp_actuel, placeholder="Ex: Nappe Face ID...", key=f"rep_supp_{tid}", label_visibility="collapsed")
        with col_rep2:
            prix_supp_actuel = float(t.get('prix_supp') or 0)
            prix_supp_val = st.number_input("€", value=prix_supp_actuel, min_value=0.0, step=5.0, key=f"prix_supp_{tid}", label_visibility="collapsed")
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Auto-save réparation supp
        if rep_supp_val != rep_supp_actuel or prix_supp_val != prix_supp_actuel:
            update_ticket(tid, reparation_supp=rep_supp_val, prix_supp=prix_supp_val)
        
        # Calcul reste
        paye = t.get('paye', 0)
        total_ttc = float(devis) + float(prix_supp_val)
        reste = max(0, total_ttc - float(acompte))
        
        if paye:
            col_paye1, col_paye2 = st.columns([2, 1])
            with col_paye1:
                st.success("✅ PAYÉ")
            with col_paye2:
                if st.button("↩️", key=f"btn_unpaye_{tid}", help="Annuler le paiement"):
                    update_ticket(tid, paye=0)
                    clear_tickets_cache()
                    clear_ticket_full_cache()
                    st.rerun()
        else:
            if reste > 0:
                st.markdown(f'<div style="background:#fef3c7;border-radius:8px;padding:10px;text-align:center;margin-top:8px;"><strong style="color:#92400e;">💳 Reste à payer: {reste:.2f} €</strong></div>', unsafe_allow_html=True)
            if st.button("✅ MARQUER PAYÉ", key=f"btn_paye_{tid}", type="primary", use_container_width=True):
                update_ticket(tid, paye=1, acompte=total_ttc)
                # Forcer l'invalidation du cache
                clear_tickets_cache()
                clear_ticket_full_cache()
                st.rerun()
        
        # Statut
        st.markdown('<div style="height:12px;"></div>', unsafe_allow_html=True)
        statut_actuel = t.get('statut', STATUTS[0])
        idx_statut = STATUTS.index(statut_actuel) if statut_actuel in STATUTS else 0
        new_statut = st.selectbox("📍 Statut", STATUTS, index=idx_statut, key=f"statut_{tid}")
        
        # Auto-save statut
        if new_statut != statut_actuel:
            changer_statut(tid, new_statut)
            st.rerun()
        
        if statut_actuel == "En attente d'accord client":
            if st.button("✅ CLIENT A ACCEPTÉ", key=f"btn_accord_{tid}", type="primary", use_container_width=True):
                update_ticket(tid, client_accord=1)
                changer_statut(tid, "En cours de réparation")
                st.rerun()
        
        # --- CONTACT CLIENT ---
        st.markdown("---")
        st.markdown("##### 📞 Contacter le client")
        
        tel_client = t.get('client_tel', '')
        email_client = t.get('client_email', '')
        nom_boutique = get_param("NOM_BOUTIQUE") or "Klikphone"
        devis_val = t.get('devis_estime') or 0
        
        msg_devis = f"""Bonjour {t.get('client_prenom', '')},

Suite au diagnostic de votre {modele_txt}, voici notre devis:

🔧 Réparation: {t.get('panne', '')}
💰 Montant: {devis_val:.2f} € TTC

Merci de nous confirmer votre accord.

{nom_boutique}
📞 {get_param('TEL_BOUTIQUE')}"""
        
        if tel_client:
            wa_url = wa_link(tel_client, msg_devis)
            sms_url = sms_link(tel_client, msg_devis)
            st.markdown(f"""
            <div style="display:flex;gap:8px;margin-bottom:12px;">
                <a href="{wa_url}" target="_blank" style="flex:1;display:block;text-align:center;padding:12px;background:#25D366;color:white;border-radius:8px;text-decoration:none;font-weight:600;">📱 WhatsApp</a>
                <a href="{sms_url}" style="flex:1;display:block;text-align:center;padding:12px;background:#3b82f6;color:white;border-radius:8px;text-decoration:none;font-weight:600;">💬 SMS</a>
            </div>
            """, unsafe_allow_html=True)
        
        if email_client and get_param("SMTP_HOST"):
            col_e1, col_e2 = st.columns(2)
            with col_e1:
                if st.button("📧 Envoyer Ticket", key=f"email_tkt_{tid}", use_container_width=True):
                    html = ticket_client_html(t, for_email=True)
                    success, _ = envoyer_email(email_client, f"Ticket {t['ticket_code']}", "Votre ticket.", html)
                    if success:
                        update_ticket(tid, msg_email=1)
                        st.success("✅ Envoyé!")
            with col_e2:
                if st.button("📧 Envoyer Devis", key=f"email_dev_{tid}", use_container_width=True):
                    html = ticket_devis_facture_html(t, "devis", for_email=True)
                    success, _ = envoyer_email(email_client, f"Devis D-{t['ticket_code']}", "Votre devis.", html)
                    if success:
                        update_ticket(tid, msg_email=1)
                        st.success("✅ Envoyé!")
        
        # --- INTÉGRATION CAISSE --- (Toujours visible)
        st.markdown("---")
        st.markdown("##### 💳 Caisse Enregistreuse")
        
        caisse_apikey = get_param("CAISSE_APIKEY") or ""
        caisse_shopid = get_param("CAISSE_SHOPID") or ""
        caisse_enabled = get_param("CAISSE_ENABLED") == "1"
        caisse_configured = caisse_enabled and caisse_apikey and caisse_shopid
        total_ticket = float(t.get('devis_estime') or 0) + float(t.get('prix_supp') or 0)
        
        if not caisse_configured:
            if not caisse_enabled:
                st.warning("⚠️ Activez l'intégration dans ⚙️ Config > 💳 Caisse")
            elif not caisse_apikey:
                st.warning("⚠️ Token API manquant - cliquez 'Obtenir le token' dans Config > Caisse")
            elif not caisse_shopid:
                st.warning("⚠️ SHOPID manquant - cliquez 'Obtenir le token' dans Config > Caisse")
        elif total_ticket <= 0:
            st.info("💡 Renseignez un devis pour envoyer vers la caisse")
        else:
            # Récupérer les IDs configurés
            cb_id = get_param("CAISSE_CB_ID") or ""
            esp_id = get_param("CAISSE_ESP_ID") or ""
            caisse_id = get_param("CAISSE_ID") or ""
            
            # Vérifier si configuré
            if not cb_id and not esp_id:
                st.error("⚠️ **Modes de paiement non configurés !**")
                st.warning("Allez dans **⚙️ Config > 💳 Caisse** et cliquez sur **'Récupérer mes MODES DE PAIEMENT'**")
            elif not caisse_id:
                st.error("⚠️ **Caisse non configurée !** (Les ventes iront vers 'webservice')")
                st.warning("Allez dans **⚙️ Config > 💳 Caisse** et cliquez sur **'Récupérer mes CAISSES'** puis entrez l'ID")
                
                # Permettre quand même d'envoyer
                mode_envoi = st.radio(
                    "Mode de paiement",
                    ["💳 Carte bancaire", "💵 Espèces", "📝 Non payée"],
                    index=0,
                    key=f"mode_paiement_envoi_{tid}",
                    horizontal=True
                )
                
                if mode_envoi == "💳 Carte bancaire":
                    mode_val = cb_id
                elif mode_envoi == "💵 Espèces":
                    mode_val = esp_id
                else:
                    mode_val = "-1"
                
                st.caption(f"🔧 payment={mode_val} | **caisse=NON CONFIGURÉE**")
                
                if st.button(f"📤 Envoyer (caisse webservice)", key=f"send_caisse_{tid}", type="secondary", use_container_width=True):
                    success, message = envoyer_vers_caisse(t, payment_override=mode_val)
                    if success:
                        st.success(f"✅ {message}")
                    else:
                        st.error(f"❌ {message}")
            else:
                # Tout est configuré !
                st.success(f"✅ Config OK: CB={cb_id} | ESP={esp_id} | Caisse={caisse_id}")
                
                mode_envoi = st.radio(
                    "Mode de paiement",
                    ["💳 Carte bancaire", "💵 Espèces", "📝 Non payée"],
                    index=0,
                    key=f"mode_paiement_envoi_{tid}",
                    horizontal=True
                )
                
                if mode_envoi == "💳 Carte bancaire":
                    mode_val = cb_id
                elif mode_envoi == "💵 Espèces":
                    mode_val = esp_id
                else:
                    mode_val = "-1"
                
                st.caption(f"🔧 payment={mode_val} | idcaisse={caisse_id}")
                
                if st.button(f"📤 Envoyer à la caisse ({total_ticket:.2f} €)", key=f"send_caisse_{tid}", type="primary", use_container_width=True):
                    success, message = envoyer_vers_caisse(t, payment_override=mode_val)
                    if success:
                        st.success(f"✅ {message}")
                    else:
                        st.error(f"❌ {message}")
    
    # =================================================================
    # ZONE BAS: Notes (gauche) + Notifications (droite)
    # =================================================================
    st.markdown('<div style="height:24px;"></div>', unsafe_allow_html=True)
    st.markdown("---")
    
    col_notes, col_notifs = st.columns([1, 1], gap="large")
    
    with col_notes:
        st.markdown("""<div class="detail-card-header">📝 Notes</div>""", unsafe_allow_html=True)
        
        # Note publique
        st.markdown('<div style="background:#ecfdf5;border-left:4px solid #22c55e;padding:8px 12px;margin-bottom:8px;border-radius:0 8px 8px 0;"><span style="font-weight:600;font-size:12px;color:#166534;">💬 Note publique</span> <span style="font-size:10px;color:#22c55e;">— visible sur ticket client</span></div>', unsafe_allow_html=True)
        
        note_pub_actuelle = t.get('commentaire_client') or ""
        new_note_pub = st.text_area("Note publique", value=note_pub_actuelle, height=80, key=f"notes_pub_{tid}", label_visibility="collapsed")
        
        # Note privée
        st.markdown('<div style="background:#fef2f2;border-left:4px solid #ef4444;padding:8px 12px;margin:12px 0 8px 0;border-radius:0 8px 8px 0;"><span style="font-weight:600;font-size:12px;color:#dc2626;">🔒 Note privée</span> <span style="font-size:10px;color:#ef4444;">— équipe uniquement</span></div>', unsafe_allow_html=True)
        
        note_int_actuelle = t.get('notes_internes') or ""
        new_note_int = st.text_area("Note privée", value=note_int_actuelle, height=80, key=f"notes_int_{tid}", label_visibility="collapsed")
        
        # Bouton enregistrer notes + précision
        if st.button("💾 Enregistrer", key=f"save_notes_{tid}", type="primary", use_container_width=True):
            # Récupérer type_ecran depuis session_state
            final_type_ecran = st.session_state.get(f"type_ecran_{tid}", t.get('type_ecran') or "")
            try:
                update_ticket(tid, commentaire_client=new_note_pub, notes_internes=new_note_int, type_ecran=final_type_ecran)
            except:
                # Si erreur (colonne type_ecran n'existe pas), sauvegarder sans
                update_ticket(tid, commentaire_client=new_note_pub, notes_internes=new_note_int)
            st.success("✅ Enregistré!")
    
    with col_notifs:
        st.markdown("""<div class="detail-card-header">📊 Centre de notifications</div>""", unsafe_allow_html=True)
        
        wa_on = t.get('msg_whatsapp')
        sms_on = t.get('msg_sms')
        email_on = t.get('msg_email')
        
        st.markdown(f"""
        <div style="background:#1e293b;border-radius:12px;padding:16px;color:white;">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
                <span style="font-size:12px;color:rgba(255,255,255,0.7);">📍 Statut actuel</span>
                <span style="background:rgba(249,115,22,0.3);color:#fdba74;padding:4px 12px;border-radius:12px;font-size:12px;font-weight:600;">{t.get('statut', '')}</span>
            </div>
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
                <span style="font-size:12px;color:rgba(255,255,255,0.7);">📨 Communications</span>
                <div style="display:flex;gap:6px;">
                    <span style="background:{'#22c55e' if wa_on else '#374151'};color:white;padding:3px 8px;border-radius:10px;font-size:10px;">WA{'✓' if wa_on else ''}</span>
                    <span style="background:{'#3b82f6' if sms_on else '#374151'};color:white;padding:3px 8px;border-radius:10px;font-size:10px;">SMS{'✓' if sms_on else ''}</span>
                    <span style="background:{'#f59e0b' if email_on else '#374151'};color:white;padding:3px 8px;border-radius:10px;font-size:10px;">Mail{'✓' if email_on else ''}</span>
                </div>
            </div>
            <div style="font-size:10px;color:rgba(255,255,255,0.4);text-align:right;">Dernière màj: {str(t.get('date_maj', ''))[:16]}</div>
        </div>
        """, unsafe_allow_html=True)
        
        # Notes du client (de dépôt)
        if t.get('notes_client'):
            st.markdown(f'<div style="background:#fff7ed;border-left:3px solid #f97316;padding:8px 12px;margin-top:12px;border-radius:0 8px 8px 0;font-size:12px;"><strong>📋 Note client (dépôt):</strong> {t.get("notes_client")}</div>', unsafe_allow_html=True)
        
        # Infos commande si applicable
        if t.get('panne_detail') and t.get('categorie') == 'Commande':
            st.markdown(f'<div style="background:#f3e8ff;border-left:3px solid #a855f7;padding:8px 12px;margin-top:8px;border-radius:0 8px 8px 0;font-size:12px;"><strong>📦 Commande:</strong> {t.get("panne_detail")}</div>', unsafe_allow_html=True)



def staff_gestion_clients():
    """Gestion des clients - Liste, modification, export, suppression"""
    
    # Header avec export
    col_title, col_export, col_export2 = st.columns([3, 1, 1])
    with col_title:
        st.markdown("""<div class="detail-card-header">👥 Gestion des Clients</div>""", unsafe_allow_html=True)
    with col_export:
        data, filename = export_clients_excel()
        if data:
            st.download_button(
                label="📥 Excel",
                data=data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if filename.endswith('.xlsx') else "text/csv",
                key="download_clients",
                type="primary",
                use_container_width=True
            )
    with col_export2:
        # Export simplifié (nom, prénom, téléphone)
        try:
            import pandas as pd
            from io import BytesIO
            conn = get_db()
            c = conn.cursor()
            c.execute("SELECT nom, prenom, telephone FROM clients ORDER BY nom, prenom")
            clients_simple = [dict(row) for row in c.fetchall()]
            conn.close()
            
            if clients_simple:
                df = pd.DataFrame(clients_simple)
                output = BytesIO()
                df.to_excel(output, index=False, sheet_name='Clients')
                st.download_button(
                    label="📋 Simple",
                    data=output.getvalue(),
                    file_name="clients_contacts.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_clients_simple",
                    use_container_width=True
                )
        except:
            pass
    
    st.markdown("---")
    
    # Mode nouvelle réparation pour client existant
    if st.session_state.get("new_repair_client_id"):
        client_id = st.session_state.new_repair_client_id
        client = get_client_by_id(client_id)
        
        if client:
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,#fff7ed,#ffedd5);border:2px solid #f97316;border-radius:12px;padding:1.5rem;margin-bottom:1rem;">
                <h3 style="color:#ea580c;margin:0 0 0.5rem 0;">➕ Nouvelle réparation</h3>
                <p style="margin:0;"><strong>{client.get('nom', '')} {client.get('prenom', '')}</strong> — 📞 {client.get('telephone', '')}</p>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            with col1:
                new_cat = st.selectbox("Type d'appareil", ["Smartphone", "Tablette", "PC Portable", "Console", "Autre"], key="new_repair_cat")
                marques_liste = list(MARQUES.get(new_cat, ["Autre"]))
                new_marque = st.selectbox("Marque", marques_liste + ["Autre"], key="new_repair_marque")
            with col2:
                modeles_liste = MODELES.get((new_cat, new_marque), ["Autre"])
                new_modele = st.selectbox("Modèle", modeles_liste + ["Autre"], key="new_repair_modele")
                new_modele_autre = ""
                if new_modele == "Autre":
                    new_modele_autre = st.text_input("Précisez le modèle", key="new_repair_modele_autre")
            
            col3, col4 = st.columns(2)
            with col3:
                new_panne = st.selectbox("Type de réparation", PANNES, key="new_repair_panne")
                new_panne_detail = ""
                if new_panne in ["Autre", "Diagnostic"]:
                    new_panne_detail = st.text_input("Détails", key="new_repair_panne_detail")
            with col4:
                new_pin = st.text_input("Code PIN", key="new_repair_pin")
                new_pattern = st.text_input("Schéma", key="new_repair_pattern")
            
            new_notes = st.text_area("Notes client", placeholder="Remarques éventuelles...", height=80, key="new_repair_notes")
            
            col_save, col_cancel = st.columns(2)
            with col_save:
                if st.button("✅ Créer la réparation", type="primary", use_container_width=True, key="save_new_repair"):
                    code = creer_ticket(client_id, new_cat, new_marque, new_modele, new_modele_autre, 
                                       new_panne, new_panne_detail, new_pin, new_pattern, new_notes, "", 0)
                    st.success(f"✅ Réparation créée: **{code}**")
                    st.session_state.new_repair_client_id = None
                    st.balloons()
                    st.rerun()
            with col_cancel:
                if st.button("❌ Annuler", type="secondary", use_container_width=True, key="cancel_new_repair"):
                    st.session_state.new_repair_client_id = None
                    st.rerun()
            
            st.markdown("---")
            return
    
    # Mode suppression client (avec PIN)
    if st.session_state.get("delete_client_id"):
        client_id = st.session_state.delete_client_id
        client = get_client_by_id(client_id)
        
        if client:
            st.markdown(f"""
            <div style="background:#fef2f2;border:2px solid #ef4444;border-radius:12px;padding:1.5rem;margin-bottom:1rem;">
                <h3 style="color:#dc2626;margin:0 0 1rem 0;">⚠️ Supprimer le client</h3>
                <p><strong>{client.get('nom', '')} {client.get('prenom', '')}</strong></p>
                <p>📞 {client.get('telephone', '')}</p>
                <p style="color:#dc2626;font-size:0.9rem;">Cette action supprimera également tous les tickets associés à ce client !</p>
            </div>
            """, unsafe_allow_html=True)
            
            pin_delete = st.text_input("🔐 Entrez le code PIN pour confirmer", type="password", key="pin_delete_client")
            
            col_del1, col_del2 = st.columns(2)
            with col_del1:
                if st.button("🗑️ Confirmer la suppression", type="primary", use_container_width=True):
                    if pin_delete == "2626":
                        supprimer_client(client_id)
                        st.success("✅ Client supprimé avec succès!")
                        st.session_state.delete_client_id = None
                        st.rerun()
                    else:
                        st.error("❌ Code PIN incorrect!")
            with col_del2:
                if st.button("❌ Annuler", use_container_width=True):
                    st.session_state.delete_client_id = None
                    st.rerun()
            
            st.markdown("---")
            return
    
    # Mode édition client
    if st.session_state.get("edit_client_id"):
        client_id = st.session_state.edit_client_id
        client = get_client_by_id(client_id)
        
        if client:
            st.markdown(f"### ✏️ Modifier le client #{client_id}")
            
            col1, col2 = st.columns(2)
            with col1:
                edit_nom = st.text_input("Nom *", value=client.get('nom', ''), key="edit_client_nom")
                edit_prenom = st.text_input("Prénom", value=client.get('prenom', ''), key="edit_client_prenom")
                edit_societe = st.text_input("Société (facultatif)", value=client.get('societe', '') or '', key="edit_client_societe")
            with col2:
                edit_tel = st.text_input("Téléphone *", value=client.get('telephone', ''), key="edit_client_tel")
                edit_email = st.text_input("Email", value=client.get('email', ''), key="edit_client_email")
            
            col_save, col_cancel = st.columns(2)
            with col_save:
                if st.button("💾 Enregistrer", type="primary", use_container_width=True, key="save_client"):
                    if edit_nom and edit_tel:
                        update_client(client_id, nom=edit_nom, prenom=edit_prenom, telephone=edit_tel, email=edit_email, societe=edit_societe)
                        st.success("✅ Client mis à jour!")
                        st.session_state.edit_client_id = None
                        st.rerun()
                    else:
                        st.error("Nom et téléphone obligatoires")
            with col_cancel:
                if st.button("❌ Annuler", type="secondary", use_container_width=True, key="cancel_edit_client"):
                    st.session_state.edit_client_id = None
                    st.rerun()
            
            st.markdown("---")
        else:
            st.session_state.edit_client_id = None
            st.rerun()
    
    # Recherche
    search = st.text_input("🔍 Rechercher un client", placeholder="Nom, prénom, téléphone ou société...", key="search_clients")
    
    # Liste des clients
    if search and len(search) >= 2:
        clients = search_clients(search)
    else:
        clients = get_all_clients()
    
    st.markdown(f"**{len(clients)} client(s)**")
    
    # Table header - ajout colonne nouvelle réparation
    st.markdown("""
    <div class="table-header">
        <div style="flex:1;">Nom</div>
        <div style="flex:1;">Prénom</div>
        <div style="flex:0.8;">Société</div>
        <div style="flex:1;">Téléphone</div>
        <div style="flex:1.1;">Email</div>
        <div style="min-width:40px;">📋</div>
        <div style="min-width:100px;">Actions</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Pagination
    ITEMS_PER_PAGE = 15
    total_pages = max(1, (len(clients) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE)
    
    if "clients_page" not in st.session_state:
        st.session_state.clients_page = 1
    
    current_page = st.session_state.clients_page
    start_idx = (current_page - 1) * ITEMS_PER_PAGE
    end_idx = start_idx + ITEMS_PER_PAGE
    clients_page = clients[start_idx:end_idx]
    
    for client in clients_page:
        col1, col2, col3, col4, col5, col6, col7, col8, col9 = st.columns([1, 1, 0.8, 1, 1.1, 0.3, 0.3, 0.3, 0.4])
        with col1:
            st.markdown(f"**{client.get('nom', '')}**")
        with col2:
            st.write(client.get('prenom', ''))
        with col3:
            societe = client.get('societe', '')
            st.write(societe[:10] if societe else "—")
        with col4:
            st.write(client.get('telephone', ''))
        with col5:
            email = client.get('email', '')
            st.write(email[:18] if email else "—")
        with col6:
            st.write(f"{client.get('nb_tickets', 0)}")
        with col7:
            if st.button("✏️", key=f"edit_client_{client['id']}", help="Modifier"):
                st.session_state.edit_client_id = client['id']
                st.rerun()
        with col8:
            if st.button("🗑️", key=f"del_client_{client['id']}", help="Supprimer"):
                st.session_state.delete_client_id = client['id']
                st.rerun()
        with col9:
            if st.button("➕", key=f"new_repair_{client['id']}", help="Nouvelle réparation", type="primary"):
                st.session_state.new_repair_client_id = client['id']
                st.rerun()
        
        st.markdown("<div style='height:1px;background:var(--neutral-100);margin:4px 0;'></div>", unsafe_allow_html=True)
    
    # Pagination
    if total_pages > 1:
        st.markdown("---")
        col_prev, col_info, col_next = st.columns([1, 2, 1])
        with col_prev:
            if current_page > 1:
                if st.button("← Précédent", key="clients_prev"):
                    st.session_state.clients_page = current_page - 1
                    st.rerun()
        with col_info:
            st.markdown(f"<div style='text-align:center;'>Page {current_page}/{total_pages}</div>", unsafe_allow_html=True)
        with col_next:
            if current_page < total_pages:
                if st.button("Suivant →", key="clients_next"):
                    st.session_state.clients_page = current_page + 1
                    st.rerun()

# Décorateur fragment conditionnel (compatibilité Streamlit < 1.33)
def optional_fragment(func):
    """Applique @st.fragment si disponible, sinon retourne la fonction telle quelle"""
    if hasattr(st, 'fragment'):
        return st.fragment(func)
    return func

@optional_fragment
def staff_commandes_pieces():
    """Gestion des commandes de pièces - FRAGMENT pour éviter les reruns globaux"""
    st.markdown("""<div class="detail-card-header">📦 Commandes de Pièces</div>""", unsafe_allow_html=True)
    
    # Onglets
    sub_tab1, sub_tab2, sub_tab3, sub_tab4 = st.tabs(["🔴 À commander", "🟡 Commandées", "✅ Reçues", "➕ Nouvelle"])
    
    with sub_tab1:
        # Pièces à commander
        commandes = get_commandes_pieces(statut="A commander")
        
        if not commandes:
            st.info("✅ Aucune pièce à commander")
        else:
            st.markdown(f"**{len(commandes)} pièce(s) à commander**")
            
            for cmd in commandes:
                cmd_id = cmd['id']
                
                # Vérifier si on est en mode édition pour cette commande
                if st.session_state.get(f"edit_cmd_{cmd_id}"):
                    with st.container():
                        st.markdown(f"**✏️ Modifier la commande #{cmd_id}**")
                        col1, col2 = st.columns(2)
                        with col1:
                            edit_desc = st.text_input("Description", value=cmd['description'], key=f"edit_desc_{cmd_id}")
                            edit_fournisseur = st.selectbox("Fournisseur", FOURNISSEURS, 
                                index=FOURNISSEURS.index(cmd.get('fournisseur')) if cmd.get('fournisseur') in FOURNISSEURS else 0,
                                key=f"edit_fourn_{cmd_id}")
                        with col2:
                            edit_ref = st.text_input("Référence", value=cmd.get('reference', ''), key=f"edit_ref_{cmd_id}")
                            edit_prix = st.number_input("Prix (€)", value=float(cmd.get('prix', 0)), min_value=0.0, step=1.0, key=f"edit_prix_{cmd_id}")
                        
                        col_save, col_cancel = st.columns(2)
                        with col_save:
                            if st.button("💾 Enregistrer", key=f"save_edit_cmd_{cmd_id}", type="primary"):
                                conn = get_db()
                                c = conn.cursor()
                                c.execute("UPDATE commandes_pieces SET description=?, fournisseur=?, reference=?, prix=? WHERE id=?",
                                         (edit_desc, edit_fournisseur, edit_ref, edit_prix, cmd_id))
                                conn.commit()
                                conn.close()
                                st.session_state[f"edit_cmd_{cmd_id}"] = False
                                st.success("✅ Commande modifiée!")
                                st.rerun()
                        with col_cancel:
                            if st.button("❌ Annuler", key=f"cancel_edit_cmd_{cmd_id}"):
                                st.session_state[f"edit_cmd_{cmd_id}"] = False
                                st.rerun()
                        st.markdown("---")
                else:
                    with st.container():
                        col1, col2, col3, col4, col5 = st.columns([2.5, 1.5, 0.8, 0.8, 0.6])
                        with col1:
                            ticket_info = f"{cmd.get('ticket_code', 'N/A')} - {cmd.get('client_nom', '')} {cmd.get('client_prenom', '')}"
                            
                            # Afficher la vraie description de la commande
                            description = cmd['description']
                            # Si c'est une commande pièce du totem, afficher aussi modele_autre/panne_detail
                            if cmd.get('modele_autre') or cmd.get('panne_detail'):
                                detail_commande = cmd.get('modele_autre') or cmd.get('panne_detail') or ''
                                if detail_commande and detail_commande not in description:
                                    st.markdown(f"**📝 {detail_commande}**")
                                else:
                                    st.markdown(f"**{description}**")
                            else:
                                st.markdown(f"**{description}**")
                            
                            st.caption(f"📋 {ticket_info}")
                            # Afficher appareil
                            appareil = ""
                            if cmd.get('marque') and cmd.get('modele'):
                                appareil = f"{cmd['marque']} {cmd['modele']}"
                            if cmd.get('modele_autre') and cmd.get('categorie') != 'Commande':
                                appareil = cmd['modele_autre']
                            if appareil:
                                st.caption(f"📱 {appareil}")
                        with col2:
                            st.write(f"🏪 {cmd.get('fournisseur', 'N/A')}")
                            if cmd.get('reference'):
                                st.caption(f"Réf: {cmd['reference']}")
                        with col3:
                            if cmd.get('prix') and cmd['prix'] > 0:
                                st.write(f"💰 {cmd['prix']:.2f} €")
                        with col4:
                            if st.button("✏️", key=f"btn_edit_cmd_{cmd_id}", help="Modifier"):
                                st.session_state[f"edit_cmd_{cmd_id}"] = True
                                st.rerun()
                            if st.button("✅", key=f"cmd_done_{cmd_id}", help="Commandée"):
                                from datetime import datetime
                                update_commande_piece(cmd_id, statut="Commandée", date_commande=datetime.now().strftime("%Y-%m-%d %H:%M"))
                                st.rerun()
                        with col5:
                            if st.button("🗑️", key=f"cmd_del_{cmd_id}", help="Supprimer"):
                                delete_commande_piece(cmd_id)
                                st.rerun()
                        
                        st.markdown("<hr style='margin:10px 0;border-color:#eee;'>", unsafe_allow_html=True)
    
    with sub_tab2:
        # Pièces commandées (en attente de réception)
        commandes = get_commandes_pieces(statut="Commandée")
        
        if not commandes:
            st.info("Aucune commande en cours")
        else:
            st.markdown(f"**{len(commandes)} commande(s) en cours**")
            
            for cmd in commandes:
                cmd_id = cmd['id']
                client_tel = cmd.get('client_tel', '')
                client_prenom = cmd.get('client_prenom', '')
                ticket_code = cmd.get('ticket_code', '')
                
                # Message pré-rempli
                msg_cmd = f"""Bonjour {client_prenom},

Votre commande est bien arrivée ! 📦

Réf: {ticket_code}
{cmd['description']}

Vous pouvez passer à la boutique.

{get_param("NOM_BOUTIQUE") or "Klikphone"}
📞 {get_param("TEL_BOUTIQUE") or "04 79 60 89 22"}"""
                
                with st.container():
                    col1, col2, col3, col4, col5 = st.columns([2, 1.2, 0.8, 1.3, 0.7])
                    with col1:
                        ticket_info = f"{cmd.get('ticket_code', 'N/A')} - {cmd.get('client_nom', '')} {cmd.get('client_prenom', '')}"
                        st.markdown(f"**{cmd['description']}**")
                        st.caption(f"📋 {ticket_info}")
                    with col2:
                        st.write(f"🏪 {cmd.get('fournisseur', 'N/A')}")
                        if cmd.get('date_commande'):
                            st.caption(f"Commandé le {cmd['date_commande'][:10]}")
                    with col3:
                        if cmd.get('prix') and cmd['prix'] > 0:
                            st.write(f"💰 {cmd['prix']:.2f} €")
                    with col4:
                        if st.button("📦 Reçue", key=f"cmd_recv_{cmd_id}", type="primary", use_container_width=True):
                            from datetime import datetime
                            update_commande_piece(cmd_id, statut="Reçue", date_reception=datetime.now().strftime("%Y-%m-%d %H:%M"))
                            st.rerun()
                    with col5:
                        # Icônes contact direct
                        if client_tel:
                            wa_url = wa_link(client_tel, msg_cmd)
                            sms_url = sms_link(client_tel, msg_cmd)
                            st.markdown(f'''
                            <div style="display:flex;gap:4px;">
                                <a href="{wa_url}" target="_blank" style="background:#25D366;color:white;padding:6px 8px;border-radius:6px;text-decoration:none;font-size:14px;" title="WhatsApp">📱</a>
                                <a href="{sms_url}" style="background:#3b82f6;color:white;padding:6px 8px;border-radius:6px;text-decoration:none;font-size:14px;" title="SMS">💬</a>
                            </div>
                            ''', unsafe_allow_html=True)
                    
                    st.markdown("<hr style='margin:10px 0;border-color:#eee;'>", unsafe_allow_html=True)
    
    with sub_tab3:
        # Commandes reçues (historique)
        commandes = get_commandes_pieces(statut="Reçue")
        
        if not commandes:
            st.info("Aucune commande reçue")
        else:
            st.markdown(f"**{len(commandes)} commande(s) reçue(s)**")
            
            # En-tête tableau
            st.markdown("""
            <div style="display:flex;background:#f1f5f9;padding:8px;border-radius:6px;margin-bottom:8px;font-weight:600;font-size:0.85rem;">
                <div style="flex:2;">Description</div>
                <div style="flex:1.2;">Ticket</div>
                <div style="flex:0.8;">Fourn.</div>
                <div style="flex:0.6;">Prix</div>
                <div style="flex:0.8;">Reçu le</div>
                <div style="flex:0.8;">Notifier</div>
            </div>
            """, unsafe_allow_html=True)
            
            for cmd in commandes:
                client_tel = cmd.get('client_tel', '')
                client_prenom = cmd.get('client_prenom', '')
                client_email = cmd.get('client_email', '')
                ticket_code = cmd.get('ticket_code', 'N/A')
                date_recv = cmd.get('date_reception', '')[:10] if cmd.get('date_reception') else '—'
                
                # Message pré-rempli pour notification
                msg_pret = f"""Bonjour {client_prenom},

Votre commande est arrivée ! 📦

Réf: {ticket_code}
{cmd['description']}

Vous pouvez passer à la boutique.

{get_param("NOM_BOUTIQUE") or "Klikphone"}
📞 {get_param("TEL_BOUTIQUE") or "04 79 60 89 22"}"""
                
                col1, col2, col3, col4, col5, col6 = st.columns([2, 1.2, 0.8, 0.6, 0.8, 0.8])
                with col1:
                    st.write(cmd['description'])
                with col2:
                    st.write(ticket_code)
                with col3:
                    st.write(cmd.get('fournisseur', '—')[:8])
                with col4:
                    st.write(f"{cmd['prix']:.0f}€")
                with col5:
                    st.write(date_recv)
                with col6:
                    if client_tel:
                        wa_url = wa_link(client_tel, msg_pret)
                        sms_url = sms_link(client_tel, msg_pret)
                        email_url = f"mailto:{client_email}?subject=Commande%20arrivée%20-%20{ticket_code}&body={urllib.parse.quote(msg_pret)}" if client_email else ""
                        st.markdown(f'''
                        <div style="display:flex;gap:4px;">
                            <a href="{wa_url}" target="_blank" style="background:#25D366;color:white;padding:5px 7px;border-radius:6px;text-decoration:none;font-size:14px;" title="WhatsApp">📱</a>
                            <a href="{sms_url}" style="background:#3b82f6;color:white;padding:5px 7px;border-radius:6px;text-decoration:none;font-size:14px;" title="SMS">💬</a>
                            {f'<a href="{email_url}" style="background:#6b7280;color:white;padding:5px 7px;border-radius:6px;text-decoration:none;font-size:14px;" title="Email">✉️</a>' if client_email else ''}
                        </div>
                        ''', unsafe_allow_html=True)
                    else:
                        st.write("—")
                st.markdown("<hr style='margin:5px 0;border:none;border-top:1px solid #eee;'>", unsafe_allow_html=True)
    
    with sub_tab4:
        # Ajouter une nouvelle commande
        st.markdown("##### Ajouter une commande de pièce")
        
        # Sélection du ticket (optionnel)
        tickets_ouverts = [t for t in chercher_tickets() if t.get('statut') not in ['Clôturé', 'Rendu au client']]
        ticket_options = ["-- Sans ticket --"] + [f"{t['ticket_code']} - {t.get('client_nom', '')} ({t.get('marque', '')} {t.get('modele', '')})" for t in tickets_ouverts]
        
        selected_ticket = st.selectbox("Associer à un ticket (optionnel)", ticket_options, key="cmd_ticket")
        
        ticket_id = None
        if selected_ticket != "-- Sans ticket --":
            idx = ticket_options.index(selected_ticket) - 1
            ticket_id = tickets_ouverts[idx]['id']
        
        col1, col2 = st.columns(2)
        with col1:
            cmd_description = st.text_input("Description de la pièce *", placeholder="Ex: Écran iPhone 11", key="cmd_desc")
            cmd_fournisseur = st.selectbox("Fournisseur", FOURNISSEURS, key="cmd_fournisseur")
            if cmd_fournisseur == "Autre":
                cmd_fournisseur = st.text_input("Précisez le fournisseur", key="cmd_fournisseur_autre")
        with col2:
            cmd_reference = st.text_input("Référence (optionnel)", placeholder="Réf. fournisseur", key="cmd_ref")
            cmd_prix = st.number_input("Prix (€)", min_value=0.0, step=1.0, key="cmd_prix")
        
        cmd_notes = st.text_area("Notes", placeholder="Informations complémentaires...", height=80, key="cmd_notes")
        
        if st.button("➕ Ajouter la commande", type="primary", use_container_width=True):
            if cmd_description:
                ajouter_commande_piece(ticket_id, cmd_description, cmd_fournisseur, cmd_reference, cmd_prix, cmd_notes)
                st.success("✅ Commande ajoutée!")
                st.session_state.cmd_new_done = True
                # Réinitialiser le formulaire (UI)
                for _k in ["cmd_desc","cmd_fourni","cmd_ref","cmd_prix","cmd_notes","cmd_ticket"]:
                    if _k in st.session_state: st.session_state[_k] = ""
                st.rerun()
            else:
                st.error("La description est obligatoire")

def staff_attestation():
    """Générer une attestation de non-reparabilite"""
    st.markdown("""<div class="detail-card-header">📄 Attestation de Non-Réparabilité</div>""", unsafe_allow_html=True)
    st.markdown("Générez une attestation pour les appareils économiquement irréparables (utile pour les assurances).")
    
    st.markdown("---")
    
    # Initialiser les valeurs si pas encore fait
    if 'att_nom' not in st.session_state:
        st.session_state.att_nom = ''
    if 'att_prenom' not in st.session_state:
        st.session_state.att_prenom = ''
    if 'att_email' not in st.session_state:
        st.session_state.att_email = ''
    
    # Option pour sélectionner un client existant
    st.markdown("##### 👤 Informations client")
    
    use_existing = st.checkbox("Rechercher un client existant", key="att_use_existing")
    
    if use_existing:
        search_query = st.text_input("🔍 Rechercher par nom, prénom ou téléphone", key="att_search", placeholder="Tapez pour rechercher...")
        
        if search_query and len(search_query) >= 2:
            clients_found = search_clients(search_query)
            if clients_found:
                options = ["-- Sélectionnez --"] + [f"{c['nom']} {c['prenom']} - {c['telephone']}" for c in clients_found]
                selected = st.selectbox("Clients trouvés", options, key="att_client_select")
                
                if selected != "-- Sélectionnez --":
                    # Trouver le client sélectionné
                    idx = options.index(selected) - 1
                    client = clients_found[idx]
                    
                    # Pré-remplir les champs directement dans session_state
                    if st.button("✅ Utiliser ce client", key="att_use_client", type="primary"):
                        st.session_state.att_nom = client['nom']
                        st.session_state.att_prenom = client['prenom']
                        st.session_state.att_email = client.get('email', '')
                        st.success(f"✅ Client sélectionné: {client['nom']} {client['prenom']}")
                        st.rerun()
            else:
                st.info("Aucun client trouvé")
    
    # Champs client (modifiables)
    col1, col2 = st.columns(2)
    with col1:
        att_nom = st.text_input("Nom du client *", key="att_nom")
        att_prenom = st.text_input("Prénom du client *", key="att_prenom")
        att_adresse = st.text_input("Adresse du client", key="att_adresse", placeholder="73000 Chambéry")
    with col2:
        att_email = st.text_input("Email du client", key="att_email", placeholder="client@email.com")
        att_marque = st.selectbox("Marque *", ["Apple", "Samsung", "Xiaomi", "Huawei", "Autre"], key="att_marque")
        att_modele = st.text_input("Modèle *", key="att_modele", placeholder="iPhone 11 Pro")
    
    att_imei = st.text_input("Numéro IMEI / Série *", key="att_imei", placeholder="353833102642466")
    
    st.markdown("---")
    st.markdown("##### 📝 Détails techniques")
    
    att_etat = st.text_area("État de l'appareil au moment du dépôt", key="att_etat", 
                           placeholder="Ex: Chassis arrière endommagé et écran fissuré", height=80)
    att_motif = st.text_area("Motif du dépôt", key="att_motif",
                            placeholder="Ex: iPhone ayant subi un choc violent", height=80)
    att_compte_rendu = st.text_area("Compte rendu du technicien *", key="att_cr",
                                   placeholder="Ex:\n- Afficheur endommagé entraînant la perte d'affichage\n- Chassis endommagé\n- Carte mère trop endommagée", height=120)
    
    st.markdown("---")
    
    if st.button("📄 GÉNÉRER L'ATTESTATION", type="primary", use_container_width=True):
        if not att_nom or not att_prenom or not att_modele or not att_imei or not att_compte_rendu:
            st.error("Veuillez remplir tous les champs obligatoires (*)")
        else:
            from datetime import datetime
            date_jour = datetime.now().strftime("%d/%m/%Y")
            
            # HTML de l'attestation (version pour affichage avec bouton imprimer)
            attestation_html = generate_attestation_html(
                att_nom, att_prenom, att_adresse, att_marque, att_modele, 
                att_imei, att_etat, att_motif, att_compte_rendu, date_jour, 
                include_print_btn=True
            )
            
            st.session_state.attestation_html = attestation_html
            st.session_state.attestation_email = att_email
            st.session_state.attestation_data = {
                'nom': att_nom, 'prenom': att_prenom, 'adresse': att_adresse,
                'marque': att_marque, 'modele': att_modele, 'imei': att_imei,
                'etat': att_etat, 'motif': att_motif, 'compte_rendu': att_compte_rendu,
                'date': date_jour
            }
            st.success("Attestation générée!")
            st.rerun()
    
    # Afficher l'attestation si générée
    if st.session_state.get("attestation_html"):
        st.markdown("---")
        st.markdown("### 📋 Aperçu de l'attestation")
        st.components.v1.html(st.session_state.attestation_html, height=800, scrolling=True)
        
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("🔄 Nouvelle attestation", key="new_attestation", type="secondary", use_container_width=True):
                for key in ['attestation_html', 'attestation_email', 'attestation_data', 
                           'att_nom_val', 'att_prenom_val', 'att_email_val']:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()
        with col2:
            att_email_saved = st.session_state.get("attestation_email", "")
            if att_email_saved and get_param("SMTP_HOST"):
                if st.button("📧 Envoyer par email", key="send_attestation_email", type="primary", use_container_width=True):
                    # Générer le PDF
                    data = st.session_state.get("attestation_data", {})
                    html_for_email = generate_attestation_html(
                        data.get('nom',''), data.get('prenom',''), data.get('adresse',''),
                        data.get('marque',''), data.get('modele',''), data.get('imei',''),
                        data.get('etat',''), data.get('motif',''), data.get('compte_rendu',''),
                        data.get('date',''), include_print_btn=False
                    )
                    
                    sujet = "Attestation de non-réparabilité - Klikphone"
                    message = "Bonjour,\n\nVeuillez trouver ci-jointe votre attestation de non-réparabilité.\n\nCordialement,\nL'équipe Klikphone\n04 79 60 89 22"
                    
                    # Convertir HTML en PDF
                    pdf_bytes = html_to_pdf(html_for_email)
                    if pdf_bytes:
                        success, result = envoyer_email_avec_pdf(att_email_saved, sujet, message, pdf_bytes, "attestation_non_reparabilite.pdf")
                    else:
                        # Fallback: envoyer en HTML
                        success, result = envoyer_email(att_email_saved, sujet, message, html_for_email)
                    
                    if success:
                        st.success(f"✅ Email envoyé à {att_email_saved}!")
                    else:
                        st.error(result)
            elif att_email_saved:
                st.info("💡 Configurez SMTP dans Config > Email")

def generate_attestation_html(nom, prenom, adresse, marque, modele, imei, etat, motif, compte_rendu, date_jour, include_print_btn=True):
    """Génère le HTML de l'attestation"""
    print_btn = '<button class="print-btn" onclick="window.print()">IMPRIMER L\'ATTESTATION</button>' if include_print_btn else ''
    
    return f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: Arial, sans-serif; font-size: 12px; margin: 40px; line-height: 1.6; }}
        .header {{ text-align: center; margin-bottom: 30px; }}
        .logo {{ font-size: 28px; font-weight: bold; color: #f97316; }}
        .contact {{ font-size: 11px; color: #666; margin-top: 5px; }}
        .titre {{ text-align: center; font-weight: bold; font-size: 16px; text-decoration: underline; margin: 30px 0; }}
        .section {{ margin: 20px 0; }}
        .section-title {{ font-weight: bold; text-decoration: underline; margin-bottom: 10px; }}
        .destinataire {{ margin: 20px 0; padding: 10px; background: #f5f5f5; }}
        ul {{ margin: 10px 0; padding-left: 20px; }}
        .footer {{ margin-top: 40px; font-style: italic; }}
        .signature {{ margin-top: 30px; text-align: right; }}
        .print-btn {{ display: block; width: 200px; margin: 30px auto; padding: 12px; background: #f97316; color: white; border: none; border-radius: 8px; font-size: 14px; cursor: pointer; }}
        @media print {{ .print-btn {{ display: none; }} }}
    </style>
</head>
<body>
    <div class="header">
        <div class="logo">KLIKPHONE</div>
        <div class="contact">
            79 Place Saint Léger, 73000 Chambéry<br>
            Tel: 04 79 60 89 22 - contact@klikphone.com<br>
            SIREN: 795334523
        </div>
    </div>
    
    <div class="destinataire">
        <u>Attestation délivrée à :</u><br>
        <strong>M. {nom.upper()} {prenom.upper()}</strong><br>
        {adresse or "73000 Chambéry"}
    </div>
    
    <div class="titre">COMPTE RENDU TECHNICIEN</div>
    
    <div class="section">
        <p>La société <strong>Klikphone</strong>, spécialiste en réparation de smartphones et tablettes, basée au 79 Place Saint Léger à Chambéry;</p>
        
        <p>Atteste que l'appareil <strong>{marque} {modele}</strong> comportant le numéro IMEI/Série suivant: <strong>{imei}</strong> a bien été déposé à notre atelier pour réparation.</p>
    </div>
    
    <div class="section">
        <div class="section-title">État de l'appareil au moment du dépôt :</div>
        <p>{etat or "Non précisé"}</p>
    </div>
    
    <div class="section">
        <div class="section-title">Motif du dépôt :</div>
        <p>{motif or "Diagnostic demandé"}</p>
    </div>
    
    <div class="section">
        <div class="section-title">Compte rendu du technicien :</div>
        <ul>
            {"".join([f"<li>{line.strip()}</li>" for line in compte_rendu.split(chr(10)) if line.strip()])}
        </ul>
    </div>
    
    <div class="section">
        <div class="section-title">Estimation des réparations :</div>
        <p><strong>Après expertise, nous attestons que cet appareil est économiquement irréparable.</strong></p>
    </div>
    
    <div class="section">
        <div class="section-title">Valeur de l'appareil :</div>
        <p>Se référer à la facture d'achat / devis de remplacement</p>
    </div>
    
    <div class="footer">
        <p>Cette attestation fait office de justificatif pour votre assurance.</p>
        <p>Tous nos diagnostics et nos réparations sont garantis et réalisés par un technicien certifié.</p>
    </div>
    
    <div class="signature">
        Fait à Chambéry, le {date_jour}
    </div>
    
    {print_btn}
</body>
</html>
"""

def staff_nouvelle_demande():
    st.markdown("<p class='section-title'>Nouvelle demande manuelle</p>", unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    with col1:
        prenom = st.text_input("Prénom *", key="n_prenom")
        tel = st.text_input("Téléphone *", key="n_tel")
    with col2:
        nom = st.text_input("Nom *", key="n_nom")
        email = st.text_input("Email", key="n_email")
    
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    with col1:
        cat = st.selectbox("Catégorie", CATEGORIES, key="n_cat")
    with col2:
        marque = st.selectbox("Marque", get_marques(cat), key="n_marque")
    with col3:
        modele = st.selectbox("Modèle", get_modeles(cat, marque), key="n_modele")
    
    modele_autre = ""
    if modele == "Autre" or marque == "Autre":
        modele_autre = st.text_input("Précisez le modèle", key="n_modele_autre")
    
    imei = st.text_input("IMEI / Numéro de série (optionnel)", key="n_imei", placeholder="Ex: 353833102642466")
    
    panne = st.selectbox("Problème", PANNES, key="n_panne")
    panne_detail = ""
    if panne in ["Autre", "Diagnostic"]:
        panne_detail = st.text_area("Détails", key="n_panne_detail")
    
    col1, col2 = st.columns(2)
    with col1:
        pin = st.text_input("Code PIN", type="password", key="n_pin")
    with col2:
        pattern = st.text_input("Schéma (ex: 1-2-3)", key="n_pattern")
    
    col1, col2 = st.columns(2)
    with col1:
        devis = st.number_input("Devis estimé (€)", min_value=0.0, step=5.0, key="n_devis")
    with col2:
        acompte = st.number_input("Acompte (€)", min_value=0.0, step=5.0, key="n_acompte")
    
    notes = st.text_area("Notes", key="n_notes")
    
    if st.button("CRÉER LA DEMANDE", type="primary", use_container_width=True):
        if not nom or not prenom or not tel:
            st.error("Nom, prénom et téléphone obligatoires")
        else:
            cid = get_or_create_client(nom, tel, prenom, email)
            code = creer_ticket(cid, cat, marque, modele, modele_autre, panne, panne_detail, pin, pattern, notes, imei, 0)
            t = get_ticket(code=code)
            if t and (devis or acompte):
                update_ticket(t['id'], devis_estime=devis, acompte=acompte)
            st.success(f"Demande créée : {code}")

def staff_config():
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(["🏪 Boutique", "📧 Email", "💬 Messages", "📚 Catalogue", "👥 Équipe", "🔒 Sécurité", "💳 Caisse"])
    
    with tab1:
        st.markdown("### Informations de la boutique")
        nom = st.text_input("Nom boutique", value=get_param("NOM_BOUTIQUE") or "Klikphone")
        adresse = st.text_input("Adresse", value=get_param("ADRESSE_BOUTIQUE") or "79 Place Saint Léger, 73000 Chambéry")
        tel = st.text_input("Téléphone", value=get_param("TEL_BOUTIQUE") or "04 79 60 89 22")
        email_boutique = st.text_input("Email boutique", value=get_param("EMAIL_BOUTIQUE") or "contact@klikphone.com")
        horaires = st.text_input("Horaires", value=get_param("HORAIRES_BOUTIQUE") or "Lundi-Samedi 10h-19h")
        url = st.text_input("URL suivi (QR code)", value=get_param("URL_SUIVI"))
        
        st.markdown("---")
        st.markdown("### Conditions générales (ticket)")
        cgv_ticket = st.text_area("Conditions affichées sur le ticket", 
            value=get_param("CGV_TICKET") or """- Klikphone ne consulte pas et n'accède pas aux données présentes dans votre appareil.
- Une perte de données reste possible — pensez à sauvegarder.
- Klikphone décline toute responsabilité en cas de perte de données ou de panne apparaissant après réparation (oxydation, choc, FaceID, etc.).""",
            height=120)
        
        if st.button("Enregistrer", key="save_config", type="primary"):
            set_param("NOM_BOUTIQUE", nom)
            set_param("ADRESSE_BOUTIQUE", adresse)
            set_param("TEL_BOUTIQUE", tel)
            set_param("EMAIL_BOUTIQUE", email_boutique)
            set_param("HORAIRES_BOUTIQUE", horaires)
            set_param("URL_SUIVI", url)
            set_param("CGV_TICKET", cgv_ticket)
            st.success("Configuration enregistrée!")
    
    with tab2:
        st.markdown("### Configuration Email (SMTP)")
        st.markdown("Pour envoyer des emails directement depuis l'application")
        
        smtp_host = st.text_input("Serveur SMTP", value=get_param("SMTP_HOST"), placeholder="smtp.gmail.com")
        smtp_port = st.text_input("Port", value=get_param("SMTP_PORT") or "587", placeholder="587")
        smtp_user = st.text_input("Email / Utilisateur", value=get_param("SMTP_USER"), placeholder="votre@email.com")
        smtp_pass = st.text_input("Mot de passe", value=get_param("SMTP_PASS"), type="password")
        smtp_from = st.text_input("Email expéditeur (optionnel)", value=get_param("SMTP_FROM"), placeholder="contact@klikphone.com")
        smtp_from_name = st.text_input("Nom expéditeur", value=get_param("SMTP_FROM_NAME") or "Klikphone")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Enregistrer config email", key="save_smtp", type="primary"):
                set_param("SMTP_HOST", smtp_host)
                set_param("SMTP_PORT", smtp_port)
                set_param("SMTP_USER", smtp_user)
                set_param("SMTP_PASS", smtp_pass)
                set_param("SMTP_FROM", smtp_from)
                set_param("SMTP_FROM_NAME", smtp_from_name)
                st.success("Configuration email enregistrée!")
        
        with col2:
            if st.button("Tester l'envoi", key="test_smtp"):
                if smtp_host and smtp_user:
                    success, msg = envoyer_email(smtp_user, "Test Klikphone", "Ceci est un email de test depuis l'application Klikphone SAV.")
                    if success:
                        st.success(msg)
                    else:
                        st.error(msg)
                else:
                    st.warning("Remplissez d'abord la configuration")
        
        st.markdown("---")
        st.markdown("""
        **Aide configuration:**
        - **Gmail**: smtp.gmail.com, port 587, activer "Mots de passe d'application"
        - **OVH**: ssl0.ovh.net, port 587
        - **Outlook**: smtp.office365.com, port 587
        """)
    
    with tab3:
        st.markdown("### Messages prédéfinis personnalisés")
        st.markdown("Personnalisez les messages envoyés aux clients. Variables disponibles: `{prenom}`, `{marque}`, `{modele}`, `{code}`, `{montant}`")
        
        msg_recu = st.text_area("Message: Appareil reçu", 
            value=get_param("MSG_APPAREIL_RECU") or "", 
            placeholder="Laissez vide pour utiliser le message par défaut",
            height=100)
        
        msg_pret = st.text_area("Message: Appareil prêt", 
            value=get_param("MSG_APPAREIL_PRET") or "", 
            placeholder="Laissez vide pour utiliser le message par défaut",
            height=100)
        
        msg_non_reparable = st.text_area("Message: Non réparable", 
            value=get_param("MSG_NON_REPARABLE") or "", 
            placeholder="Laissez vide pour utiliser le message par défaut",
            height=100)
        
        signature = st.text_area("Signature des messages", 
            value=get_param("SIGNATURE_MSG") or "Cordialement,\nL'équipe Klikphone\n04 79 60 89 22",
            height=80)
        
        if st.button("Enregistrer messages", key="save_messages", type="primary"):
            set_param("MSG_APPAREIL_RECU", msg_recu)
            set_param("MSG_APPAREIL_PRET", msg_pret)
            set_param("MSG_NON_REPARABLE", msg_non_reparable)
            set_param("SIGNATURE_MSG", signature)
            st.success("Messages enregistrés!")
    
    with tab4:
        st.markdown("### Ajouter une marque")
        col1, col2 = st.columns(2)
        with col1:
            cat_m = st.selectbox("Catégorie", CATEGORIES, key="cat_marque")
            new_m = st.text_input("Nouvelle marque", key="new_marque")
        with col2:
            if st.button("Ajouter marque", key="add_marque"):
                if new_m and ajouter_marque(cat_m, new_m):
                    st.success("Marque ajoutée!")
                    st.rerun()
        
        st.markdown("---")
        st.markdown("### Ajouter un modèle")
        col1, col2 = st.columns(2)
        with col1:
            cat_mo = st.selectbox("Catégorie", CATEGORIES, key="cat_modele")
            marque_mo = st.selectbox("Marque", get_marques(cat_mo), key="marque_modele")
            new_mo = st.text_input("Nouveau modèle", key="new_modele")
        with col2:
            if st.button("Ajouter modèle", key="add_modele"):
                if new_mo and ajouter_modele(cat_mo, marque_mo, new_mo):
                    st.success("Modèle ajouté!")
                    st.rerun()
        
        st.markdown("---")
        st.markdown("### Types de pannes")
        st.markdown("Pannes actuelles: " + ", ".join(PANNES[:5]) + "...")
        new_panne = st.text_input("Nouvelle panne (avec description)", placeholder="Ex: Caméra floue (photos pas nettes)")
        if st.button("Ajouter panne", key="add_panne"):
            st.info("Pour ajouter des pannes, modifiez la liste PANNES dans le code source.")
    
    with tab5:
        st.markdown("### 👥 Gestion de l'équipe")
        st.markdown("Ajoutez et gérez les membres de l'équipe pour assigner les réparations.")
        
        # Liste des membres actuels
        membres = get_membres_equipe()
        
        st.markdown("#### Membres actuels")
        if membres:
            for m in membres:
                col1, col2, col3, col4 = st.columns([0.3, 1.5, 1.5, 0.8])
                with col1:
                    st.markdown(f"<div style='width:30px;height:30px;background:{m['couleur']};border-radius:50%;margin-top:5px;'></div>", unsafe_allow_html=True)
                with col2:
                    st.write(f"**{m['nom']}**")
                with col3:
                    st.write(m['role'])
                with col4:
                    if st.button("🗑️", key=f"del_membre_{m['id']}", help="Supprimer"):
                        supprimer_membre_equipe(m['id'])
                        st.rerun()
        else:
            st.info("Aucun membre dans l'équipe")
        
        st.markdown("---")
        st.markdown("#### Ajouter un membre")
        
        col1, col2 = st.columns(2)
        with col1:
            new_nom = st.text_input("Nom *", key="new_membre_nom", placeholder="Ex: Marina")
            new_role = st.selectbox("Rôle", ["Technicien Apple", "Technicien Multimarque", "Manager", "Accueil", "Autre"], key="new_membre_role")
        with col2:
            couleurs = {
                "🔴 Rouge": "#EF4444",
                "🟠 Orange": "#F97316",
                "🟡 Jaune": "#EAB308",
                "🟢 Vert": "#22C55E",
                "🔵 Bleu": "#3B82F6",
                "🟣 Violet": "#8B5CF6",
                "💗 Rose": "#EC4899",
                "⚫ Gris": "#6B7280"
            }
            new_couleur_nom = st.selectbox("Couleur", list(couleurs.keys()), key="new_membre_couleur")
            new_couleur = couleurs[new_couleur_nom]
            st.markdown(f"<div style='width:40px;height:40px;background:{new_couleur};border-radius:8px;margin-top:10px;'></div>", unsafe_allow_html=True)
        
        if st.button("➕ Ajouter le membre", type="primary"):
            if new_nom:
                ajouter_membre_equipe(new_nom, new_role, new_couleur)
                st.success(f"✅ {new_nom} ajouté à l'équipe!")
                st.rerun()
            else:
                st.error("Le nom est obligatoire")
    
    with tab6:
        st.markdown("### 🔒 Codes PIN d'accès")
        pin_acc = st.text_input("PIN Accueil", type="password", value=get_param("PIN_ACCUEIL"))
        pin_tech = st.text_input("PIN Technicien", type="password", value=get_param("PIN_TECH"))
        if st.button("Enregistrer PIN", key="save_pin", type="primary"):
            set_param("PIN_ACCUEIL", pin_acc)
            set_param("PIN_TECH", pin_tech)
            st.success("PIN mis à jour!")
        
        st.markdown("---")
        st.markdown("### 💾 Sauvegarde et restauration de la BASE DE DONNÉES")
        st.markdown("""
        <div style="background:#dbeafe;border:1px solid #3b82f6;border-radius:8px;padding:1rem;margin-bottom:1rem;">
            <strong>💡 Important:</strong> Faites une sauvegarde régulière de vos données !<br>
            Le fichier JSON contient TOUTE votre base de données (clients, tickets, paramètres, équipe, catalogue, commandes).
        </div>
        """, unsafe_allow_html=True)
        
        # === EXPORT COMPLET ===
        st.markdown("#### 📤 Exporter la base de données complète")
        
        col_exp1, col_exp2, col_exp3 = st.columns(3)
        
        with col_exp1:
            if st.button("💾 SAUVEGARDE COMPLÈTE (JSON)", type="primary", use_container_width=True):
                import json
                from datetime import datetime as dt
                
                conn = get_db()
                c = conn.cursor()
                
                # Récupérer TOUTES les tables
                c.execute("SELECT * FROM clients ORDER BY id")
                clients_data = [dict(row) for row in c.fetchall()]
                
                c.execute("SELECT * FROM tickets ORDER BY id")
                tickets_data = [dict(row) for row in c.fetchall()]
                
                c.execute("SELECT cle, valeur FROM params")
                params_data = {row['cle']: row['valeur'] for row in c.fetchall()}
                
                c.execute("SELECT * FROM membres_equipe")
                membres_data = [dict(row) for row in c.fetchall()]
                
                c.execute("SELECT * FROM catalog_marques ORDER BY categorie, marque")
                marques_data = [dict(row) for row in c.fetchall()]
                
                c.execute("SELECT * FROM catalog_modeles ORDER BY categorie, marque, modele")
                modeles_data = [dict(row) for row in c.fetchall()]
                
                # Commandes pièces si table existe
                commandes_data = []
                try:
                    c.execute("SELECT * FROM commandes_pieces ORDER BY id")
                    commandes_data = [dict(row) for row in c.fetchall()]
                except:
                    pass
                
                conn.close()
                
                # Créer le fichier JSON complet
                backup_data = {
                    "version": "2.0",
                    "type": "full_backup",
                    "date_export": dt.now().isoformat(),
                    "stats": {
                        "clients": len(clients_data),
                        "tickets": len(tickets_data),
                        "params": len(params_data),
                        "membres_equipe": len(membres_data),
                        "marques": len(marques_data),
                        "modeles": len(modeles_data),
                        "commandes_pieces": len(commandes_data)
                    },
                    "data": {
                        "clients": clients_data,
                        "tickets": tickets_data,
                        "params": params_data,
                        "membres_equipe": membres_data,
                        "catalog_marques": marques_data,
                        "catalog_modeles": modeles_data,
                        "commandes_pieces": commandes_data
                    }
                }
                
                json_str = json.dumps(backup_data, indent=2, ensure_ascii=False, default=str)
                
                st.download_button(
                    label="⬇️ Télécharger la sauvegarde",
                    data=json_str,
                    file_name=f"klikphone_DB_COMPLETE_{dt.now().strftime('%Y%m%d_%H%M')}.json",
                    mime="application/json",
                    key="download_full_backup"
                )
                st.success(f"✅ Sauvegarde complète: {len(clients_data)} clients, {len(tickets_data)} tickets, {len(marques_data)} marques, {len(modeles_data)} modèles")
        
        with col_exp2:
            if st.button("📊 Export Excel", use_container_width=True):
                try:
                    import pandas as pd
                    from io import BytesIO
                    from datetime import datetime as dt
                    
                    conn = get_db()
                    c = conn.cursor()
                    
                    c.execute("SELECT * FROM clients ORDER BY id")
                    clients_df = pd.DataFrame([dict(row) for row in c.fetchall()])
                    
                    c.execute("""SELECT t.*, c.nom as client_nom, c.prenom as client_prenom, 
                                c.telephone as client_tel, c.email as client_email
                                FROM tickets t LEFT JOIN clients c ON t.client_id = c.id ORDER BY t.id""")
                    tickets_df = pd.DataFrame([dict(row) for row in c.fetchall()])
                    
                    conn.close()
                    
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        if not clients_df.empty:
                            clients_df.to_excel(writer, sheet_name='Clients', index=False)
                        if not tickets_df.empty:
                            tickets_df.to_excel(writer, sheet_name='Tickets', index=False)
                    
                    st.download_button(
                        label="⬇️ Télécharger Excel",
                        data=output.getvalue(),
                        file_name=f"klikphone_export_{dt.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_excel"
                    )
                    st.success("✅ Export Excel prêt!")
                except ImportError:
                    st.error("Module pandas ou openpyxl non installé.")
        
        with col_exp3:
            if st.button("📋 Contacts simples", use_container_width=True):
                try:
                    import pandas as pd
                    from io import BytesIO
                    
                    conn = get_db()
                    c = conn.cursor()
                    c.execute("SELECT nom, prenom, telephone, email FROM clients ORDER BY nom, prenom")
                    contacts = [dict(row) for row in c.fetchall()]
                    conn.close()
                    
                    if contacts:
                        df = pd.DataFrame(contacts)
                        output = BytesIO()
                        df.to_excel(output, index=False, sheet_name='Contacts')
                        
                        st.download_button(
                            label="⬇️ Télécharger contacts",
                            data=output.getvalue(),
                            file_name="klikphone_contacts.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_contacts"
                        )
                        st.success(f"✅ {len(contacts)} contacts!")
                except:
                    st.error("Erreur export contacts")
        
        st.markdown("---")
        
        # === IMPORT / RESTAURATION ===
        st.markdown("#### 📥 Restaurer la base de données")
        
        st.markdown("""
        <div style="background:#fef3c7;border:1px solid #f59e0b;border-radius:8px;padding:1rem;margin-bottom:1rem;">
            <strong>⚠️ Attention:</strong><br>
            • <strong>Mode AJOUTER:</strong> Les nouvelles données s'ajoutent aux données existantes (doublons ignorés)<br>
            • <strong>Mode REMPLACER:</strong> Efface TOUTES les données actuelles et les remplace par la sauvegarde (nécessite PIN)
        </div>
        """, unsafe_allow_html=True)
        
        uploaded_file = st.file_uploader("📁 Choisir un fichier de sauvegarde (.json)", type=['json'], key="upload_backup")
        
        if uploaded_file is not None:
            import json
            try:
                backup_data = json.load(uploaded_file)
                
                # Vérifier le format (v1 ou v2)
                is_v2 = backup_data.get('version') == '2.0' and 'data' in backup_data
                
                if is_v2:
                    stats = backup_data.get('stats', {})
                    data = backup_data.get('data', {})
                else:
                    # Format v1 (ancien)
                    stats = {
                        "clients": len(backup_data.get('clients', [])),
                        "tickets": len(backup_data.get('tickets', [])),
                        "params": len(backup_data.get('params', {})),
                        "membres_equipe": len(backup_data.get('membres_equipe', []))
                    }
                    data = backup_data
                
                st.info(f"""
                **📁 Contenu du fichier de sauvegarde:**
                - 📅 Date: {backup_data.get('date_export', 'N/A')}
                - 👥 Clients: **{stats.get('clients', 0)}**
                - 🎫 Tickets: **{stats.get('tickets', 0)}**
                - ⚙️ Paramètres: {stats.get('params', 0)}
                - 👷 Équipe: {stats.get('membres_equipe', 0)}
                - 🏷️ Marques: {stats.get('marques', 0)}
                - 📱 Modèles: {stats.get('modeles', 0)}
                """)
                
                mode_import = st.radio("Mode d'importation:", 
                    ["➕ AJOUTER aux données existantes", "🔄 REMPLACER toutes les données"],
                    key="import_mode")
                
                if "REMPLACER" in mode_import:
                    st.error("⚠️ ATTENTION: Cette action va EFFACER toutes vos données actuelles!")
                    pin_confirm = st.text_input("🔐 Entrez le code PIN pour confirmer:", type="password", key="pin_restore")
                else:
                    pin_confirm = "ok"
                
                col_imp1, col_imp2 = st.columns(2)
                
                with col_imp1:
                    if st.button("✅ RESTAURER", type="primary", use_container_width=True):
                        if "REMPLACER" in mode_import and pin_confirm != "2626":
                            st.error("❌ Code PIN incorrect!")
                        else:
                            conn = get_db()
                            c = conn.cursor()
                            
                            imported_clients = 0
                            imported_tickets = 0
                            
                            # Si mode REMPLACER, vider les tables d'abord
                            if "REMPLACER" in mode_import:
                                c.execute("DELETE FROM tickets")
                                c.execute("DELETE FROM clients")
                                c.execute("DELETE FROM params")
                                c.execute("DELETE FROM membres_equipe")
                                try:
                                    c.execute("DELETE FROM catalog_marques")
                                    c.execute("DELETE FROM catalog_modeles")
                                    c.execute("DELETE FROM commandes_pieces")
                                except:
                                    pass
                                conn.commit()
                            
                            # Importer les clients
                            clients_list = data.get('clients', [])
                            for client in clients_list:
                                try:
                                    c.execute("""INSERT INTO clients (nom, prenom, telephone, email, societe, carte_camby)
                                                VALUES (?, ?, ?, ?, ?, ?)""",
                                             (client.get('nom'), client.get('prenom'), client.get('telephone'),
                                              client.get('email'), client.get('societe'), client.get('carte_camby', 0)))
                                    imported_clients += 1
                                except:
                                    pass
                            
                            # Importer les tickets
                            tickets_list = data.get('tickets', [])
                            for ticket in tickets_list:
                                try:
                                    c.execute("SELECT id FROM tickets WHERE ticket_code=?", (ticket.get('ticket_code'),))
                                    if not c.fetchone():
                                        c.execute("""INSERT INTO tickets (ticket_code, client_id, categorie, marque, modele, 
                                                    modele_autre, imei, panne, panne_detail, pin, pattern, notes_client, 
                                                    notes_internes, devis_estime, acompte, tarif_final, statut, date_depot,
                                                    technicien_assigne, date_recuperation, commande_piece, paye)
                                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                                 (ticket.get('ticket_code'), ticket.get('client_id'), ticket.get('categorie'),
                                                  ticket.get('marque'), ticket.get('modele'), ticket.get('modele_autre'),
                                                  ticket.get('imei'), ticket.get('panne'), ticket.get('panne_detail'),
                                                  ticket.get('pin'), ticket.get('pattern'), ticket.get('notes_client'),
                                                  ticket.get('notes_internes'), ticket.get('devis_estime'), ticket.get('acompte'),
                                                  ticket.get('tarif_final'), ticket.get('statut'), ticket.get('date_depot'),
                                                  ticket.get('technicien_assigne'), ticket.get('date_recuperation'),
                                                  ticket.get('commande_piece', 0), ticket.get('paye', 0)))
                                        imported_tickets += 1
                                except:
                                    pass
                            
                            # Importer les paramètres
                            params_dict = data.get('params', {})
                            for cle, valeur in params_dict.items():
                                try:
                                    c.execute("INSERT OR REPLACE INTO params (cle, valeur) VALUES (?, ?)", (cle, valeur))
                                except:
                                    pass
                            
                            # Importer les membres équipe
                            membres_list = data.get('membres_equipe', [])
                            for membre in membres_list:
                                try:
                                    c.execute("SELECT id FROM membres_equipe WHERE nom=?", (membre.get('nom'),))
                                    if not c.fetchone():
                                        c.execute("INSERT INTO membres_equipe (nom, role, couleur, actif) VALUES (?, ?, ?, ?)",
                                                 (membre.get('nom'), membre.get('role'), membre.get('couleur'), membre.get('actif', 1)))
                                except:
                                    pass
                            
                            # Importer le catalogue (v2 uniquement)
                            if is_v2:
                                for marque in data.get('catalog_marques', []):
                                    try:
                                        c.execute("INSERT OR IGNORE INTO catalog_marques (categorie, marque) VALUES (?, ?)",
                                                 (marque.get('categorie'), marque.get('marque')))
                                    except:
                                        pass
                                
                                for modele in data.get('catalog_modeles', []):
                                    try:
                                        c.execute("INSERT OR IGNORE INTO catalog_modeles (categorie, marque, modele) VALUES (?, ?, ?)",
                                                 (modele.get('categorie'), modele.get('marque'), modele.get('modele')))
                                    except:
                                        pass
                            
                            conn.commit()
                            conn.close()
                            
                            # Vider les caches
                            keys_to_delete = [k for k in st.session_state.keys() if k.startswith('_cache_')]
                            for k in keys_to_delete:
                                del st.session_state[k]
                            
                            st.success(f"✅ Restauration terminée: {imported_clients} clients, {imported_tickets} tickets importés!")
                            st.balloons()
                
                with col_imp2:
                    if st.button("❌ Annuler", use_container_width=True):
                        st.rerun()
                        
            except Exception as e:
                st.error(f"❌ Erreur lors de la lecture du fichier: {str(e)}")
    
    with tab7:
        st.markdown("### 💳 Intégration Caisse Enregistreuse")
        st.markdown("""
        <div style="background:#dbeafe;border:1px solid #3b82f6;border-radius:8px;padding:1rem;margin-bottom:1rem;">
            <strong>🔗 caisse.enregistreuse.fr</strong><br>
            Synchronisez vos tickets de réparation avec votre logiciel de caisse.<br>
            <span style="font-size:12px;color:#64748b;">Nécessite une licence étendue pour l'accès API</span>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("#### Configuration API")
        
        caisse_enabled = st.checkbox("Activer l'intégration Caisse Enregistreuse", 
                                     value=get_param("CAISSE_ENABLED") == "1",
                                     key="caisse_enabled")
        
        # Afficher l'état actuel
        current_apikey = get_param("CAISSE_APIKEY") or ""
        current_shopid = get_param("CAISSE_SHOPID") or ""
        if current_apikey and current_shopid:
            st.success(f"✅ Configuré - SHOPID: {current_shopid} | Token: {current_apikey[:15]}...")
        else:
            st.warning("⚠️ Non configuré - Obtenez le token ci-dessous")
        
        col_c1, col_c2 = st.columns(2)
        with col_c1:
            caisse_login = st.text_input("Email de connexion", 
                                         value=get_param("CAISSE_LOGIN") or "",
                                         placeholder="votre@email.com",
                                         key="caisse_login")
            caisse_password = st.text_input("Mot de passe", 
                                            type="password",
                                            value=get_param("CAISSE_PASSWORD") or "",
                                            key="caisse_password")
        
        with col_c2:
            caisse_apikey = st.text_input("Token API (APIKEY)", 
                                          value=current_apikey,
                                          placeholder="Obtenu automatiquement ou manuellement",
                                          key="caisse_apikey")
            caisse_shopid = st.text_input("ID Boutique (SHOPID)", 
                                          value=current_shopid,
                                          placeholder="Obtenu automatiquement",
                                          key="caisse_shopid")
        
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("🔑 Obtenir le token API", type="secondary", use_container_width=True):
                if caisse_login and caisse_password:
                    try:
                        import requests
                        response = requests.post(
                            "https://caisse.enregistreuse.fr/workers/getAuthToken.php",
                            data={"login": caisse_login, "password": caisse_password},
                            headers={"Content-Type": "application/x-www-form-urlencoded"},
                            timeout=10
                        )
                        data = response.json()
                        if data.get("result") == "OK":
                            new_apikey = data.get("APIKEY", "")
                            new_shopid = data.get("SHOPID", "")
                            set_param("CAISSE_APIKEY", new_apikey)
                            set_param("CAISSE_SHOPID", new_shopid)
                            set_param("CAISSE_ENABLED", "1")
                            st.success(f"✅ Token obtenu et sauvegardé !")
                            st.info(f"SHOPID: {new_shopid}")
                            st.info(f"Token: {new_apikey[:20]}...")
                            st.rerun()
                        else:
                            st.error(f"❌ Erreur: {data.get('errorMessage', data)}")
                    except Exception as e:
                        st.error(f"❌ Erreur de connexion: {str(e)}")
                else:
                    st.warning("Renseignez l'email et le mot de passe")
        
        with col_btn2:
            if st.button("💾 Enregistrer la configuration", type="primary", use_container_width=True):
                set_param("CAISSE_ENABLED", "1" if caisse_enabled else "0")
                set_param("CAISSE_LOGIN", caisse_login)
                set_param("CAISSE_PASSWORD", caisse_password)
                set_param("CAISSE_APIKEY", caisse_apikey)
                set_param("CAISSE_SHOPID", caisse_shopid)
                st.success("✅ Configuration enregistrée!")
                st.rerun()
        
        # Récupérer les modes de paiement de votre caisse
        st.markdown("---")
        st.markdown("#### 💳 Configuration des modes de paiement et caisse")
        
        col_btn1, col_btn2, col_btn3 = st.columns(3)
        
        with col_btn1:
            if st.button("🔄 MODES PAIEMENT", use_container_width=True):
                apikey = get_param("CAISSE_APIKEY")
                shopid = get_param("CAISSE_SHOPID")
                if apikey and shopid:
                    try:
                        import requests
                        response = requests.get(
                            f"https://caisse.enregistreuse.fr/workers/getPaymentModes.php?idboutique={shopid}&key={apikey}&format=json",
                            timeout=10
                        )
                        if response.status_code == 200:
                            data = response.json()
                            st.success(f"✅ {len(data)} modes trouvés !")
                            for m in data:
                                mode_id = m.get('id', '')
                                mode_nom = m.get('nomlong', m.get('nom', str(m)))
                                st.write(f"- **ID {mode_id}** : {mode_nom}")
                    except Exception as e:
                        st.error(f"❌ {e}")
        
        with col_btn2:
            if st.button("🔄 CAISSES", use_container_width=True):
                apikey = get_param("CAISSE_APIKEY")
                shopid = get_param("CAISSE_SHOPID")
                if apikey and shopid:
                    try:
                        import requests
                        response = requests.get(
                            f"https://caisse.enregistreuse.fr/workers/getCashbox.php?idboutique={shopid}&key={apikey}&format=json",
                            timeout=10
                        )
                        if response.status_code == 200:
                            data = response.json()
                            st.success(f"✅ {len(data)} caisses !")
                            for c in data:
                                st.write(f"- **ID {c.get('id', '')}** : {c.get('nom', c)}")
                    except Exception as e:
                        st.error(f"❌ {e}")
        
        with col_btn3:
            if st.button("🔄 UTILISATEURS", use_container_width=True):
                apikey = get_param("CAISSE_APIKEY")
                shopid = get_param("CAISSE_SHOPID")
                if apikey and shopid:
                    try:
                        import requests
                        response = requests.get(
                            f"https://caisse.enregistreuse.fr/workers/getUsers.php?idboutique={shopid}&key={apikey}&format=json",
                            timeout=10
                        )
                        if response.status_code == 200:
                            data = response.json()
                            st.success(f"✅ {len(data)} utilisateurs !")
                            for u in data:
                                user_id = u.get('id', '')
                                user_nom = u.get('login', u.get('nom', str(u)))
                                st.write(f"- **ID {user_id}** : {user_nom}")
                            st.error("⚠️ NE PAS utiliser 'Webservices' !")
                    except Exception as e:
                        st.error(f"❌ {e}")
        
        # Saisie des IDs
        st.markdown("---")
        st.markdown("##### 📝 Configurer les IDs")
        
        # DEBUG - Lire directement de la BDD
        conn = get_db()
        cursor = conn.cursor()
        st.write("**🔍 Valeurs actuelles en BDD:**")
        for param in ["CAISSE_CB_ID", "CAISSE_ESP_ID", "CAISSE_ID", "CAISSE_USER_ID"]:
            r = cursor.execute("SELECT valeur FROM params WHERE cle=?", (param,)).fetchone()
            val = r["valeur"] if r else "(VIDE)"
            st.write(f"• {param} = `{val}`")
        conn.close()
        
        current_cb = get_param("CAISSE_CB_ID") or ""
        current_esp = get_param("CAISSE_ESP_ID") or ""
        current_caisse = get_param("CAISSE_ID") or ""
        current_user = get_param("CAISSE_USER_ID") or ""
        
        col_id1, col_id2 = st.columns(2)
        with col_id1:
            new_cb_id = st.text_input("💳 ID Carte bancaire", value=current_cb, placeholder="528273", key="manual_cb_id")
        with col_id2:
            new_esp_id = st.text_input("💵 ID Espèces", value=current_esp, placeholder="528275", key="manual_esp_id")
        
        col_id3, col_id4 = st.columns(2)
        with col_id3:
            new_caisse_id = st.text_input("🏪 ID Caisse (GENERALE)", value=current_caisse, placeholder="49343", key="manual_caisse_id")
        with col_id4:
            new_user_id = st.text_input("👤 ID Utilisateur", value=current_user, placeholder="42867", key="manual_user_id")
        
        if st.button("💾 SAUVEGARDER TOUS LES IDs", type="primary", use_container_width=True):
            # Sauvegarder directement en BDD
            conn = get_db()
            cursor = conn.cursor()
            
            params_to_save = [
                ("CAISSE_CB_ID", new_cb_id.strip()),
                ("CAISSE_ESP_ID", new_esp_id.strip()),
                ("CAISSE_ID", new_caisse_id.strip()),
                ("CAISSE_USER_ID", new_user_id.strip()),
            ]
            
            for key, val in params_to_save:
                cursor.execute("INSERT OR REPLACE INTO params (cle, valeur) VALUES (?, ?)", (key, val))
                st.write(f"✅ Sauvegardé: {key} = `{val}`")
            
            conn.commit()
            conn.close()
            
            # Invalider tout le cache
            keys_to_clear = [k for k in st.session_state.keys() if k.startswith("_cache_param_")]
            for k in keys_to_clear:
                del st.session_state[k]
            
            st.success("✅ TOUS LES IDs SAUVEGARDÉS !")
            st.rerun()
        
        # Méthode de livraison
        st.markdown("---")
        delivery_options = ["4 (Vente au comptoir)", "0 (À emporter)", "2 (Sur place)"]
        delivery_values = ["4", "0", "2"]
        current_delivery = get_param("CAISSE_DELIVERY_METHOD") or "4"
        delivery_index = delivery_values.index(current_delivery) if current_delivery in delivery_values else 0
        
        caisse_delivery = st.selectbox(
            "Méthode de livraison",
            delivery_options,
            index=delivery_index,
            key="caisse_delivery"
        )
        delivery_val = caisse_delivery.split(" ")[0]
        
        if delivery_val != current_delivery:
            set_param("CAISSE_DELIVERY_METHOD", delivery_val)
        
        # Test de connexion
        st.markdown("---")
        st.markdown("#### Test de connexion")
        
        if st.button("🔄 Tester la connexion API", use_container_width=True):
            apikey = get_param("CAISSE_APIKEY")
            shopid = get_param("CAISSE_SHOPID")
            if apikey and shopid:
                try:
                    import requests
                    response = requests.get(
                        f"https://caisse.enregistreuse.fr/workers/getPlus.php?idboutique={shopid}&key={apikey}&format=json",
                        timeout=10
                    )
                    if response.status_code == 200:
                        data = response.json()
                        if isinstance(data, list):
                            st.success(f"✅ Connexion réussie ! {len(data)} articles trouvés dans votre catalogue")
                        else:
                            st.warning(f"⚠️ Réponse inattendue: {data}")
                    else:
                        st.error(f"❌ Erreur HTTP: {response.status_code}")
                except Exception as e:
                    st.error(f"❌ Erreur: {str(e)}")
            else:
                st.warning("⚠️ Configurez d'abord l'API (token + SHOPID)")
        
        # Instructions
        st.markdown("---")
        st.markdown("""
        <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:1rem;font-size:13px;">
            <strong>📖 Comment ça marche :</strong>
            <ol style="margin:10px 0 0 0;padding-left:20px;">
                <li>Créez un compte sur <a href="https://caisse.enregistreuse.fr/caisse-gratuite/" target="_blank">caisse.enregistreuse.fr</a></li>
                <li>Souscrivez à la licence étendue (28€/mois) pour accéder à l'API</li>
                <li>Entrez vos identifiants ci-dessus et cliquez "Obtenir le token"</li>
                <li>Une fois configuré, un bouton "📤 Envoyer à la caisse" apparaîtra sur les tickets clôturés</li>
            </ol>
        </div>
        """, unsafe_allow_html=True)

# =============================================================================
# INTERFACE TECHNICIEN
# =============================================================================
def ui_tech():
    col1, col2, col3 = st.columns([6, 2, 2])
    with col1:
        st.markdown("<h1 class='page-title'>🔧 Espace Technicien</h1>", unsafe_allow_html=True)
    with col2:
        if st.button("🏠 Accueil", key="goto_accueil", type="secondary", use_container_width=True):
            st.session_state.mode = "accueil"
            st.rerun()
    with col3:
        if st.button("🚪 Sortir", key="logout_tech", use_container_width=True):
            st.session_state.mode = None
            st.session_state.auth = False
            st.rerun()
    
    # Si un ticket est sélectionné, afficher directement le detail
    if st.session_state.get("tech_selected"):
        tech_detail_ticket(st.session_state.tech_selected)
        return
    
    # Filtres améliorés
    col_f1, col_f2, col_f3, col_f4 = st.columns([1.5, 1.5, 1.5, 1.5])
    with col_f1:
        filtre_statut = st.selectbox("Statut", ["Tous"] + STATUTS, key="tech_filtre_statut")
    with col_f2:
        # Filtre par technicien
        membres = get_membres_equipe()
        tech_options = ["👥 Tous", "🔴 Non assignés"] + [m['nom'] for m in membres]
        filtre_tech = st.selectbox("Technicien", tech_options, key="tech_filtre_tech")
    with col_f3:
        tri = st.selectbox("Tri", ["📅 Récent", "📅 Ancien", "🏷️ Statut"], key="tech_tri")
    with col_f4:
        recherche = st.text_input("Recherche", placeholder="Ticket, nom...", key="tech_recherche")
    
    st.markdown("---")
    
    # Récupérer les tickets
    if filtre_statut == "Tous":
        tickets = chercher_tickets()
    else:
        tickets = chercher_tickets(statut=filtre_statut)
    
    # Filtrer par technicien
    if filtre_tech == "🔴 Non assignés":
        tickets = [t for t in tickets if not t.get('technicien_assigne')]
    elif filtre_tech not in ["👥 Tous"]:
        tickets = [t for t in tickets if t.get('technicien_assigne') and filtre_tech in t.get('technicien_assigne', '')]
    
    # Filtrer par recherche
    if recherche:
        recherche_lower = recherche.lower()
        tickets = [t for t in tickets if 
                   recherche_lower in t.get('ticket_code', '').lower() or
                   recherche_lower in t.get('client_nom', '').lower() or
                   recherche_lower in t.get('client_prenom', '').lower() or
                   recherche_lower in t.get('marque', '').lower() or
                   recherche_lower in t.get('modele', '').lower() or
                   recherche_lower in t.get('client_tel', '').lower()]
    
    # Trier
    if "Ancien" in tri:
        tickets = sorted(tickets, key=lambda x: x.get('date_depot', ''))
    elif "Statut" in tri:
        ordre_statut = {s: i for i, s in enumerate(STATUTS)}
        tickets = sorted(tickets, key=lambda x: ordre_statut.get(x.get('statut', ''), 99))
    
    # Pagination
    ITEMS_PER_PAGE = 6
    total_pages = max(1, (len(tickets) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE)
    
    if "tech_page" not in st.session_state:
        st.session_state.tech_page = 1
    
    current_page = st.session_state.tech_page
    start_idx = (current_page - 1) * ITEMS_PER_PAGE
    end_idx = start_idx + ITEMS_PER_PAGE
    tickets_page = tickets[start_idx:end_idx]
    
    st.markdown(f"**{len(tickets)} réparation(s)** • Page {current_page}/{total_pages}")
    
    # En-tete du tableau avec st.columns (aligné avec les lignes)
    col_props_tech = [1, 1.4, 1.6, 1, 1.4, 0.6]
    
    header_cols = st.columns(col_props_tech)
    headers_tech = ["Ticket", "Client", "Appareil", "Tech", "Statut", ""]
    
    st.markdown("""
    <style>
    .tech-header {
        background: linear-gradient(135deg, #1e293b 0%, #334155 100%);
        margin: -12px -16px 12px -16px;
        padding: 14px 16px;
        border-radius: 12px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    for i, col in enumerate(header_cols):
        with col:
            st.markdown(f'<div style="font-size:11px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:0.5px;">{headers_tech[i]}</div>', unsafe_allow_html=True)
    
    st.markdown('<hr style="margin:8px 0;border:none;border-top:1px solid #e5e5e5;">', unsafe_allow_html=True)
    
    # Affichage en liste avec boutons
    for t in tickets_page:
        tid = t['id']
        status_class = get_status_class(t.get('statut', ''))
        
        # Modèle COMPLET VISIBLE
        marque = t.get('marque', '')
        modele_nom = t.get('modele', '')
        if t.get('modele_autre'):
            modele_nom = t['modele_autre']
        
        # Afficher marque + modèle ensemble
        appareil_complet = f"{marque} {modele_nom}".strip()
        if len(appareil_complet) > 24:
            appareil_display = appareil_complet[:22] + "..."
        else:
            appareil_display = appareil_complet
        
        # Catégorie pour icône
        categorie = t.get('categorie', 'Smartphone')
        device_icons = {"Smartphone": "📱", "Tablette": "📟", "PC Portable": "💻", "Console": "🎮", "Commande": "📦"}
        device_icon = device_icons.get(categorie, "📱")
        
        # Technicien avec couleur
        tech = t.get('technicien_assigne', '')
        tech_display = "—"
        tech_color = "#9CA3AF"
        if tech:
            for m in get_membres_equipe():
                if m['nom'] in tech:
                    tech_display = m['nom']
                    tech_color = m['couleur']
                    break
        
        # Indicateur accord client
        accord_icon = ""
        if t.get('statut') == "En attente d'accord client":
            accord_icon = "⚠️ "
        
        # Colonnes alignées avec le header
        row_cols = st.columns(col_props_tech)
        
        with row_cols[0]:
            st.markdown(f'''
            <div style="font-family:monospace;font-size:12px;font-weight:600;color:#171717;">{t['ticket_code']}</div>
            ''', unsafe_allow_html=True)
        
        with row_cols[1]:
            client_nom = f"{t.get('client_nom','')} {t.get('client_prenom','')}".strip()
            if len(client_nom) > 16:
                client_nom = client_nom[:14] + "..."
            st.markdown(f'<div style="font-size:12px;color:#374151;">{client_nom}</div>', unsafe_allow_html=True)
        
        with row_cols[2]:
            # Appareil avec icône et modèle VISIBLE
            st.markdown(f'''
            <div style="display:flex;align-items:center;gap:6px;">
                <span style="font-size:14px;">{device_icon}</span>
                <span style="font-size:12px;font-weight:500;color:#171717;" title="{appareil_complet}">{appareil_display}</span>
            </div>
            ''', unsafe_allow_html=True)
        
        with row_cols[3]:
            if tech_display != "—":
                st.markdown(f'<span style="background:{tech_color};color:white;padding:3px 8px;border-radius:12px;font-size:10px;font-weight:500;">{tech_display}</span>', unsafe_allow_html=True)
            else:
                st.markdown('<span style="color:#a3a3a3;font-size:11px;font-style:italic;">—</span>', unsafe_allow_html=True)
        
        with row_cols[4]:
            statut = t.get('statut', '')
            statut_short = statut[:14] + "..." if len(statut) > 14 else statut
            st.markdown(f'{accord_icon}<span class="badge {status_class}" style="font-size:10px;" title="{statut}">{statut_short}</span>', unsafe_allow_html=True)
        
        with row_cols[5]:
            if st.button("→", key=f"tech_open_{tid}", use_container_width=True):
                st.session_state.tech_selected = tid
                st.rerun()
        
        st.markdown('<hr style="margin:6px 0;border:none;border-top:1px solid #f0f0f0;">', unsafe_allow_html=True)
    
    # Navigation pagination
    if total_pages > 1:
        col_prev, col_pages, col_next = st.columns([1, 3, 1])
        with col_prev:
            if current_page > 1:
                if st.button("◀ Préc", key="tech_prev"):
                    st.session_state.tech_page = current_page - 1
                    st.rerun()
        with col_pages:
            st.markdown(f"<div style='text-align:center;'>Page {current_page} / {total_pages}</div>", unsafe_allow_html=True)
        with col_next:
            if current_page < total_pages:
                if st.button("Suiv ▶", key="tech_next"):
                    st.session_state.tech_page = current_page + 1
                    st.rerun()

def tech_detail_ticket(tid):
    t = get_ticket_full(tid=tid)
    if not t:
        st.error("Ticket non trouvé")
        return
    
    statut_actuel = t.get('statut', STATUTS[0])
    
    # === HEADER PREMIUM ===
    col_back, col_spacer = st.columns([1, 5])
    with col_back:
        if st.button("← Retour", key="tech_close_detail", type="secondary"):
            del st.session_state.tech_selected
            st.rerun()
    
    # Card header avec infos essentielles
    modele_txt = t.get('modele_autre') if t.get('modele_autre') else f"{t.get('marque','')} {t.get('modele','')}"
    panne = t.get('panne_detail') if t.get('panne_detail') else t.get('panne', '')
    status_class = get_status_class(statut_actuel)
    camby_badge = '<span style="background:#8b5cf6;color:white;padding:2px 8px;border-radius:10px;font-size:11px;margin-left:8px;">CAMBY</span>' if t.get('client_carte_camby') else ""
    
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#1e293b 0%,#334155 100%);border-radius:16px;padding:24px;margin-bottom:20px;color:white;">
        <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:16px;">
            <div>
                <div style="font-size:2rem;font-weight:800;letter-spacing:-0.02em;">🎫 {t['ticket_code']}</div>
                <div style="font-size:1.1rem;opacity:0.9;margin-top:4px;">{t.get('client_nom','')} {t.get('client_prenom','')}{camby_badge}</div>
            </div>
            <span class="badge {status_class}" style="font-size:0.95rem;padding:10px 18px;">{statut_actuel}</span>
        </div>
        <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:20px;padding-top:16px;border-top:1px solid rgba(255,255,255,0.15);">
            <div>
                <div style="font-size:0.75rem;text-transform:uppercase;letter-spacing:0.08em;opacity:0.7;">Appareil</div>
                <div style="font-size:1.05rem;font-weight:600;margin-top:4px;">📱 {modele_txt}</div>
            </div>
            <div>
                <div style="font-size:0.75rem;text-transform:uppercase;letter-spacing:0.08em;opacity:0.7;">Réparation</div>
                <div style="font-size:1.05rem;font-weight:600;margin-top:4px;">🔧 {panne}</div>
            </div>
            <div>
                <div style="font-size:0.75rem;text-transform:uppercase;letter-spacing:0.08em;opacity:0.7;">Téléphone</div>
                <div style="font-size:1.05rem;font-weight:600;margin-top:4px;">📞 {t.get('client_tel','N/A')}</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # === WORKFLOW VISUEL ===
    # Déterminer l'étape actuelle
    etapes = {
        "En attente de diagnostic": 1,
        "En attente de pièce": 2,
        "En attente d'accord client": 2,
        "En cours de réparation": 3,
        "Réparation terminée": 4,
        "Rendu au client": 5,
        "Clôturé": 5
    }
    etape_actuelle = etapes.get(statut_actuel, 1)
    
    # Générer le HTML du workflow
    def step_style(num):
        if num < etape_actuelle:
            return "background:linear-gradient(135deg,#22c55e,#16a34a);color:white;"
        elif num == etape_actuelle:
            return "background:linear-gradient(135deg,#f97316,#ea580c);color:white;box-shadow:0 4px 12px rgba(249,115,22,0.4);"
        else:
            return "background:#e2e8f0;color:#94a3b8;"
    
    def label_style(num):
        if num == etape_actuelle:
            return "color:#f97316;font-weight:700;"
        else:
            return "color:#64748b;font-weight:400;"
    
    def bar_style(num):
        return "#22c55e" if num < etape_actuelle else "#e2e8f0"
    
    labels = ["Diagnostic", "Devis", "Réparation", "Terminé", "Rendu"]
    icons = ["🔍", "💰", "🔧", "✅", "🤝"]
    
    st.markdown(f"""
    <div style="background:white;border-radius:16px;padding:24px;margin-bottom:20px;box-shadow:0 4px 20px rgba(0,0,0,0.06);">
        <div style="display:flex;align-items:center;justify-content:space-between;">
            <div style="text-align:center;flex:1;">
                <div style="width:44px;height:44px;margin:0 auto;border-radius:50%;{step_style(1)}display:flex;align-items:center;justify-content:center;font-weight:700;font-size:1.1rem;">{icons[0]}</div>
                <div style="font-size:0.75rem;margin-top:8px;{label_style(1)}">{labels[0]}</div>
            </div>
            <div style="flex:0.8;height:4px;background:{bar_style(1)};border-radius:2px;margin:0 -5px;margin-top:-20px;"></div>
            <div style="text-align:center;flex:1;">
                <div style="width:44px;height:44px;margin:0 auto;border-radius:50%;{step_style(2)}display:flex;align-items:center;justify-content:center;font-weight:700;font-size:1.1rem;">{icons[1]}</div>
                <div style="font-size:0.75rem;margin-top:8px;{label_style(2)}">{labels[1]}</div>
            </div>
            <div style="flex:0.8;height:4px;background:{bar_style(2)};border-radius:2px;margin:0 -5px;margin-top:-20px;"></div>
            <div style="text-align:center;flex:1;">
                <div style="width:44px;height:44px;margin:0 auto;border-radius:50%;{step_style(3)}display:flex;align-items:center;justify-content:center;font-weight:700;font-size:1.1rem;">{icons[2]}</div>
                <div style="font-size:0.75rem;margin-top:8px;{label_style(3)}">{labels[2]}</div>
            </div>
            <div style="flex:0.8;height:4px;background:{bar_style(3)};border-radius:2px;margin:0 -5px;margin-top:-20px;"></div>
            <div style="text-align:center;flex:1;">
                <div style="width:44px;height:44px;margin:0 auto;border-radius:50%;{step_style(4)}display:flex;align-items:center;justify-content:center;font-weight:700;font-size:1.1rem;">{icons[3]}</div>
                <div style="font-size:0.75rem;margin-top:8px;{label_style(4)}">{labels[3]}</div>
            </div>
            <div style="flex:0.8;height:4px;background:{bar_style(4)};border-radius:2px;margin:0 -5px;margin-top:-20px;"></div>
            <div style="text-align:center;flex:1;">
                <div style="width:44px;height:44px;margin:0 auto;border-radius:50%;{step_style(5)}display:flex;align-items:center;justify-content:center;font-weight:700;font-size:1.1rem;">{icons[4]}</div>
                <div style="font-size:0.75rem;margin-top:8px;{label_style(5)}">{labels[4]}</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # === COLONNES PRINCIPALES ===
    col1, col2 = st.columns([1, 1], gap="large")
    
    with col1:
        # --- INFOS & SÉCURITÉ ---
        st.markdown("""
        <div style="background:linear-gradient(135deg,rgba(251,191,36,0.1),rgba(245,158,11,0.05));border:1px solid rgba(251,191,36,0.3);border-radius:14px;padding:16px;margin-bottom:16px;">
            <div style="font-weight:700;color:#92400e;margin-bottom:10px;">🔐 Codes de sécurité</div>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;">
        """, unsafe_allow_html=True)
        
        st.markdown(f"""
                <div style="background:white;padding:10px;border-radius:8px;text-align:center;">
                    <div style="font-size:0.7rem;color:#64748b;text-transform:uppercase;">Code PIN</div>
                    <div style="font-size:1.3rem;font-weight:700;color:#1e293b;font-family:monospace;">{t.get('pin') or '—'}</div>
                </div>
                <div style="background:white;padding:10px;border-radius:8px;text-align:center;">
                    <div style="font-size:0.7rem;color:#64748b;text-transform:uppercase;">Schéma</div>
                    <div style="font-size:1.3rem;font-weight:700;color:#1e293b;font-family:monospace;">{t.get('pattern') or '—'}</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Date de récupération
        if t.get('date_recuperation'):
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,#dcfce7,#bbf7d0);border:1px solid #22c55e;border-radius:12px;padding:14px;margin-bottom:16px;text-align:center;">
                <span style="font-size:1.1rem;">📅</span>
                <span style="font-weight:600;color:#166534;margin-left:8px;">Récupération: {t.get('date_recuperation')}</span>
            </div>
            """, unsafe_allow_html=True)

        # --- APERÇU TICKETS (EN HAUT) ---
        st.markdown("""<div style="margin-top:16px;margin-bottom:8px;font-weight:600;color:#374151;">🎫 Aperçu & Impression</div>""", unsafe_allow_html=True)
        
        col_tk1, col_tk2 = st.columns(2)
        with col_tk1:
            if st.button("🎫 Ticket Client", key=f"tech_show_client_{tid}", use_container_width=True):
                st.session_state[f"tech_show_ticket_{tid}"] = "client"
                st.rerun()
        with col_tk2:
            if st.button("📋 Ticket Staff", key=f"tech_show_staff_{tid}", use_container_width=True):
                st.session_state[f"tech_show_ticket_{tid}"] = "staff"
                st.rerun()

        # Affichage du ticket
        tech_ticket_type = st.session_state.get(f"tech_show_ticket_{tid}")
        if tech_ticket_type:
            st.markdown("---")
            col_hd, col_cl = st.columns([4, 1])
            with col_hd:
                if tech_ticket_type == "client":
                    st.markdown('<div style="background:linear-gradient(135deg,rgba(251,146,60,0.2),rgba(249,115,22,0.1));padding:10px;border-radius:8px;border-left:4px solid #fb923c;"><strong>🎫 TICKET CLIENT</strong></div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div style="background:linear-gradient(135deg,rgba(107,114,128,0.2),rgba(75,85,99,0.1));padding:10px;border-radius:8px;border-left:4px solid #6b7280;"><strong>📋 TICKET STAFF</strong></div>', unsafe_allow_html=True)
            with col_cl:
                if st.button("✕", key=f"tech_close_ticket_{tid}", type="secondary", use_container_width=True):
                    del st.session_state[f"tech_show_ticket_{tid}"]
                    st.rerun()
            
            if tech_ticket_type == "client":
                st.components.v1.html(ticket_client_html(t), height=700, scrolling=True)
            else:
                st.components.v1.html(ticket_staff_html(t), height=750, scrolling=True)

        # --- CENTRE DE NOTIFICATIONS (EN BAS) ---
        st.markdown("""<div style="margin-top:16px;margin-bottom:8px;font-weight:600;color:#374151;">📊 Centre de notifications</div>""", unsafe_allow_html=True)
        
        wa_on = t.get('msg_whatsapp')
        sms_on = t.get('msg_sms')
        email_on = t.get('msg_email')
        date_maj_raw = t.get('date_maj')
        date_maj = str(date_maj_raw)[:16] if date_maj_raw else 'N/A'
        
        # Construire les notes
        notif_parts_tech = []
        if t.get('notes_client'):
            notif_parts_tech.append(f'<div style="background:rgba(249,115,22,0.15);border-left:3px solid #f97316;border-radius:0 8px 8px 0;padding:10px 12px;margin-bottom:8px;"><div style="font-size:10px;color:#fdba74;margin-bottom:3px;font-weight:600;">📋 NOTE CLIENT</div><div style="font-size:12px;color:rgba(255,255,255,0.9);">{t.get("notes_client")}</div></div>')
        if t.get('panne_detail') and t.get('categorie') == 'Commande':
            notif_parts_tech.append(f'<div style="background:rgba(168,85,247,0.15);border-left:3px solid #a855f7;border-radius:0 8px 8px 0;padding:10px 12px;margin-bottom:8px;"><div style="font-size:10px;color:#c4b5fd;margin-bottom:3px;font-weight:600;">📦 COMMANDE</div><div style="font-size:12px;color:rgba(255,255,255,0.9);">{t.get("panne_detail")}</div></div>')
        if t.get('commentaire_client'):
            notif_parts_tech.append(f'<div style="background:rgba(34,197,94,0.15);border-left:3px solid #22c55e;border-radius:0 8px 8px 0;padding:10px 12px;margin-bottom:8px;"><div style="font-size:10px;color:#86efac;margin-bottom:3px;font-weight:600;">💬 NOTE PUBLIQUE</div><div style="font-size:12px;color:rgba(255,255,255,0.9);">{t.get("commentaire_client")}</div></div>')
        if t.get('notes_internes'):
            notif_parts_tech.append(f'<div style="background:rgba(239,68,68,0.15);border-left:3px solid #ef4444;border-radius:0 8px 8px 0;padding:10px 12px;margin-bottom:8px;"><div style="font-size:10px;color:#fca5a5;margin-bottom:3px;font-weight:600;">🔒 NOTE PRIVÉE</div><div style="font-size:12px;color:rgba(255,255,255,0.9);">{t.get("notes_internes")}</div></div>')
        
        notes_content_tech = "".join(notif_parts_tech)
        
        wa_bg = "#22c55e" if wa_on else "#374151"
        wa_color = "white" if wa_on else "#6b7280"
        sms_bg = "#3b82f6" if sms_on else "#374151"
        sms_color = "white" if sms_on else "#6b7280"
        email_bg = "#f59e0b" if email_on else "#374151"
        email_color = "white" if email_on else "#6b7280"
        
        notif_html_tech = f'''<div style="background:linear-gradient(135deg,#1e293b 0%,#334155 100%);border-radius:12px;padding:16px;margin-bottom:16px;color:white;">
<div style="display:flex;justify-content:space-between;align-items:center;background:rgba(255,255,255,0.08);border-radius:8px;padding:10px 12px;margin-bottom:8px;">
<span style="font-size:12px;color:rgba(255,255,255,0.7);">📍 Statut</span>
<span style="background:rgba(249,115,22,0.3);color:#fdba74;padding:3px 10px;border-radius:12px;font-size:11px;font-weight:600;">{statut_actuel}</span>
</div>
<div style="display:flex;justify-content:space-between;align-items:center;background:rgba(255,255,255,0.08);border-radius:8px;padding:10px 12px;margin-bottom:8px;">
<span style="font-size:12px;color:rgba(255,255,255,0.7);">📨 Communications</span>
<div style="display:flex;gap:4px;">
<span style="background:{wa_bg};color:{wa_color};padding:2px 8px;border-radius:10px;font-size:9px;">WA{"✓" if wa_on else ""}</span>
<span style="background:{sms_bg};color:{sms_color};padding:2px 8px;border-radius:10px;font-size:9px;">SMS{"✓" if sms_on else ""}</span>
<span style="background:{email_bg};color:{email_color};padding:2px 8px;border-radius:10px;font-size:9px;">Email{"✓" if email_on else ""}</span>
</div>
</div>
{notes_content_tech}
<div style="text-align:right;font-size:9px;color:rgba(255,255,255,0.4);margin-top:6px;">Màj: {date_maj}</div>
</div>'''
        
        st.markdown(notif_html_tech, unsafe_allow_html=True)

        # --- NOTES (INTERNE + PUBLIC) ---
        st.markdown("""<div style="margin-top:16px;margin-bottom:8px;font-weight:600;color:#374151;">📝 Modifier les notes</div>""", unsafe_allow_html=True)

        notes_internes_edit = st.text_area(
            "Note interne (équipe)",
            value=t.get('notes_internes') or "",
            height=130,
            key=f"tech_notes_int_edit_{tid}",
            help="Visible uniquement par l'équipe (atelier/accueil).",
        )


        # Note privée (équipe uniquement) — mémo interne (non imprimé)
        note_privee_tech_val = st.text_area(
            "Note privée (équipe uniquement)",
            value=get_param("NOTE_PRIVEE_TECH") or "",
            height=80,
            key=f"note_privee_tech_{tid}",
            help="Non imprimée, non visible client. Mémo interne technicien.",
        )
        col_np_t1, col_np_t2 = st.columns([4, 1], gap="small")
        with col_np_t2:
            if st.button("OK", key=f"save_note_privee_tech_{tid}", type="secondary", use_container_width=True):
                set_param("NOTE_PRIVEE_TECH", note_privee_tech_val.strip())
                st.toast("✅ Note privée enregistrée")
                st.rerun()

        commentaire_public_edit = st.text_area(
            "Commentaire public (visible sur le ticket client)",
            value=t.get('commentaire_client') or "",
            height=90,
            key=f"tech_comment_pub_edit_{tid}",
            help="S'imprime sur le ticket/reçu remis au client.",
        )

        col_save_notes, _sp = st.columns([1, 3], gap="small")
        with col_save_notes:
            if st.button("OK", key=f"tech_save_notes_{tid}", type="primary", use_container_width=True):
                update_ticket(tid, notes_internes=notes_internes_edit, commentaire_client=commentaire_public_edit)
                st.toast("✅ Notes mises à jour")
                st.rerun()

    with col2:

        # === ZONE D'ACTION CONTEXTUELLE ===
        
        # Tarification actuelle
        devis = t.get('devis_estime') or 0
        acompte = t.get('acompte') or 0
        prix_supp = t.get('prix_supp') or 0
        total_ttc = devis + prix_supp
        reste = max(0, total_ttc - acompte)
        
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#fff7ed,#ffedd5);border:2px solid #f97316;border-radius:14px;padding:18px;margin-bottom:20px;">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
                <span style="font-weight:700;color:#9a3412;">💰 Tarification</span>
                <span style="font-size:1.8rem;font-weight:800;color:#ea580c;">{total_ttc:.2f} €</span>
            </div>
            <div style="display:flex;justify-content:space-between;font-size:0.9rem;color:#78350f;">
                <span>Acompte versé</span>
                <span style="color:#16a34a;">- {acompte:.2f} €</span>
            </div>
            <div style="height:1px;background:rgba(249,115,22,0.3);margin:10px 0;"></div>
            <div style="display:flex;justify-content:space-between;font-weight:700;">
                <span style="color:#9a3412;">Reste à payer</span>
                <span style="color:#dc2626;font-size:1.2rem;">{reste:.2f} €</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # === ACTIONS SELON LE STATUT ===
        
        # Si diagnostic en cours → Créer devis + envoyer
        if statut_actuel in ["En attente de diagnostic", "En attente de pièce"]:
            st.markdown("""
            <div style="background:linear-gradient(135deg,#dbeafe,#bfdbfe);border-radius:14px;padding:16px;margin-bottom:16px;">
                <div style="font-weight:700;color:#1e40af;margin-bottom:12px;">📋 ÉTAPE 1: Créer et envoyer le devis</div>
            </div>
            """, unsafe_allow_html=True)
            
            col_d1, col_d2 = st.columns(2)
            with col_d1:
                new_devis = st.number_input("💶 Devis TTC", value=float(devis), min_value=0.0, step=5.0, key=f"tech_devis_{tid}")
            with col_d2:
                new_acompte = st.number_input("💳 Acompte", value=float(acompte), min_value=0.0, step=5.0, key=f"tech_acompte_{tid}")
            
            if st.button("💾 Enregistrer le devis", key=f"tech_save_devis_{tid}", type="primary", use_container_width=True):
                update_ticket(tid, devis_estime=new_devis, acompte=new_acompte)
                st.success("✅ Devis enregistré!")
                st.rerun()
            
            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
            
            # Envoyer le devis
            tel = t.get('client_tel', '')
            email = t.get('client_email', '')
            nom_boutique = get_param("NOM_BOUTIQUE") or "Klikphone"
            devis_montant = new_devis if new_devis else devis
            
            msg_devis = f"""Bonjour {t.get('client_prenom', '')},

Suite au diagnostic de votre {modele_txt}, voici notre devis:

🔧 Réparation: {panne}
💰 Montant: {devis_montant:.2f} € TTC

Merci de nous confirmer votre accord pour procéder à la réparation.

{nom_boutique}
📞 {get_param('TEL_BOUTIQUE')}"""
            
            if tel:
                wa_url = wa_link(tel, msg_devis)
                st.markdown(f"""
                <a href="{wa_url}" target="_blank" style="
                    display:flex;align-items:center;justify-content:center;gap:10px;
                    padding:14px;background:linear-gradient(135deg,#25D366,#128C7E);
                    color:white;text-decoration:none;border-radius:12px;
                    font-weight:700;font-size:1rem;margin-bottom:10px;
                    box-shadow:0 4px 15px rgba(37,211,102,0.3);
                ">📱 Envoyer le devis par WhatsApp</a>
                """, unsafe_allow_html=True)
            
            # Bouton pour passer en attente d'accord
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            if st.button("⏳ DEMANDER L'ACCORD CLIENT", key=f"tech_ask_accord_{tid}", type="primary", use_container_width=True):
                changer_statut(tid, "En attente d'accord client")
                ajouter_note(tid, "[TECH] Devis envoyé - En attente d'accord client")
                st.success("✅ Statut mis à jour!")
                st.rerun()
        
        # Si en attente d'accord → Valider l'accord
        elif statut_actuel == "En attente d'accord client":
            st.markdown("""
            <div style="background:linear-gradient(135deg,#fef3c7,#fde68a);border:2px solid #f59e0b;border-radius:14px;padding:20px;margin-bottom:16px;text-align:center;">
                <div style="font-size:2.5rem;margin-bottom:8px;">⏳</div>
                <div style="font-weight:700;color:#92400e;font-size:1.1rem;">En attente de la réponse du client</div>
                <div style="color:#78350f;font-size:0.9rem;margin-top:4px;">Devis: {total_ttc:.2f} € TTC</div>
            </div>
            """.format(total_ttc=total_ttc), unsafe_allow_html=True)
            
            col_acc1, col_acc2 = st.columns(2)
            with col_acc1:
                if st.button("✅ CLIENT A ACCEPTÉ", key=f"tech_accord_ok_{tid}", type="primary", use_container_width=True):
                    update_ticket(tid, client_accord=1)
                    changer_statut(tid, "En cours de réparation")
                    ajouter_note(tid, "[TECH] ✅ Client a accepté le devis - Réparation lancée")
                    st.success("✅ Accord validé! Réparation en cours...")
                    st.rerun()
            with col_acc2:
                if st.button("❌ CLIENT REFUSE", key=f"tech_accord_no_{tid}", type="secondary", use_container_width=True):
                    ajouter_note(tid, "[TECH] ❌ Client a refusé le devis")
                    st.warning("Devis refusé - Que faire?")
            
            # Relancer le client
            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
            tel = t.get('client_tel', '')
            if tel:
                msg_relance = f"Bonjour {t.get('client_prenom', '')}, nous attendons votre confirmation pour le devis de {total_ttc:.2f}€. Merci de nous répondre. {get_param('NOM_BOUTIQUE') or 'Klikphone'}"
                wa_url = wa_link(tel, msg_relance)
                st.markdown(f"""
                <a href="{wa_url}" target="_blank" style="
                    display:block;text-align:center;padding:12px;
                    background:#f1f5f9;color:#475569;text-decoration:none;
                    border-radius:10px;font-weight:600;border:1px solid #e2e8f0;
                ">🔄 Relancer le client par WhatsApp</a>
                """, unsafe_allow_html=True)
        
        # Si en cours de réparation → Terminer
        elif statut_actuel == "En cours de réparation":
            st.markdown("""
            <div style="background:linear-gradient(135deg,#dbeafe,#bfdbfe);border:2px solid #3b82f6;border-radius:14px;padding:20px;margin-bottom:16px;text-align:center;">
                <div style="font-size:2.5rem;margin-bottom:8px;">🔧</div>
                <div style="font-weight:700;color:#1e40af;font-size:1.1rem;">Réparation en cours</div>
            </div>
            """, unsafe_allow_html=True)
            
            # Réparation supplémentaire
            st.markdown("**Réparation supplémentaire (optionnel):**")
            col_rep, col_prix = st.columns([3, 1])
            with col_rep:
                rep_supp = st.text_input("Description", value=t.get('reparation_supp') or "", placeholder="Ex: Nappe Face ID...", key=f"tech_rep_supp_{tid}", label_visibility="collapsed")
            with col_prix:
                new_prix_supp = st.number_input("€", min_value=0.0, step=5.0, value=float(t.get('prix_supp') or 0), key=f"tech_prix_supp_{tid}", label_visibility="collapsed")
            
            if rep_supp and st.button("💾 Enregistrer supplément", key=f"tech_save_supp_{tid}"):
                update_ticket(tid, reparation_supp=rep_supp, prix_supp=new_prix_supp)
                st.success("✅ Supplément enregistré!")
                st.rerun()
            
            st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
            
            if st.button("✅ RÉPARATION TERMINÉE", key=f"tech_finish_{tid}", type="primary", use_container_width=True):
                changer_statut(tid, "Réparation terminée")
                ajouter_note(tid, "[TECH] ✅ Réparation terminée")
                st.success("✅ Réparation terminée!")
                st.rerun()
        
        # Si terminé → Informer le client
        elif statut_actuel == "Réparation terminée":
            st.markdown("""
            <div style="background:linear-gradient(135deg,#dcfce7,#bbf7d0);border:2px solid #22c55e;border-radius:14px;padding:20px;margin-bottom:16px;text-align:center;">
                <div style="font-size:2.5rem;margin-bottom:8px;">✅</div>
                <div style="font-weight:700;color:#166534;font-size:1.1rem;">Réparation terminée!</div>
                <div style="color:#15803d;font-size:0.9rem;margin-top:4px;">En attente de récupération client</div>
            </div>
            """, unsafe_allow_html=True)
            
            # Message pour informer le client
            tel = t.get('client_tel', '')
            if tel:
                msg_pret = f"Bonjour {t.get('client_prenom', '')}, votre {modele_txt} est prêt! Vous pouvez venir le récupérer. Reste à payer: {reste:.2f}€. {get_param('NOM_BOUTIQUE') or 'Klikphone'}"
                wa_url = wa_link(tel, msg_pret)
                st.markdown(f"""
                <a href="{wa_url}" target="_blank" style="
                    display:flex;align-items:center;justify-content:center;gap:10px;
                    padding:14px;background:linear-gradient(135deg,#25D366,#128C7E);
                    color:white;text-decoration:none;border-radius:12px;
                    font-weight:700;font-size:1rem;
                    box-shadow:0 4px 15px rgba(37,211,102,0.3);
                ">📱 Informer le client (WhatsApp)</a>
                """, unsafe_allow_html=True)
        
        # Rendu ou Clôturé
        elif statut_actuel in ["Rendu au client", "Clôturé"]:
            st.markdown("""
            <div style="background:linear-gradient(135deg,#f1f5f9,#e2e8f0);border-radius:14px;padding:20px;text-align:center;">
                <div style="font-size:2.5rem;margin-bottom:8px;">🏁</div>
                <div style="font-weight:700;color:#475569;font-size:1.1rem;">Dossier clôturé</div>
            </div>
            """, unsafe_allow_html=True)
    
    # === SECTION INFÉRIEURE ===
    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
    
    # Changer statut - Toggle avec session_state
    if f"show_statuts_{tid}" not in st.session_state:
        st.session_state[f"show_statuts_{tid}"] = False
    
    col_toggle2, col_spacer2 = st.columns([3, 1])
    with col_toggle2:
        if st.button("⚙️ Changer le statut manuellement" + (" ▼" if st.session_state[f"show_statuts_{tid}"] else " ▶"), 
                     key=f"toggle_statuts_{tid}", use_container_width=True, type="secondary"):
            st.session_state[f"show_statuts_{tid}"] = not st.session_state[f"show_statuts_{tid}"]
            st.rerun()
    
    if st.session_state[f"show_statuts_{tid}"]:
        st.markdown("""<div style="background:white;border:1px solid #e2e8f0;border-radius:0 0 12px 12px;padding:16px;margin-top:-10px;">""", unsafe_allow_html=True)
        st.caption("⚠️ Utilisez les boutons d'action ci-dessus pour un workflow optimal.")
        cols = st.columns(3)
        for i, s in enumerate(STATUTS):
            with cols[i % 3]:
                is_current = (s == statut_actuel)
                if st.button(s, key=f"tech_status_{tid}_{s}", use_container_width=True, disabled=is_current, type="primary" if is_current else "secondary"):
                    changer_statut(tid, s)
                    st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
    
    # Footer
    st.markdown("""
    <div style="text-align:center;padding:20px 0;margin-top:40px;border-top:1px solid #e2e8f0;color:#94a3b8;font-size:12px;">
        Créé par <strong>TkConcept26</strong>
    </div>
    """, unsafe_allow_html=True)

# =============================================================================
# PAGE SUIVI CLIENT
# =============================================================================
def ui_suivi():
    st.markdown(f"""
    <div style="text-align:center; padding:1.5rem 0; background:white; border-radius:16px; margin-bottom:1rem; box-shadow: 0 2px 10px rgba(0,0,0,0.05);">
        <img src="data:image/png;base64,{LOGO_B64}" style="width:60px; height:60px; margin-bottom:0.5rem;">
        <div style="color:#f97316; font-size: 1.8rem; font-weight: 800;">KLIKPHONE</div>
        <p style="color:#6b7280; font-size:0.9rem; margin:0;">Suivi de votre réparation</p>
    </div>
    """, unsafe_allow_html=True)
    
    params = st.query_params
    code_url = params.get("ticket", "")
    
    # Si ticket dans URL, accéder directement au suivi
    if code_url:
        t = get_ticket_full(code=code_url)
        if t:
            # Afficher directement le suivi
            afficher_suivi_ticket(t)
            
            st.markdown("---")
            if st.button("← Retour à l'accueil", use_container_width=True):
                st.query_params.clear()
                st.session_state.mode = None
                st.rerun()
            return
    
    # Formulaire de recherche manuel
    st.markdown("""
    <div style="background:#f8fafc;border-radius:12px;padding:1.5rem;margin-bottom:1rem;">
        <h3 style="color:#1e293b;margin-bottom:0.5rem;">🔍 Rechercher votre réparation</h3>
        <p style="color:#64748b;font-size:0.9rem;">Entrez votre numéro de ticket ou votre téléphone</p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    with col1:
        code = st.text_input("N° de ticket", placeholder="KP-000001")
    with col2:
        tel = st.text_input("OU votre téléphone", placeholder="06 12 34 56 78")
    
    rechercher = st.button("🔍 RECHERCHER", type="primary", use_container_width=True)
    
    if rechercher:
        t = None
        
        # Recherche par code
        if code:
            t = get_ticket_full(code=code)
        # Recherche par téléphone
        elif tel:
            tel_clean = "".join(filter(str.isdigit, tel))
            tickets = chercher_tickets(tel=tel_clean)
            if tickets:
                # Prendre le ticket le plus récent
                t = get_ticket_full(tid=tickets[0]['id'])
        
        if t:
            afficher_suivi_ticket(t)
        else:
            st.error("❌ Aucun ticket trouvé. Vérifiez votre numéro de ticket ou téléphone.")
    
    st.markdown("---")
    if st.button("← Retour à l'accueil"):
        st.session_state.mode = None
        st.rerun()

def afficher_suivi_ticket(t):
    """Affiche le suivi d'un ticket"""
    status_class = get_status_class(t.get('statut', ''))
    modele_txt = f"{t.get('marque','')} {t.get('modele','')}"
    if t.get('modele_autre'): modele_txt += f" ({t['modele_autre']})"
    
    panne = t.get('panne', '')
    if t.get('panne_detail'): panne = t.get('panne_detail')
    
    devis = t.get('devis_estime') or 0
    prix_supp = t.get('prix_supp') or 0
    acompte = t.get('acompte') or 0
    total_ttc = devis + prix_supp
    reste = max(0, total_ttc - acompte)
    
    statut = t.get('statut', '')
    progress_map = {
        "En attente de diagnostic": 20,
        "En attente de pièce": 30,
        "En attente d'accord client": 35,
        "En cours de réparation": 50,
        "Réparation terminée": 80,
        "Rendu au client": 100, 
        "Clôturé": 100
    }
    progress = progress_map.get(statut, 10)
    
    # Carte principale - Utiliser st.container et colonnes Streamlit
    st.markdown(f"""
<div style="background:white;border-radius:16px;padding:1.5rem;box-shadow:0 4px 20px rgba(0,0,0,0.08);margin-bottom:1rem;">
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:1rem;">
<span style="font-size:1.5rem;font-weight:700;color:#1e293b;">🎫 {t['ticket_code']}</span>
<span class="badge {status_class}" style="font-size:0.9rem;padding:8px 16px;">{statut}</span>
</div>
</div>
""", unsafe_allow_html=True)
    
    # Infos en colonnes Streamlit native
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Client**")
        st.write(f"{t.get('client_nom','')} {t.get('client_prenom','')}")
        st.markdown("**Réparation**")
        st.write(panne)
    with col2:
        st.markdown("**Appareil**")
        st.write(modele_txt)
        st.markdown("**Déposé le**")
        st.write(fmt_date(t.get('date_depot','')))
    
    # Date de récupération si définie
    if t.get('date_recuperation'):
        st.success(f"📅 **Récupération prévue :** {t.get('date_recuperation')}")
    
    # Tarification si définie
    if total_ttc > 0:
        st.markdown(f"""
<div style="background:#fff7ed;border:1px solid #f97316;border-radius:12px;padding:1rem;margin:1rem 0;">
<div style="display:flex;justify-content:space-between;align-items:center;">
<span style="color:#9a3412;">💰 Total TTC</span>
<span style="color:#1e293b;font-size:1.2rem;font-weight:700;">{total_ttc:.2f} €</span>
</div>
</div>
""", unsafe_allow_html=True)
        
        if acompte > 0:
            st.info(f"✅ Acompte versé : {acompte:.2f} €")
        
        st.warning(f"💳 **Reste à payer : {reste:.2f} €**")
    
    # Barre de progression
    st.markdown(f"**Progression : {progress}%**")
    st.progress(progress / 100)
    
    # Étapes visuelles
    cols = st.columns(5)
    etapes = ["📥 Déposé", "🔍 Diagnostic", "🔧 Réparation", "✅ Terminé", "🤝 Récupéré"]
    etapes_done = {
        "En attente de diagnostic": 1,
        "En attente de pièce": 2,
        "En attente d'accord client": 2,
        "En cours de réparation": 3,
        "Réparation terminée": 4,
        "Rendu au client": 5,
        "Clôturé": 5
    }
    done_count = etapes_done.get(statut, 0)
    
    for i, (col, etape) in enumerate(zip(cols, etapes)):
        with col:
            if i < done_count:
                st.markdown(f"<div style='text-align:center;color:#16a34a;font-size:0.75rem;font-weight:600;'>{etape}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='text-align:center;color:#d1d5db;font-size:0.75rem;'>{etape}</div>", unsafe_allow_html=True)

# =============================================================================
# ÉCRAN D'ACCUEIL
# =============================================================================
def ui_home():
    """Page d'accueil - tuiles premium (UI only)"""

    st.markdown(f"""
    <div style="text-align:center;padding:2.2rem 1rem 1.2rem;">
        <img src="data:image/png;base64,{LOGO_B64}" style="width:74px;height:74px;margin-bottom:0.9rem;">
        <div style="font-size:2.35rem;font-weight:900;letter-spacing:-1px;color:#0f172a;line-height:1;">KLIKPHONE</div>
        <div style="margin-top:0.35rem;color:#64748b;font-size:0.95rem;">
            SAV Manager • 79 Place Saint Léger, Chambéry • 04 79 60 89 22
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="home-grid">', unsafe_allow_html=True)
    c1, c2 = st.columns(2, gap="large")

    with c1:
        st.markdown('<div class="home-tile">', unsafe_allow_html=True)
        if st.button("👤  Client  ·  Déposer un appareil", key="go_client", use_container_width=True):
            st.session_state.mode = "client"
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div style="height:14px;"></div>', unsafe_allow_html=True)

        st.markdown('<div class="home-tile">', unsafe_allow_html=True)
        if st.button("🧾  Accueil  ·  Gestion SAV", key="go_accueil", use_container_width=True):
            st.session_state.mode = "auth_accueil"
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    with c2:
        st.markdown('<div class="home-tile">', unsafe_allow_html=True)
        if st.button("🔧  Technicien  ·  Atelier", key="go_tech", use_container_width=True):
            st.session_state.mode = "auth_tech"
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div style="height:14px;"></div>', unsafe_allow_html=True)

        st.markdown('<div class="home-tile">', unsafe_allow_html=True)
        if st.button("🔍  Suivi  ·  Voir l'avancement", key="go_suivi", use_container_width=True):
            st.session_state.mode = "suivi"
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)


def ui_auth(mode):
    titre = "Accès Accueil" if mode == "accueil" else "Accès Technicien"
    target = "accueil" if mode == "accueil" else "tech"
    
    st.markdown(f"""
    <div style="text-align:center; padding:1.5rem 0;">
        <div style="color:#f97316; font-size: 1.5rem; font-weight: 700;">{titre}</div>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        # Script pour gérer les cookies
        st.markdown("""
        <script>
        function setCookie(name, value, days) {
            var expires = "";
            if (days) {
                var date = new Date();
                date.setTime(date.getTime() + (days*24*60*60*1000));
                expires = "; expires=" + date.toUTCString();
            }
            document.cookie = name + "=" + (value || "") + expires + "; path=/";
        }
        function getCookie(name) {
            var nameEQ = name + "=";
            var ca = document.cookie.split(';');
            for(var i=0; i < ca.length; i++) {
                var c = ca[i];
                while (c.charAt(0)==' ') c = c.substring(1,c.length);
                if (c.indexOf(nameEQ) == 0) return c.substring(nameEQ.length,c.length);
            }
            return null;
        }
        </script>
        """, unsafe_allow_html=True)
        
        # Vérifier si le PIN est déjà mémorisé dans session_state
        saved_key = f"saved_pin_{target}"
        if saved_key in st.session_state and st.session_state[saved_key]:
            st.session_state.mode = target
            st.session_state.auth = True
            st.rerun()
        
        pin = st.text_input("Code PIN", type="password", placeholder="••••", key="auth_pin_input", label_visibility="collapsed")
        
        # Checkbox pour mémoriser
        remember = st.checkbox("Se souvenir de moi (cette session)", key="remember_pin", value=True)
        
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("← Retour", use_container_width=True):
                st.session_state.mode = None
                st.rerun()
        with col_btn2:
            if st.button("Valider →", type="primary", use_container_width=True):
                if pin == "2626":
                    st.session_state.mode = target
                    st.session_state.auth = True
                    if remember:
                        st.session_state[saved_key] = True
                        # Marquer l'authentification globale
                        st.session_state.global_auth = True
                    st.rerun()
                else:
                    st.error("❌ Code incorrect")

# =============================================================================
# MAIN
# =============================================================================
def main():
    init_db()
    load_css()
    
    if "mode" not in st.session_state: st.session_state.mode = None
    if "auth" not in st.session_state: st.session_state.auth = False
    
    # Si l'URL contient un ticket, aller directement vers le suivi
    params = st.query_params
    if params.get("ticket") and st.session_state.mode is None:
        st.session_state.mode = "suivi"
    
    mode = st.session_state.mode
    
    if mode is None: ui_home()
    elif mode == "client": ui_client()
    elif mode == "auth_accueil": ui_auth("accueil")
    elif mode == "auth_tech": ui_auth("tech")
    elif mode == "accueil":
        if st.session_state.auth: ui_accueil()
        else: st.session_state.mode = "auth_accueil"; st.rerun()
    elif mode == "tech":
        if st.session_state.auth: ui_tech()
        else: st.session_state.mode = "auth_tech"; st.rerun()
    elif mode == "suivi": ui_suivi()

if __name__ == "__main__":
    main()
